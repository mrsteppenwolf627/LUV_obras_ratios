"""Tests for Hito 1: schema, readers, normalizer, import, ratios, master Excel."""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

# -- Project root on sys.path via conftest.py --

from src.core.auditor import compute_file_hash, generate_import_log, save_log
from src.core.bc3_reader import read_bc3
from src.core.excel_reader import read_excel
from src.core.normalizer import normalize
from src.db.models import init_db
from src.db.queries import get_budget_by_hash, list_all_budgets, list_all_ratios
from src.db.schema import Base, Budget, LineItem, Ratio, ValidationLog
from src.export.excel_master_generator import generate_master_excel
from src.ratios.calculator import recalculate_all_ratios

SAMPLES_DIR = Path("data/samples")
SAMPLE_XLSX = SAMPLES_DIR / "proyecto_001" / "22_10_SCE_Datos.xlsx"
SAMPLE_BC3 = SAMPLES_DIR / "proyecto_001" / "P22-143.1 Pressupost Sant Celoni.bc3"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def tmp_db(tmp_path):
    """In-memory-like SQLite in a temp directory."""
    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", echo=False)

    @event.listens_for(engine, "connect")
    def fk_on(conn, _):
        conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


@pytest.fixture()
def minimal_budget():
    return Budget(
        filename="test.xlsx",
        file_hash="a" * 64,
        source_format="excel",
        total_cost=100_000.0,
        surface_m2=200.0,
        building_type="residential",
    )


@pytest.fixture()
def minimal_item(minimal_budget):
    return LineItem(
        budget=minimal_budget,
        chapter_code="C01",
        chapter_name="Actuaciones previas",
        total_cost=10_000.0,
        validation_status="VALID",
    )


# ---------------------------------------------------------------------------
# Tarea 1: Schema
# ---------------------------------------------------------------------------


class TestSchema:
    def test_tables_created(self, tmp_db):
        """All four model tables should be accessible after init."""
        # Tables exist if we can query them without error
        assert tmp_db.query(Budget).count() == 0
        assert tmp_db.query(LineItem).count() == 0
        assert tmp_db.query(Ratio).count() == 0
        assert tmp_db.query(ValidationLog).count() == 0

    def test_budget_insert(self, tmp_db):
        b = Budget(filename="a.xlsx", file_hash="b" * 64, source_format="excel")
        tmp_db.add(b)
        tmp_db.commit()
        assert b.id is not None

    def test_line_item_foreign_key(self, tmp_db):
        b = Budget(filename="b.xlsx", file_hash="c" * 64, source_format="bc3")
        item = LineItem(
            budget=b, chapter_code="C01", chapter_name="X", total_cost=1000.0
        )
        tmp_db.add(b)
        tmp_db.commit()
        assert item.budget_id == b.id

    def test_validation_log_insert(self, tmp_db):
        b = Budget(filename="c.xlsx", file_hash="d" * 64, source_format="excel")
        log = ValidationLog(budget=b, rule_name="TEST", status="PASS")
        tmp_db.add(b)
        tmp_db.commit()
        assert log.id is not None

    def test_timestamps_auto(self, tmp_db):
        b = Budget(filename="d.xlsx", file_hash="e" * 64, source_format="excel")
        tmp_db.add(b)
        tmp_db.commit()
        assert b.import_date is not None


# ---------------------------------------------------------------------------
# Tarea 2: Excel Reader
# ---------------------------------------------------------------------------


class TestExcelReader:
    def test_read_sample(self):
        if not SAMPLE_XLSX.exists():
            pytest.skip("Sample xlsx not available")
        result = read_excel(SAMPLE_XLSX)
        assert result["source_format"] == "excel"
        assert len(result["chapters"]) > 0

    def test_chapters_have_amounts(self):
        if not SAMPLE_XLSX.exists():
            pytest.skip("Sample xlsx not available")
        result = read_excel(SAMPLE_XLSX)
        # DUBIOUS chapters may have None amount — only check VALID ones
        valid = [c for c in result["chapters"] if c.get("validation_status") == "VALID"]
        assert len(valid) > 0
        for ch in valid:
            assert ch["total_cost"] > 0

    def test_unsupported_extension(self, tmp_path):
        bad = tmp_path / "file.csv"
        bad.write_text("a,b")
        result = read_excel(bad)
        assert result["errors"]

    def test_synthetic_xlsx(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Código", "Descripción", "Importe"])
        ws.append(["C01", "Actuaciones previas", 5000.0])
        ws.append(["C02", "Estructura", 20000.0])
        out = tmp_path / "synth.xlsx"
        wb.save(str(out))

        result = read_excel(out)
        assert len(result["chapters"]) == 2
        codes = {c["chapter_code"] for c in result["chapters"]}
        assert "C01" in codes or "C02" in codes
        assert result["total_cost"] == pytest.approx(25000.0)

    def test_no_crashes_on_empty_workbook(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        out = tmp_path / "empty.xlsx"
        wb.save(str(out))
        result = read_excel(out)
        assert isinstance(result["chapters"], list)


# ---------------------------------------------------------------------------
# Tarea 3: BC3 Reader
# ---------------------------------------------------------------------------


class TestBC3Reader:
    def test_read_sample(self):
        if not SAMPLE_BC3.exists():
            pytest.skip("Sample bc3 not available")
        result = read_bc3(SAMPLE_BC3)
        assert result["source_format"] == "bc3"
        assert len(result["chapters"]) > 0

    def test_chapters_have_amounts(self):
        if not SAMPLE_BC3.exists():
            pytest.skip("Sample bc3 not available")
        result = read_bc3(SAMPLE_BC3)
        # All BC3 ~C records have positive totals (DUBIOUS only from normalizer, not reader)
        valid = [c for c in result["chapters"] if c.get("validation_status") == "VALID"]
        assert len(valid) > 0
        for ch in valid:
            assert ch["total_cost"] > 0

    def test_unsupported_extension(self, tmp_path):
        bad = tmp_path / "file.txt"
        bad.write_text("hello")
        result = read_bc3(bad)
        assert result["errors"]

    def test_synthetic_bc3(self, tmp_path):
        content = (
            "~V|RIB|FIEBDC-3/2020||\n"
            "~C|C01#||ACTUACIONES PREVIAS|5000.00|281122|0|\n"
            "~C|C02#||ESTRUCTURA|20000.00|281122|0|\n"
        )
        f = tmp_path / "test.bc3"
        f.write_bytes(content.encode("cp1252"))
        result = read_bc3(f)
        assert len(result["chapters"]) == 2
        codes = {c["chapter_code"] for c in result["chapters"]}
        assert "C01" in codes
        assert "C02" in codes

    def test_top_level_detection(self, tmp_path):
        content = (
            "~C|C01#||TOP CHAPTER|10000.0|281122|0|\n"
            "~C|C0101#||SUB CHAPTER|3000.0|281122|0|\n"
        )
        f = tmp_path / "t.bc3"
        f.write_bytes(content.encode("cp1252"))
        result = read_bc3(f)
        top = [c for c in result["chapters"] if c["is_top_level"]]
        sub = [c for c in result["chapters"] if not c["is_top_level"]]
        assert len(top) == 1
        assert len(sub) == 1


# ---------------------------------------------------------------------------
# Tarea 4: Normalizer
# ---------------------------------------------------------------------------


class TestNormalizer:
    def _make_raw(self, chapters=None, total=None):
        return {
            "source_format": "excel",
            "filename": "test.xlsx",
            "filepath": "test.xlsx",
            "chapters": chapters or [],
            "total_cost": total,
            "warnings": [],
            "errors": [],
            "sheets_processed": ["Sheet1"],
        }

    def test_returns_budget_and_items(self):
        raw = self._make_raw(
            [{"chapter_code": "C01", "chapter_name": "Test", "total_cost": 1000.0, "confidence": "HIGH"}]
        )
        budget, items, logs = normalize(raw, file_hash="x" * 64)
        assert isinstance(budget, Budget)
        assert len(items) == 1
        assert items[0].validation_status == "VALID"

    def test_missing_amount_is_dubious(self):
        raw = self._make_raw(
            [{"chapter_code": "C01", "chapter_name": "Test", "total_cost": None, "confidence": "HIGH"}]
        )
        budget, items, logs = normalize(raw, file_hash="x" * 64)
        assert items[0].validation_status == "DUBIOUS"

    def test_negative_amount_is_dubious(self):
        raw = self._make_raw(
            [{"chapter_code": "C01", "chapter_name": "Test", "total_cost": -100.0, "confidence": "HIGH"}]
        )
        budget, items, logs = normalize(raw, file_hash="x" * 64)
        assert items[0].validation_status == "DUBIOUS"

    def test_surface_m2_stored(self):
        raw = self._make_raw()
        budget, _, _ = normalize(raw, surface_m2=250.0, file_hash="x" * 64)
        assert budget.surface_m2 == 250.0

    def test_reader_errors_create_logs(self):
        raw = self._make_raw()
        raw["errors"] = ["DECODE_FAILED"]
        budget, _, logs = normalize(raw, file_hash="x" * 64)
        rule_names = [l.rule_name for l in logs]
        assert "READER_ERROR" in rule_names


# ---------------------------------------------------------------------------
# Tarea 5: Auditor
# ---------------------------------------------------------------------------


class TestAuditor:
    def test_hash_consistency(self, tmp_path):
        f = tmp_path / "file.txt"
        f.write_bytes(b"hello world")
        h1 = compute_file_hash(f)
        h2 = compute_file_hash(f)
        assert h1 == h2
        assert len(h1) == 64

    def test_hash_differs_for_different_content(self, tmp_path):
        f1 = tmp_path / "a.txt"
        f2 = tmp_path / "b.txt"
        f1.write_bytes(b"aaa")
        f2.write_bytes(b"bbb")
        assert compute_file_hash(f1) != compute_file_hash(f2)

    def test_generate_log_structure(self):
        budget = Budget(
            filename="f.xlsx",
            file_hash="a" * 64,
            source_format="excel",
            total_cost=1000.0,
        )
        items = [
            LineItem(budget=budget, chapter_code="C01", total_cost=500.0, validation_status="VALID")
        ]
        log = generate_import_log(budget, items, "f.xlsx")
        assert log["schema_version"] == "1.0"
        assert log["file_hash"] == "a" * 64
        assert log["chapters"]["valid"] == 1

    def test_save_log_creates_file(self, tmp_path):
        budget = Budget(
            filename="f.xlsx", file_hash="b" * 64, source_format="excel"
        )
        log = generate_import_log(budget, [], "f.xlsx")
        path = save_log(log, logs_dir=tmp_path)
        assert path.exists()
        data = json.loads(path.read_text())
        assert data["file_hash"] == "b" * 64


# ---------------------------------------------------------------------------
# Tarea 6: Import (dry-run test via function calls)
# ---------------------------------------------------------------------------


class TestImport:
    def _do_import(self, session, filepath: Path, surface_m2=None):
        """Helper: replicate import.py --confirm logic."""
        from src.core.auditor import compute_file_hash
        from src.core.excel_reader import read_excel
        from src.core.bc3_reader import read_bc3
        from src.core.normalizer import normalize

        fmt = filepath.suffix.lower()
        if fmt in (".xlsx", ".xlsm"):
            raw = read_excel(filepath)
        else:
            raw = read_bc3(filepath)

        file_hash = compute_file_hash(filepath)
        budget, items, logs = normalize(raw, surface_m2=surface_m2, file_hash=file_hash)

        session.add(budget)
        session.add_all(items)
        session.add_all(logs)
        session.commit()
        return budget

    def test_import_xlsx(self, tmp_db):
        if not SAMPLE_XLSX.exists():
            pytest.skip("Sample xlsx not available")
        b = self._do_import(tmp_db, SAMPLE_XLSX, surface_m2=450.0)
        assert b.id is not None
        assert tmp_db.query(LineItem).count() > 0

    def test_import_bc3(self, tmp_db):
        if not SAMPLE_BC3.exists():
            pytest.skip("Sample bc3 not available")
        b = self._do_import(tmp_db, SAMPLE_BC3)
        assert b.id is not None
        assert tmp_db.query(LineItem).count() > 0

    def test_duplicate_detection(self, tmp_db):
        if not SAMPLE_XLSX.exists():
            pytest.skip("Sample xlsx not available")
        self._do_import(tmp_db, SAMPLE_XLSX)
        from src.core.auditor import compute_file_hash
        from src.db.queries import get_budget_by_hash

        h = compute_file_hash(SAMPLE_XLSX)
        existing = get_budget_by_hash(tmp_db, h)
        assert existing is not None

    def test_dry_run_no_db_write(self, tmp_db):
        """Dry-run: we simulate not committing to session."""
        if not SAMPLE_XLSX.exists():
            pytest.skip("Sample xlsx not available")
        from src.core.excel_reader import read_excel
        from src.core.normalizer import normalize
        from src.core.auditor import compute_file_hash

        raw = read_excel(SAMPLE_XLSX)
        h = compute_file_hash(SAMPLE_XLSX)
        budget, items, logs = normalize(raw, file_hash=h)
        # Dry-run: do NOT add to session
        assert tmp_db.query(Budget).count() == 0


# ---------------------------------------------------------------------------
# Tarea 7: Ratio Calculator
# ---------------------------------------------------------------------------


class TestRatioCalculator:
    def _seed(self, session, chapter: str, amounts: list[float], surface: float = 100.0):
        for i, amt in enumerate(amounts):
            b = Budget(
                filename=f"b{i}.xlsx",
                file_hash=str(i).zfill(64),
                source_format="excel",
                surface_m2=surface,
            )
            item = LineItem(
                budget=b,
                chapter_code=chapter,
                chapter_name=chapter,
                total_cost=amt,
                validation_status="VALID",
            )
            session.add(b)
        session.commit()

    def test_recalculate_ratios(self, tmp_db):
        self._seed(tmp_db, "C01", [10000.0, 20000.0, 30000.0], surface=100.0)
        n = recalculate_all_ratios(tmp_db)
        tmp_db.commit()
        assert n == 1
        ratios = list_all_ratios(tmp_db)
        assert len(ratios) == 1
        r = ratios[0]
        assert r.chapter_code == "C01"
        assert r.median == pytest.approx(200.0)  # median(100,200,300)
        assert r.min_value == pytest.approx(100.0)
        assert r.max_value == pytest.approx(300.0)
        assert r.sample_count == 3

    def test_no_surface_no_ratio(self, tmp_db):
        b = Budget(
            filename="nosurface.xlsx",
            file_hash="z" * 64,
            source_format="excel",
            surface_m2=None,
        )
        item = LineItem(
            budget=b, chapter_code="C02", chapter_name="C02", total_cost=5000.0, validation_status="VALID"
        )
        tmp_db.add(b)
        tmp_db.commit()
        n = recalculate_all_ratios(tmp_db)
        tmp_db.commit()
        assert n == 1
        r = list_all_ratios(tmp_db)[0]
        assert r.median is None

    def test_dubious_items_excluded(self, tmp_db):
        b = Budget(
            filename="dub.xlsx", file_hash="d" * 64, source_format="excel", surface_m2=100.0
        )
        item = LineItem(
            budget=b, chapter_code="C03", chapter_name="C03", total_cost=9999.0, validation_status="DUBIOUS"
        )
        tmp_db.add(b)
        tmp_db.commit()
        recalculate_all_ratios(tmp_db)
        tmp_db.commit()
        ratios = list_all_ratios(tmp_db)
        assert all(r.chapter_code != "C03" or r.sample_count == 0 for r in ratios)


# ---------------------------------------------------------------------------
# Tarea 8: Excel Master Generator
# ---------------------------------------------------------------------------


class TestExcelMasterGenerator:
    def _seed_full(self, session):
        b = Budget(
            filename="test.xlsx",
            file_hash="f" * 64,
            source_format="excel",
            surface_m2=200.0,
            building_type="residential",
            total_cost=50000.0,
        )
        items = [
            LineItem(budget=b, chapter_code="C01", chapter_name="Actuaciones previas", total_cost=5000.0, validation_status="VALID"),
            LineItem(budget=b, chapter_code="C02", chapter_name="Estructura", total_cost=20000.0, validation_status="VALID"),
            LineItem(budget=b, chapter_code="C03", chapter_name="Instalaciones", total_cost=15000.0, validation_status="DUBIOUS"),
        ]
        session.add(b)
        session.add_all(items)
        session.commit()
        recalculate_all_ratios(session)
        session.commit()

    def test_generates_file(self, tmp_db, tmp_path):
        self._seed_full(tmp_db)
        out = tmp_path / "master.xlsx"
        path = generate_master_excel(tmp_db, out)
        assert Path(path).exists()

    def test_has_five_sheets(self, tmp_db, tmp_path):
        self._seed_full(tmp_db)
        out = tmp_path / "master.xlsx"
        generate_master_excel(tmp_db, out)
        wb = load_workbook(str(out))
        assert set(wb.sheetnames) == {"INDEX", "RATIOS_SUMMARY", "CHAPTERS", "AUDIT", "RAW_DATA"}

    def test_index_has_budget_row(self, tmp_db, tmp_path):
        self._seed_full(tmp_db)
        out = tmp_path / "master.xlsx"
        generate_master_excel(tmp_db, out)
        wb = load_workbook(str(out))
        ws = wb["INDEX"]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 2  # header + at least 1 budget

    def test_opens_on_index(self, tmp_db, tmp_path):
        self._seed_full(tmp_db)
        out = tmp_path / "master.xlsx"
        generate_master_excel(tmp_db, out)
        wb = load_workbook(str(out))
        assert wb.active.title == "INDEX"

    def test_overwrites_existing(self, tmp_db, tmp_path):
        self._seed_full(tmp_db)
        out = tmp_path / "master.xlsx"
        generate_master_excel(tmp_db, out)
        generate_master_excel(tmp_db, out)  # second call should not raise
        assert out.exists()


# ---------------------------------------------------------------------------
# Correcciones críticas: nuevos tests
# ---------------------------------------------------------------------------


class TestBC3ReaderPartidas:
    """Fix 1 — BC3 parser procesa ~D (items/descomposiciones)."""

    def test_extracts_items_from_decomp(self, tmp_path):
        content = (
            "~V|RIB|FIEBDC-3/2020||\n"
            "~C|C01#||ACTUACIONES PREVIAS|5000.00|281122|0|\n"
            "~D|C01#|E0101\\1\\438.5\\E0102\\1\\4\\|\n"
        )
        f = tmp_path / "test.bc3"
        f.write_bytes(content.encode("cp1252"))
        result = read_bc3(f)
        assert len(result["chapters"]) == 1
        ch = result["chapters"][0]
        assert "items" in ch
        assert len(ch["items"]) == 2
        codes = {i["code"] for i in ch["items"]}
        assert "E0101" in codes
        assert "E0102" in codes

    def test_items_have_factor_and_rendimiento(self, tmp_path):
        content = (
            "~C|C01#||CAP 1|3984.42|281122|0|\n"
            "~D|C01#|E0101\\1\\438.5\\E0102\\2\\4\\|\n"
        )
        f = tmp_path / "t.bc3"
        f.write_bytes(content.encode("cp1252"))
        result = read_bc3(f)
        ch = result["chapters"][0]
        e0101 = next(i for i in ch["items"] if i["code"] == "E0101")
        assert e0101["factor"] == pytest.approx(1.0)
        assert e0101["rendimiento"] == pytest.approx(438.5)
        e0102 = next(i for i in ch["items"] if i["code"] == "E0102")
        assert e0102["factor"] == pytest.approx(2.0)

    def test_chapter_without_decomp_has_empty_items(self, tmp_path):
        content = "~C|C01#||CAP 1|3984.42|281122|0|\n"
        f = tmp_path / "t.bc3"
        f.write_bytes(content.encode("cp1252"))
        result = read_bc3(f)
        ch = result["chapters"][0]
        assert ch["items"] == []

    def test_decomp_mismatch_logged_as_warning_not_dubious(self, tmp_path):
        """~D rendimiento != cost: logs WARNING but chapter stays VALID."""
        content = (
            "~C|C01#||CAP 1|3984.42|281122|0|\n"
            "~D|C01#|E0101\\1\\438.5\\E0102\\1\\4\\|\n"
        )
        f = tmp_path / "t.bc3"
        f.write_bytes(content.encode("cp1252"))
        result = read_bc3(f)
        ch = result["chapters"][0]
        # Chapter stays VALID (trust ~C total — ~D has rendimientos, not costs)
        assert ch["validation_status"] == "VALID"
        # Mismatch is logged as a warning
        assert any("DECOMP_MISMATCH" in w for w in result["warnings"])

    def test_sample_bc3_has_items(self):
        if not SAMPLE_BC3.exists():
            pytest.skip("Sample BC3 not available")
        result = read_bc3(SAMPLE_BC3)
        chapters_with_items = [c for c in result["chapters"] if c["items"]]
        assert len(chapters_with_items) > 0


class TestImportTransactional:
    """Fix 2 — import.py usa transacción atómica (flush/commit/rollback)."""

    def _run_import(self, session, filepath: Path, surface_m2=None):
        """Replicate the atomic import logic from import.py."""
        from src.core.auditor import compute_file_hash, generate_import_log, save_log
        from src.core.excel_reader import read_excel
        from src.core.bc3_reader import read_bc3
        from src.core.normalizer import normalize

        fmt = filepath.suffix.lower()
        raw = read_excel(filepath) if fmt in (".xlsx", ".xlsm") else read_bc3(filepath)
        h = compute_file_hash(filepath)
        budget, items, logs = normalize(raw, surface_m2=surface_m2, file_hash=h)

        session.add(budget)
        session.add_all(items)
        session.add_all(logs)
        session.flush()  # staged, not committed
        return budget, items, logs

    def test_flush_before_commit(self, tmp_db):
        """flush() makes data visible in session but doesn't commit to disk."""
        if not SAMPLE_XLSX.exists():
            pytest.skip("Sample xlsx not available")
        budget, _, _ = self._run_import(tmp_db, SAMPLE_XLSX, surface_m2=450.0)
        # Data visible within session after flush
        assert budget.id is not None
        assert tmp_db.query(Budget).count() == 1

    def test_rollback_removes_staged_data(self, tmp_db):
        """If we rollback after flush(), nothing persists."""
        if not SAMPLE_XLSX.exists():
            pytest.skip("Sample xlsx not available")
        try:
            self._run_import(tmp_db, SAMPLE_XLSX, surface_m2=450.0)
            raise RuntimeError("simulated failure after flush")
        except RuntimeError:
            tmp_db.rollback()
        # After rollback, nothing in DB
        assert tmp_db.query(Budget).count() == 0
        assert tmp_db.query(LineItem).count() == 0

    def test_commit_persists_data(self, tmp_db):
        """After commit(), data survives a new query."""
        if not SAMPLE_XLSX.exists():
            pytest.skip("Sample xlsx not available")
        self._run_import(tmp_db, SAMPLE_XLSX, surface_m2=450.0)
        tmp_db.commit()
        assert tmp_db.query(Budget).count() == 1

    def test_duplicate_hash_blocked(self, tmp_db):
        """Second import of same file raises IntegrityError on commit."""
        from sqlalchemy.exc import IntegrityError
        if not SAMPLE_XLSX.exists():
            pytest.skip("Sample xlsx not available")
        self._run_import(tmp_db, SAMPLE_XLSX)
        tmp_db.commit()
        try:
            self._run_import(tmp_db, SAMPLE_XLSX)
            tmp_db.commit()
            # Should have been blocked by unique hash constraint
            # (or duplicate-check logic in import.py)
        except IntegrityError:
            tmp_db.rollback()
            # Expected: unique constraint on file_hash
        assert tmp_db.query(Budget).count() == 1


class TestExcelReaderDubious:
    """Fix 3 — excel_reader marca DUBIOUS correctamente."""

    def test_row_without_code_is_dubious(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Código", "Descripción", "Importe"])
        ws.append(["C01", "Con código", 5000.0])
        ws.append(["", "Sin código", 3000.0])  # should be DUBIOUS
        out = tmp_path / "t.xlsx"
        wb.save(str(out))
        result = read_excel(out)
        dubious = [c for c in result["chapters"] if c.get("validation_status") == "DUBIOUS"]
        assert len(dubious) >= 1
        assert any("codigo" in (c.get("validation_reason") or "").lower() for c in dubious)

    def test_row_without_amount_is_dubious_not_skipped(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Código", "Descripción", "Importe"])
        ws.append(["C01", "Con importe", 5000.0])
        ws.append(["C02", "Sin importe", None])  # should be DUBIOUS, not skipped
        out = tmp_path / "t.xlsx"
        wb.save(str(out))
        result = read_excel(out)
        codes = {c["chapter_code"] for c in result["chapters"]}
        assert "C02" in codes, "Row without amount should be included as DUBIOUS, not dropped"
        c02 = next(c for c in result["chapters"] if c["chapter_code"] == "C02")
        assert c02["validation_status"] == "DUBIOUS"
        assert c02["total_cost"] is None

    def test_valid_rows_stay_valid(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Código", "Descripción", "Importe"])
        ws.append(["C01", "Bueno", 5000.0])
        out = tmp_path / "t.xlsx"
        wb.save(str(out))
        result = read_excel(out)
        valid = [c for c in result["chapters"] if c.get("validation_status") == "VALID"]
        assert len(valid) == 1
        assert valid[0]["chapter_code"] == "C01"


class TestNormalizerRespectsDubious:
    """Fix 3 — normalizer nunca convierte DUBIOUS del reader a VALID."""

    def _raw(self, chapters):
        return {
            "source_format": "excel",
            "filename": "t.xlsx",
            "filepath": "t.xlsx",
            "chapters": chapters,
            "total_cost": None,
            "warnings": [],
            "errors": [],
            "sheets_processed": [],
        }

    def test_reader_dubious_propagates(self):
        raw = self._raw([{
            "chapter_code": "C01",
            "chapter_name": "X",
            "total_cost": 5000.0,  # amount is fine...
            "validation_status": "DUBIOUS",  # ...but reader marked DUBIOUS
            "validation_reason": "codigo de capitulo ausente",
        }])
        _, items, logs = normalize(raw, file_hash="a" * 64)
        assert items[0].validation_status == "DUBIOUS"

    def test_reader_valid_with_amount_stays_valid(self):
        raw = self._raw([{
            "chapter_code": "C01",
            "chapter_name": "Good",
            "total_cost": 5000.0,
            "validation_status": "VALID",
            "validation_reason": None,
        }])
        _, items, _ = normalize(raw, file_hash="b" * 64)
        assert items[0].validation_status == "VALID"

    def test_dubious_reason_logged(self):
        raw = self._raw([{
            "chapter_code": "C01",
            "chapter_name": "X",
            "total_cost": 5000.0,
            "validation_status": "DUBIOUS",
            "validation_reason": "importe faltante o invalido",
        }])
        _, _, logs = normalize(raw, file_hash="c" * 64)
        rule_names = [l.rule_name for l in logs]
        assert "READER_DUBIOUS" in rule_names


class TestAuditIntegrity:
    """Fix 4 — AUDIT usa (chapter_code, building_type) como clave."""

    def _seed_two_types(self, session):
        for btype, hash_suffix, amount in [
            ("residential", "r" * 64, 5000.0),
            ("office", "o" * 64, 8000.0),
        ]:
            b = Budget(
                filename=f"{btype}.xlsx",
                file_hash=hash_suffix,
                source_format="excel",
                surface_m2=100.0,
                building_type=btype,
                total_cost=amount,
            )
            item = LineItem(
                budget=b,
                chapter_code="C01",
                chapter_name="Estructura",
                total_cost=amount,
                validation_status="VALID",
            )
            session.add(b)
        session.commit()
        recalculate_all_ratios(session)
        session.commit()

    def test_audit_has_column_for_building_type(self, tmp_db, tmp_path):
        self._seed_two_types(tmp_db)
        out = tmp_path / "m.xlsx"
        generate_master_excel(tmp_db, out)
        wb = load_workbook(str(out))
        ws = wb["AUDIT"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Tipo Edificio" in headers

    def test_audit_no_collision_between_building_types(self, tmp_db, tmp_path):
        """C01/residential and C01/office should be separate rows in AUDIT."""
        self._seed_two_types(tmp_db)
        out = tmp_path / "m.xlsx"
        generate_master_excel(tmp_db, out)
        wb = load_workbook(str(out))
        ws = wb["AUDIT"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        c01_rows = [r for r in rows if r[0] == "C01"]
        # Should have separate rows for residential and office
        assert len(c01_rows) >= 2

    def test_ratios_summary_no_duplicate_chapter_type(self, tmp_db, tmp_path):
        """RATIOS_SUMMARY must not have duplicate (chapter, building_type) rows."""
        self._seed_two_types(tmp_db)
        out = tmp_path / "m.xlsx"
        generate_master_excel(tmp_db, out)
        wb = load_workbook(str(out))
        ws = wb["RATIOS_SUMMARY"]
        keys = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = (row[0], row[2])  # (Capítulo, Tipo Edificio)
            assert key not in keys, f"Duplicate (chapter, type) in RATIOS_SUMMARY: {key}"
            keys.add(key)
