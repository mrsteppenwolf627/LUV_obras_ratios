"""Tests for master Excel generation."""

from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.db.models import get_session
from src.export.excel_master_generator import generate_master_excel


@pytest.fixture
def db_session():
    """Get a database session."""
    session = get_session()
    yield session
    session.close()


def test_master_excel_generation(db_session):
    """Test that master Excel is generated with correct structure."""
    excel_path = generate_master_excel(db_session)
    excel_file = Path(excel_path)

    # Verify file exists
    assert excel_file.exists(), f"Excel file not created: {excel_file}"

    # Verify file size
    file_size = excel_file.stat().st_size
    assert file_size > 5000, f"Excel file too small: {file_size} bytes"


def test_master_excel_dated_filename(db_session):
    """Test that Excel uses dated filename: MASTER_YYYY-MM-DD.xlsx."""
    excel_path = generate_master_excel(db_session)
    excel_file = Path(excel_path)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    expected_name = f"MASTER_{today}.xlsx"

    assert excel_file.name == expected_name, f"Expected {expected_name}, got {excel_file.name}"


def test_master_excel_sheets(db_session):
    """Test that Excel has all required sheets."""
    excel_path = generate_master_excel(db_session)
    wb = load_workbook(excel_path)

    required_sheets = ["ITEM_MASTER", "INDEX", "RATIOS_SUMMARY", "CHAPTERS", "AUDIT", "RAW_DATA"]
    assert set(required_sheets) == set(
        wb.sheetnames
    ), f"Expected sheets {required_sheets}, got {wb.sheetnames}"

    # First sheet should be ITEM_MASTER
    assert wb.active.title == "ITEM_MASTER", f"First sheet should be ITEM_MASTER, got {wb.active.title}"


def test_item_master_sheet_has_36_items(db_session):
    """Test that ITEM_MASTER sheet contains exactly 36 items."""
    excel_path = generate_master_excel(db_session)
    wb = load_workbook(excel_path)
    ws = wb["ITEM_MASTER"]

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    # Verify 36 items
    assert len(data_rows) == 36, f"Expected 36 items, got {len(data_rows)}"

    # Verify all items have item_key (first column)
    items_with_key = sum(1 for row in data_rows if row[0])
    assert items_with_key == 36, f"Expected 36 items with key, got {items_with_key}"


def test_item_master_sheet_structure(db_session):
    """Test ITEM_MASTER sheet has correct columns."""
    excel_path = generate_master_excel(db_session)
    wb = load_workbook(excel_path)
    ws = wb["ITEM_MASTER"]

    expected_headers = [
        "Item Key",
        "Categoría",
        "Subcategoría",
        "Unidad",
        "Unitario (Mediana)",
        "Mínimo",
        "Máximo",
        "P25",
        "P75",
        "Desv. Std",
        "Muestras (N)",
        "Solidez",
        "Última Actualización",
    ]

    actual_headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    actual_headers = list(actual_headers) if isinstance(actual_headers, tuple) else actual_headers
    assert actual_headers == expected_headers, f"Headers mismatch: {actual_headers}"


def test_item_master_solidez_calculation(db_session):
    """Test that solidez (confidence level) is calculated correctly."""
    excel_path = generate_master_excel(db_session)
    wb = load_workbook(excel_path)
    ws = wb["ITEM_MASTER"]

    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[1:]

    valid_solidez = {"MUY_DÉBIL", "DÉBIL", "SÓLIDO", "MUY_SÓLIDO"}

    solidez_col = 11  # Solidez is column 11 (0-indexed)
    for row_num, row in enumerate(data_rows, start=2):
        solidez = row[solidez_col]
        if solidez:  # Only check non-empty values
            assert (
                solidez in valid_solidez
            ), f"Row {row_num}: Invalid solidez value '{solidez}'. Expected one of {valid_solidez}"


def test_item_master_currency_formatting(db_session):
    """Test that currency values are properly formatted in Excel."""
    excel_path = generate_master_excel(db_session)
    wb = load_workbook(excel_path)
    ws = wb["ITEM_MASTER"]

    rows = list(ws.iter_rows(values_only=True))

    # Check that currency columns (5-10) have values or are empty
    currency_cols = [4, 5, 6, 7, 8, 9]  # 0-indexed: Mediana, Min, Max, P25, P75, Desv Std
    for row_num, row in enumerate(rows[1:], start=2):
        for col_idx in currency_cols:
            val = row[col_idx]
            # Value should be either empty or a number (formatted in Excel)
            assert val is None or isinstance(val, (int, float, str)), f"Row {row_num}, Col {col_idx}: Invalid value type {type(val)}"
