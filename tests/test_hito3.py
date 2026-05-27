"""Tests for Hito 3: Presto parser + space calculator + Excel generator."""

from __future__ import annotations

import math
import struct
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Fixtures — paths to real sample files
# ---------------------------------------------------------------------------

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
_PRESTO_SCE = _PROJECT_ROOT / "data/samples/proyecto_001/22_10_SCE_Presupuesto final.Presto"
_PRESTO_MED = _PROJECT_ROOT / "data/samples/proyecto_002/MED_PSJ25_V1.Presto"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_minimal_presto(spaces: list[tuple[str, float]], tmp_path: Path) -> Path:
    """Create a minimal Presto binary with the given spaces and costs."""
    buf = bytearray()

    # Budget code header
    buf += b"\x04\x01TEST_CODE\x00"

    for name, cost in spaces:
        name_b = name.encode("latin-1")
        # Space marker
        buf += b"_SPC_\x00\x02\x01" + name_b + b"\x00"
        # Cost pattern
        buf += b"\x06\x01" + name_b + b"\x00\x07\x00"
        buf += struct.pack("<d", cost)

    fpath = tmp_path / "test.Presto"
    fpath.write_bytes(bytes(buf))
    return fpath


def _presupuesto_no_m2(spaces: list[tuple[str, float]]) -> dict:
    """Build a minimal presupuesto dict (no m²)."""
    from src.core.presto_reader import _resolve_zone
    return {
        "filename": "test.Presto",
        "budget_code": "TST",
        "total_coste": sum(c for _, c in spaces),
        "total_m2": 0.0,
        "espacios": [
            {
                "nombre": name,
                "zona": _resolve_zone(name),
                "planta": "TOTAL",
                "coste": cost,
                "m2": 0.0,
                "partidas": [],
            }
            for name, cost in spaces
        ],
    }


# ===========================================================================
# 1. PRESTO READER — read_presto()
# ===========================================================================

class TestReadPresto:
    def test_returns_dict(self, tmp_path):
        from src.core.presto_reader import read_presto
        f = _make_minimal_presto([("SALA", 10000.0)], tmp_path)
        result = read_presto(f)
        assert isinstance(result, dict)

    def test_has_required_keys(self, tmp_path):
        from src.core.presto_reader import read_presto
        f = _make_minimal_presto([("SALA", 5000.0)], tmp_path)
        result = read_presto(f)
        for key in ("filename", "source_format", "budget_code", "spaces", "total_coste",
                    "has_space_breakdown", "errors", "warnings"):
            assert key in result, f"Missing key: {key}"

    def test_source_format_is_presto(self, tmp_path):
        from src.core.presto_reader import read_presto
        f = _make_minimal_presto([("SALA", 1000.0)], tmp_path)
        assert read_presto(f)["source_format"] == "presto"

    def test_extracts_space_names(self, tmp_path):
        from src.core.presto_reader import read_presto
        spaces = [("SALA", 1000.0), ("COMEDOR", 2000.0)]
        f = _make_minimal_presto(spaces, tmp_path)
        result = read_presto(f)
        names = [s["nombre"] for s in result["spaces"]]
        assert "SALA" in names
        assert "COMEDOR" in names

    def test_extracts_costs(self, tmp_path):
        from src.core.presto_reader import read_presto
        f = _make_minimal_presto([("COCINA", 12345.67)], tmp_path)
        result = read_presto(f)
        assert any(abs(s["coste"] - 12345.67) < 0.01 for s in result["spaces"])

    def test_total_coste_equals_sum(self, tmp_path):
        from src.core.presto_reader import read_presto
        spaces = [("SALA", 5000.0), ("COCINA", 3000.0)]
        f = _make_minimal_presto(spaces, tmp_path)
        result = read_presto(f)
        expected = sum(c for _, c in spaces)
        assert abs(result["total_coste"] - expected) < 0.01

    def test_missing_file_returns_error(self):
        from src.core.presto_reader import read_presto
        result = read_presto("/nonexistent/file.Presto")
        assert result["errors"]

    def test_real_presto_sce(self):
        """Smoke test against the real SCE file."""
        if not _PRESTO_SCE.exists():
            pytest.skip("Sample file not found")
        from src.core.presto_reader import read_presto
        result = read_presto(_PRESTO_SCE)
        assert result["has_space_breakdown"]
        assert len(result["spaces"]) > 1
        assert result["total_coste"] > 0

    def test_real_presto_med_fallback(self):
        """MED file has no breakdown — should use fallback."""
        if not _PRESTO_MED.exists():
            pytest.skip("Sample file not found")
        from src.core.presto_reader import read_presto
        result = read_presto(_PRESTO_MED)
        assert not result["has_space_breakdown"]
        assert result["warnings"]

    def test_zone_assigned(self, tmp_path):
        from src.core.presto_reader import read_presto
        f = _make_minimal_presto([("SALA", 1000.0)], tmp_path)
        result = read_presto(f)
        sala = next(s for s in result["spaces"] if s["nombre"] == "SALA")
        assert sala["zona"] == "NOBLE"

    def test_cocina_zone_servicio(self, tmp_path):
        from src.core.presto_reader import read_presto
        f = _make_minimal_presto([("COCINA", 1000.0)], tmp_path)
        result = read_presto(f)
        spc = next(s for s in result["spaces"] if s["nombre"] == "COCINA")
        assert spc["zona"] == "SERVICIO"


# ===========================================================================
# 2. PRESTO READER — parse_presto()
# ===========================================================================

class TestParsePresto:
    def test_returns_dict_with_espacios(self, tmp_path):
        from src.core.presto_reader import parse_presto
        f = _make_minimal_presto([("SALA", 5000.0)], tmp_path)
        result = parse_presto(f)
        assert "espacios" in result

    def test_espacios_have_required_fields(self, tmp_path):
        from src.core.presto_reader import parse_presto
        f = _make_minimal_presto([("SALA", 1000.0)], tmp_path)
        result = parse_presto(f)
        for esp in result["espacios"]:
            for k in ("nombre", "zona", "planta", "coste", "m2", "partidas"):
                assert k in esp, f"Missing key in espacio: {k}"

    def test_planta_is_total(self, tmp_path):
        from src.core.presto_reader import parse_presto
        f = _make_minimal_presto([("SALA", 1000.0), ("COCINA", 2000.0)], tmp_path)
        result = parse_presto(f)
        for esp in result["espacios"]:
            assert esp["planta"] == "TOTAL", f"Expected TOTAL, got {esp['planta']}"

    def test_m2_is_zero(self, tmp_path):
        from src.core.presto_reader import parse_presto
        f = _make_minimal_presto([("SALA", 1000.0)], tmp_path)
        result = parse_presto(f)
        for esp in result["espacios"]:
            assert esp["m2"] == 0.0

    def test_partidas_is_list(self, tmp_path):
        from src.core.presto_reader import parse_presto
        f = _make_minimal_presto([("SALA", 1000.0)], tmp_path)
        result = parse_presto(f)
        for esp in result["espacios"]:
            assert isinstance(esp["partidas"], list)
            assert len(esp["partidas"]) >= 1

    def test_partida_coste_matches_space(self, tmp_path):
        from src.core.presto_reader import parse_presto
        f = _make_minimal_presto([("SALA", 9999.0)], tmp_path)
        result = parse_presto(f)
        sala = next(e for e in result["espacios"] if e["nombre"] == "SALA")
        assert abs(sala["partidas"][0]["coste"] - 9999.0) < 0.01

    def test_total_m2_is_zero(self, tmp_path):
        from src.core.presto_reader import parse_presto
        f = _make_minimal_presto([("SALA", 1000.0)], tmp_path)
        result = parse_presto(f)
        assert result["total_m2"] == 0.0

    def test_total_coste_matches(self, tmp_path):
        from src.core.presto_reader import parse_presto
        spaces = [("SALA", 1000.0), ("COCINA", 2000.0)]
        f = _make_minimal_presto(spaces, tmp_path)
        result = parse_presto(f)
        assert abs(result["total_coste"] - 3000.0) < 0.01

    def test_real_presto_sce(self):
        if not _PRESTO_SCE.exists():
            pytest.skip("Sample file not found")
        from src.core.presto_reader import parse_presto
        result = parse_presto(_PRESTO_SCE)
        assert len(result["espacios"]) > 1
        assert result["total_coste"] > 0


# ===========================================================================
# 3. SPACE CALCULATOR — calculate_space_ratios()
# ===========================================================================

class TestCalculateSpaceRatios:
    def _pres(self, spaces):
        return _presupuesto_no_m2(spaces)

    def test_returns_dict(self):
        from src.ratios.space_calculator import calculate_space_ratios
        result = calculate_space_ratios(self._pres([("SALA", 5000.0)]))
        assert isinstance(result, dict)

    def test_has_required_keys(self):
        from src.ratios.space_calculator import calculate_space_ratios
        result = calculate_space_ratios(self._pres([("SALA", 5000.0)]))
        for k in ("filename", "budget_code", "total_coste", "total_m2", "espacios"):
            assert k in result

    def test_espacio_has_required_keys(self):
        from src.ratios.space_calculator import calculate_space_ratios
        result = calculate_space_ratios(self._pres([("SALA", 5000.0)]))
        esp = result["espacios"][0]
        for k in ("nombre", "zona", "plantas", "total"):
            assert k in esp

    def test_plantas_have_ps_pb_pp(self):
        from src.ratios.space_calculator import calculate_space_ratios
        result = calculate_space_ratios(self._pres([("SALA", 5000.0)]))
        plantas = result["espacios"][0]["plantas"]
        assert "PS" in plantas and "PB" in plantas and "PP" in plantas

    def test_without_areas_m2_is_zero(self):
        from src.ratios.space_calculator import calculate_space_ratios
        result = calculate_space_ratios(self._pres([("SALA", 5000.0)]))
        assert result["total_m2"] == 0.0

    def test_without_m2_ratio_is_none(self):
        from src.ratios.space_calculator import calculate_space_ratios
        result = calculate_space_ratios(self._pres([("SALA", 5000.0)]))
        assert result["espacios"][0]["total"]["ratio"] is None

    def test_with_areas_total_m2(self):
        from src.ratios.space_calculator import calculate_space_ratios
        areas = {"SALA": {"PS": 30.0, "PB": 0.0, "PP": 0.0}}
        result = calculate_space_ratios(self._pres([("SALA", 6000.0)]), areas=areas)
        assert result["total_m2"] == pytest.approx(30.0, abs=0.01)

    def test_with_areas_ratio_calculated(self):
        from src.ratios.space_calculator import calculate_space_ratios
        areas = {"SALA": {"PS": 50.0, "PB": 0.0, "PP": 0.0}}
        result = calculate_space_ratios(self._pres([("SALA", 10000.0)]), areas=areas)
        esp = result["espacios"][0]
        assert esp["total"]["ratio"] == pytest.approx(200.0, abs=0.01)

    def test_with_areas_per_plant_split(self):
        from src.ratios.space_calculator import calculate_space_ratios
        areas = {"SALA": {"PS": 25.0, "PB": 25.0, "PP": 0.0}}
        result = calculate_space_ratios(self._pres([("SALA", 10000.0)]), areas=areas)
        plantas = result["espacios"][0]["plantas"]
        assert plantas["PS"]["m2"] == pytest.approx(25.0)
        assert plantas["PB"]["m2"] == pytest.approx(25.0)
        assert plantas["PP"]["m2"] == 0.0

    def test_proration_pct_sums_to_100(self):
        from src.ratios.space_calculator import calculate_space_ratios
        spaces = [("SALA", 5000.0), ("COCINA", 5000.0)]
        areas = {
            "SALA": {"PS": 50.0, "PB": 0.0, "PP": 0.0},
            "COCINA": {"PS": 50.0, "PB": 0.0, "PP": 0.0},
        }
        result = calculate_space_ratios(_presupuesto_no_m2(spaces), areas=areas)
        total_pct = sum(e["total"]["pct_m2"] for e in result["espacios"])
        assert total_pct == pytest.approx(100.0, abs=0.01)

    def test_proration_coste_sums_to_total(self):
        from src.ratios.space_calculator import calculate_space_ratios
        spaces = [("SALA", 3000.0), ("COCINA", 7000.0)]
        areas = {
            "SALA": {"PS": 30.0, "PB": 0.0, "PP": 0.0},
            "COCINA": {"PS": 70.0, "PB": 0.0, "PP": 0.0},
        }
        result = calculate_space_ratios(_presupuesto_no_m2(spaces), areas=areas)
        total_pro = sum(e["total"]["coste_prorrateado"] for e in result["espacios"])
        assert total_pro == pytest.approx(10000.0, abs=1.0)

    def test_plant_cost_split_proportional(self):
        from src.ratios.space_calculator import calculate_space_ratios
        areas = {"SALA": {"PS": 10.0, "PB": 30.0, "PP": 0.0}}
        result = calculate_space_ratios(_presupuesto_no_m2([("SALA", 4000.0)]), areas=areas)
        plantas = result["espacios"][0]["plantas"]
        assert plantas["PS"]["coste"] == pytest.approx(1000.0, abs=0.01)
        assert plantas["PB"]["coste"] == pytest.approx(3000.0, abs=0.01)

    def test_scalar_area_value(self):
        from src.ratios.space_calculator import calculate_space_ratios
        # Scalar area goes to PS, PB and PP = 0
        areas = {"SALA": 100.0}
        result = calculate_space_ratios(_presupuesto_no_m2([("SALA", 20000.0)]), areas=areas)
        assert result["total_m2"] == pytest.approx(100.0, abs=0.01)

    def test_zero_cost_space(self):
        from src.ratios.space_calculator import calculate_space_ratios
        result = calculate_space_ratios(_presupuesto_no_m2([("SALA", 0.0)]))
        assert result["espacios"][0]["total"]["ratio"] is None

    def test_filename_preserved(self):
        from src.ratios.space_calculator import calculate_space_ratios
        pres = _presupuesto_no_m2([("SALA", 1000.0)])
        pres["filename"] = "mi_archivo.Presto"
        result = calculate_space_ratios(pres)
        assert result["filename"] == "mi_archivo.Presto"

    def test_multiple_spaces_all_returned(self):
        from src.ratios.space_calculator import calculate_space_ratios
        spaces = [("SALA", 1000.0), ("COCINA", 2000.0), ("JARDIN", 3000.0)]
        result = calculate_space_ratios(_presupuesto_no_m2(spaces))
        assert len(result["espacios"]) == 3


# ===========================================================================
# 4. CALCULATE PRORATION
# ===========================================================================

class TestCalculateProration:
    def test_basic_proration(self):
        from src.ratios.space_calculator import calculate_proration
        espacios = [
            {"nombre": "SALA", "m2": 50.0, "coste": 5000.0},
            {"nombre": "COCINA", "m2": 50.0, "coste": 5000.0},
        ]
        result = calculate_proration(espacios, total_coste=20000.0)
        assert result[0]["coste_prorrateado"] == pytest.approx(10000.0, abs=0.01)
        assert result[1]["coste_prorrateado"] == pytest.approx(10000.0, abs=0.01)

    def test_pct_sums_to_100(self):
        from src.ratios.space_calculator import calculate_proration
        espacios = [
            {"nombre": "A", "m2": 30.0, "coste": 3000.0},
            {"nombre": "B", "m2": 70.0, "coste": 7000.0},
        ]
        result = calculate_proration(espacios, total_coste=10000.0)
        assert sum(e["pct_m2"] for e in result) == pytest.approx(100.0, abs=0.01)

    def test_zero_m2_fallback(self):
        from src.ratios.space_calculator import calculate_proration
        espacios = [{"nombre": "SALA", "m2": 0.0, "coste": 5000.0}]
        result = calculate_proration(espacios, total_coste=5000.0)
        assert result[0]["coste_prorrateado"] == 5000.0

    def test_ratio_computed_when_m2_nonzero(self):
        from src.ratios.space_calculator import calculate_proration
        espacios = [{"nombre": "SALA", "m2": 100.0, "coste": 10000.0}]
        result = calculate_proration(espacios, total_coste=10000.0)
        assert result[0]["ratio"] == pytest.approx(100.0, abs=0.01)

    def test_returns_same_list(self):
        from src.ratios.space_calculator import calculate_proration
        espacios = [{"nombre": "A", "m2": 10.0, "coste": 1000.0}]
        result = calculate_proration(espacios, total_coste=1000.0)
        assert result is espacios


# ===========================================================================
# 5. EXCEL GENERATOR
# ===========================================================================

class TestGenerateSpaceRatiosExcel:
    def _minimal_data(self):
        from src.ratios.space_calculator import calculate_space_ratios
        pres = _presupuesto_no_m2([
            ("SALA", 10000.0),
            ("COCINA", 5000.0),
            ("JARDIN", 3000.0),
        ])
        ratios = calculate_space_ratios(pres)
        return ratios, pres

    def test_creates_file(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "test_output.xlsx"
        result = generate_space_ratios_excel(ratios, pres, str(out))
        assert Path(result).exists()

    def test_returns_absolute_path(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        result = generate_space_ratios_excel(ratios, pres, str(out))
        assert Path(result).is_absolute()

    def test_has_areas_sheet(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out)
        assert "AREAS" in wb.sheetnames

    def test_has_ratios_estancia_sheet(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out)
        assert "RATIOS_ESTANCIA" in wb.sheetnames

    def test_areas_has_two_sheets_total(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 2

    def test_areas_sheet_has_plant_labels(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out, data_only=True)
        ws = wb["AREAS"]
        col_a_values = [ws.cell(r, 1).value for r in range(1, 20)]
        assert "PS" in col_a_values
        assert "PB" in col_a_values
        assert "PP" in col_a_values

    def test_areas_sheet_has_space_headers(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out, data_only=True)
        ws = wb["AREAS"]
        row3_values = [ws.cell(3, c).value for c in range(1, 30)]
        assert "SALA" in row3_values

    def test_areas_raw_data_has_plant_rows(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out, data_only=True)
        ws = wb["AREAS"]
        # Raw data starts at row 15; check plant values
        col_a_raw = [ws.cell(r, 1).value for r in range(15, 75)]
        assert "PS" in col_a_raw
        assert "PB" in col_a_raw
        assert "PP" in col_a_raw

    def test_areas_raw_data_has_space_names(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out, data_only=True)
        ws = wb["AREAS"]
        col_d_raw = [ws.cell(r, 4).value for r in range(15, 75)]
        assert "SALA" in col_d_raw

    def test_ratios_estancia_has_headers(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out, data_only=True)
        ws = wb["RATIOS_ESTANCIA"]
        row4_vals = [ws.cell(4, c).value for c in range(1, 11)]
        assert "Espacio" in row4_vals
        assert "Zona" in row4_vals

    def test_ratios_estancia_has_data_rows(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out, data_only=True)
        ws = wb["RATIOS_ESTANCIA"]
        # Row 5+ should have space names
        row5_a = ws.cell(5, 1).value
        assert row5_a is not None

    def test_areas_sumifs_formula_in_ps_row(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out, data_only=False)
        ws = wb["AREAS"]
        # Row 5, space columns should have formulas containing SUMIFS
        cell_val = ws.cell(5, 6).value  # first space column (SALA)
        assert cell_val is not None
        assert "SUMIFS" in str(cell_val).upper()

    def test_ratios_estancia_cross_ref_to_areas(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out, data_only=False)
        ws = wb["RATIOS_ESTANCIA"]
        # Col 3 (m² PS) in row 5 should reference AREAS
        cell_val = ws.cell(5, 3).value
        assert "AREAS" in str(cell_val)

    def test_coste_values_in_areas_row11(self, tmp_path):
        from src.export.space_ratios_generator import generate_space_ratios_excel
        ratios, pres = self._minimal_data()
        out = tmp_path / "out.xlsx"
        generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(out, data_only=True)
        ws = wb["AREAS"]
        # Row 11 = coste row; col 2 = total coste
        total_cell = ws.cell(11, 2).value
        assert total_cell == pytest.approx(18000.0, abs=1.0)

    def test_output_with_real_presto_sce(self, tmp_path):
        if not _PRESTO_SCE.exists():
            pytest.skip("Sample file not found")
        from src.core.presto_reader import parse_presto
        from src.ratios.space_calculator import calculate_space_ratios
        from src.export.space_ratios_generator import generate_space_ratios_excel
        pres = parse_presto(_PRESTO_SCE)
        ratios = calculate_space_ratios(pres)
        out = tmp_path / "sce_ratios.xlsx"
        result = generate_space_ratios_excel(ratios, pres, str(out))
        wb = load_workbook(result)
        assert "AREAS" in wb.sheetnames
        assert "RATIOS_ESTANCIA" in wb.sheetnames


# ===========================================================================
# 6. Zone / space resolution helpers
# ===========================================================================

class TestZoneResolution:
    @pytest.mark.parametrize("space,expected_zone", [
        ("SALA", "NOBLE"),
        ("COMEDOR", "NOBLE"),
        ("HABITACION MASTER", "NOBLE"),
        ("BAÑO MASTER", "NOBLE"),
        ("COCINA", "SERVICIO"),
        ("COCINA SERVICIO", "SERVICIO"),
        ("ZONAS DE SERVICIOS", "SERVICIO"),
        ("BALCONES", "EXTERIORES"),
        ("TERRAZAS", "EXTERIORES"),
        ("PISCINA", "EXTERIORES"),
        ("JARDIN", "EXTERIORES"),
        ("AMENITIES", "COMUNES"),
        ("PASILLOS", "COMUNES"),
        ("INSTALACIONES", "COMUNES"),
    ])
    def test_zone_mapping(self, space, expected_zone):
        from src.core.presto_reader import _resolve_zone
        assert _resolve_zone(space) == expected_zone

    def test_unknown_space_defaults_to_comunes(self):
        from src.core.presto_reader import _resolve_zone
        assert _resolve_zone("CUARTO_DESCONOCIDO") == "COMUNES"
