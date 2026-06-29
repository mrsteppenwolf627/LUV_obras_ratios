"""Contract tests for FASE MASTER — export filter (T2/T6).

All tests in this file are marked xfail(strict=True) because they depend on
generate_master_excel_approved() which will be implemented in T6.

xfail(strict=True) semantics:
  - While T6 is pending: ImportError from the missing function causes the
    test to FAIL → pytest records it as XFAIL → CI stays green.
  - Once T6 is implemented and a test PASSES: pytest records XPASS → CI
    fails → that is the signal to remove the xfail marker from that test.

The contract being specified:
  generate_master_excel_approved(session, output_path) must:
    - Accept a SQLAlchemy session and a file path.
    - Generate an Excel workbook at output_path.
    - Include in INDEX only Budget rows whose file_hash matches a
      BudgetImport with approval_status="APPROVED".
    - Exclude PENDING_REVIEW and REJECTED budgets.
    - Always name the output "LUV_RATIOS_MASTER.xlsx" when using the
      canonical path.

ADR-002: el dato bruto nunca se sobrescribe.
ADR-007: exclusión sin borrado del histórico.
ADR-019: Excel maestro vivo como salida principal del sistema.
ADR-020: identidad inequívoca del artefacto XLSX.
"""
from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from src.db.schema import Base, Budget, BudgetImport


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine, "connect")
    def _fk_on(conn, _):
        conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    s = Session()
    yield s
    s.close()


@pytest.fixture
def tmp_excel(tmp_path: Path) -> Path:
    return tmp_path / "LUV_RATIOS_MASTER.xlsx"


def _sha256(seed: str) -> str:
    return hashlib.sha256(seed.encode()).hexdigest()


def _seed_budget(session, seed: str, approval_status: str, technical_status: str = "success"):
    """Insert a BudgetImport + matching Budget row (same file_hash).

    The Budget row is what get_all_budgets() queries; the BudgetImport is what
    generate_master_excel_approved() must filter on.
    """
    file_hash = _sha256(seed)
    filename = f"{seed}.xlsx"

    bi = BudgetImport(
        filename=filename,
        file_hash=file_hash,
        building_type="residencial",
        status=technical_status,
        approval_status=approval_status,
    )
    session.add(bi)
    session.flush()

    b = Budget(
        filename=filename,
        file_hash=file_hash,
        source_format="excel",
        total_cost=50_000.0,
    )
    session.add(b)
    session.flush()
    return bi, b


def _filenames_in_index(wb) -> list[str]:
    """Extract filename column from the INDEX sheet (col 2, rows 2+)."""
    ws = wb["INDEX"]
    return [
        ws.cell(row=r, column=2).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=2).value
    ]


# ---------------------------------------------------------------------------
# Test 5 — REJECTED no aparece en el export
# ---------------------------------------------------------------------------

@pytest.mark.xfail(
    strict=True,
    reason=(
        "generate_master_excel_approved() not implemented yet (T6 pending). "
        "src/export/excel_master_generator.py does not yet filter by approval_status."
    ),
)
def test_rejected_import_never_feeds_master_export(session, tmp_excel):
    """Un presupuesto REJECTED no debe aparecer en LUV_RATIOS_MASTER.xlsx.

    ADR-007: exclusión sin borrado del histórico.
    El dato bruto sigue en la BD; el artefacto exportado solo refleja datos aprobados.
    """
    from src.export.excel_master_generator import generate_master_excel_approved  # not yet

    _seed_budget(session, "mi_t2_exp5_rejected", approval_status="REJECTED")
    _seed_budget(session, "mi_t2_exp5_approved", approval_status="APPROVED")

    output = generate_master_excel_approved(session, output_path=tmp_excel)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    names = _filenames_in_index(wb)

    assert "mi_t2_exp5_rejected.xlsx" not in names, (
        "Un presupuesto REJECTED no debe aparecer en el INDEX del master export."
    )
    assert "mi_t2_exp5_approved.xlsx" in names, (
        "Un presupuesto APPROVED debe aparecer en el INDEX del master export."
    )


# ---------------------------------------------------------------------------
# Test 6 — export solo incluye APPROVED
# ---------------------------------------------------------------------------

@pytest.mark.xfail(
    strict=True,
    reason=(
        "generate_master_excel_approved() not implemented yet (T6 pending). "
        "src/export/excel_master_generator.py does not yet filter by approval_status."
    ),
)
def test_master_export_uses_only_approved_imports(session, tmp_excel):
    """El exportador oficial debe incluir únicamente presupuestos APPROVED.

    PENDING_REVIEW y REJECTED quedan excluidos del LUV_RATIOS_MASTER.xlsx.
    Solo los datos validados y aprobados consolidan el master.
    """
    from src.export.excel_master_generator import generate_master_excel_approved  # not yet

    _seed_budget(session, "mi_t2_exp6_pending", approval_status="PENDING_REVIEW")
    _seed_budget(session, "mi_t2_exp6_rejected", approval_status="REJECTED")
    _seed_budget(session, "mi_t2_exp6_approved_a", approval_status="APPROVED")
    _seed_budget(session, "mi_t2_exp6_approved_b", approval_status="APPROVED")

    output = generate_master_excel_approved(session, output_path=tmp_excel)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    names = _filenames_in_index(wb)

    assert "mi_t2_exp6_pending.xlsx" not in names
    assert "mi_t2_exp6_rejected.xlsx" not in names
    assert "mi_t2_exp6_approved_a.xlsx" in names
    assert "mi_t2_exp6_approved_b.xlsx" in names
    assert len(names) == 2, (
        f"Solo 2 presupuestos APPROVED deben aparecer en el INDEX, "
        f"pero hay {len(names)}: {names}"
    )


@pytest.mark.xfail(
    strict=True,
    reason=(
        "generate_master_excel_approved() not implemented yet (T6 pending). "
        "Export with zero approved budgets behaviour undefined."
    ),
)
def test_master_export_with_no_approved_imports_generates_empty_sheets(session, tmp_excel):
    """Con cero presupuestos APPROVED el export debe generar hojas vacías (sin crash).

    El sistema debe ser robusto: exportar aunque no haya datos aprobados aún,
    devolviendo un workbook válido con encabezados pero sin filas de datos.
    """
    from src.export.excel_master_generator import generate_master_excel_approved  # not yet

    _seed_budget(session, "mi_t2_exp6b_pending_only", approval_status="PENDING_REVIEW")

    output = generate_master_excel_approved(session, output_path=tmp_excel)

    import openpyxl
    wb = openpyxl.load_workbook(output)
    names = _filenames_in_index(wb)
    assert names == [], (
        f"Con cero APPROVED, el INDEX debe estar vacío, pero tiene: {names}"
    )


# ---------------------------------------------------------------------------
# Test — nombre del archivo oficial
# ---------------------------------------------------------------------------

@pytest.mark.xfail(
    strict=True,
    reason="generate_master_excel_approved() not implemented yet (T6 pending)",
)
def test_master_export_filename_is_luv_ratios_master(session, tmp_excel):
    """El archivo de exportación oficial debe llamarse LUV_RATIOS_MASTER.xlsx.

    ADR-020: identidad inequívoca del artefacto XLSX entregado para revisión humana.
    El nombre canónico no varía entre generaciones para que la última versión
    siempre sobrescriba a la anterior en la carpeta de entrega.
    """
    from src.export.excel_master_generator import generate_master_excel_approved  # not yet

    _seed_budget(session, "mi_t2_exp_name", approval_status="APPROVED")

    output_path_str = generate_master_excel_approved(session, output_path=tmp_excel)

    assert Path(output_path_str).name == "LUV_RATIOS_MASTER.xlsx", (
        f"El archivo oficial debe llamarse LUV_RATIOS_MASTER.xlsx, "
        f"pero se generó como {Path(output_path_str).name!r}."
    )
