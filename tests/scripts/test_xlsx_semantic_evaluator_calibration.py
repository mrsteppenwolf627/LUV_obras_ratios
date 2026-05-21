from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import generate_preview_from_real_xlsx
from scripts.live_excel_dry_run_evaluator import evaluate_dry_run_workbook


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"xlsx_eval_semantic_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_summary_input(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(["Codigo", "Descripcion", "Formula", "Importe"])
    ws.append(["LUV_AP", "EQUIPAMIENTO", "=D2/PEM", 37297.09])
    ws.append(["LUV_REV", "13 abril", "", 2026])
    wb.save(path)
    wb.close()


def _build_comparison_input(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(["Cap.", "Nombre del capítulo", "Importe (€)", "Nombre equivalente", "Importe equivalente", "Diferencia"])
    ws.append([2, "Demoliciones", 687.5, "DEMOLICIONES", 550, "=+H4-F4"])
    wb.save(path)
    wb.close()


def test_evaluator_does_not_block_summary_for_classic_amount_split_rules():
    root = _make_root()
    try:
        src = root / "input_summary.xlsx"
        _build_summary_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview_summary.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_eval_summary")

        wb = load_workbook(out)
        try:
            view = wb["IMPORTED_BUDGET_VIEW"]
            headers = [str(view.cell(row=1, column=i).value or "").strip() for i in range(1, view.max_column + 1)]
            idx = {h: pos + 1 for pos, h in enumerate(headers)}
            view.cell(row=2, column=idx["item_description"], value="13 abril")
            view.cell(row=2, column=idx["amount"], value="2026")
            view.cell(
                row=2,
                column=idx["notes"],
                value="DESCRIPTION_AMOUNT_SPLIT_FAILED|SHEET_TYPE=BUDGET_SUMMARY|NUMERIC_PARSE_AMBIGUOUS",
            )
            wb.save(out)
        finally:
            wb.close()

        evaluation = evaluate_dry_run_workbook(out, run_id="eval_summary")
        assert "description_amount_split_failed" not in evaluation.reasons
        assert "amount_mixed_in_description" not in evaluation.reasons
    finally:
        _cleanup(root)


def test_evaluator_allows_comparison_without_classic_cost_item_mapping():
    root = _make_root()
    try:
        src = root / "input_comparison.xlsx"
        _build_comparison_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview_comparison.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_eval_comparison")

        evaluation = evaluate_dry_run_workbook(out, run_id="eval_comparison")
        assert evaluation.state in {"OPERATIVE_CANDIDATE", "MANUAL_REVIEW_REQUIRED"}
        assert "cost_item_mapping_ambiguous" not in evaluation.reasons
        assert "cost_item_mapping_low_confidence" not in evaluation.reasons
    finally:
        _cleanup(root)
