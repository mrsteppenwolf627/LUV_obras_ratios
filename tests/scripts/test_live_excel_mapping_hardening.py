from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import generate_preview_from_real_xlsx
from scripts.live_excel_dry_run_evaluator import evaluate_dry_run_workbook


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"mapping_hardening_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_budget(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws.append(["Codigo", "Descripcion", "Unidad", "Cantidad", "Precio Unitario", "Importe"])
    ws.append(["1", "Capitulo sintetico", "", "", "", ""])
    ws.append(["1.1.01", "Partida valida", "ud", 2, 10, 20])
    ws.append(["", "Subtotal capitulo", "", "", "", 20])
    ws.append(["", "Total presupuesto", "", "", "", 20])
    ws.append(["", "", "", "", "", ""])
    ws.append(["texto auxiliar", "", "", "", "", ""])
    wb.save(path)
    wb.close()


def test_mapping_distinguishes_cost_items_from_non_budget_rows():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_budget(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_mapping_001")

        wb = load_workbook(out)
        try:
            ws = wb["PRESERVED_TO_COST_ITEMS_MAP"]
            statuses = [
                str(ws.cell(row=row, column=10).value or "")
                for row in range(2, ws.max_row + 1)
            ]
            assert "MAPPED" in statuses
            assert "NOT_COST_ITEM" in statuses
            assert statuses.count("MAPPED") == 1
        finally:
            wb.close()

        result = evaluate_dry_run_workbook(out)
        assert result.metrics["mapped_rows"] == 1.0
        assert result.metrics["not_cost_item_rows"] >= 3.0
        assert result.metrics["mapping_rate_on_candidate_cost_items"] == 1.0
    finally:
        _cleanup(root)


def test_ambiguous_mapping_status_requires_manual_review():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_budget(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_mapping_002")

        wb = load_workbook(out)
        try:
            ws = wb["PRESERVED_TO_COST_ITEMS_MAP"]
            ws.cell(row=2, column=10, value="AMBIGUOUS")
            wb.save(out)
        finally:
            wb.close()

        result = evaluate_dry_run_workbook(out)
        assert result.state == "MANUAL_REVIEW_REQUIRED"
        assert "ambiguous_mapping" in result.reasons
        assert "cost_item_mapping_ambiguous" in result.reasons
    finally:
        _cleanup(root)


def test_no_mapping_is_invented_when_description_is_missing():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws.append(["Codigo", "Descripcion", "Importe"])
        ws.append(["1.1.01", "", 100])
        wb.save(src)
        wb.close()

        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_mapping_003")
        result = evaluate_dry_run_workbook(out)
        assert result.metrics["mapped_rows"] == 0.0
        assert result.metrics["not_cost_item_rows"] >= 1.0
    finally:
        _cleanup(root)
