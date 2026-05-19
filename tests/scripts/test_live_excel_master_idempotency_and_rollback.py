from pathlib import Path
import shutil
import uuid

import pytest
from openpyxl import load_workbook

from scripts.generate_live_excel_master import (
    load_synthetic_incremental,
    rollback_master_from_snapshot,
    validate_workbook_file,
)


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"live_master_idempotent_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def test_idempotent_synthetic_load_with_same_run_id():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        run_id = "run_fixed_idempotency_001"
        first = load_synthetic_incremental(output_path=output, retention_max=5, run_id=run_id)
        second = load_synthetic_incremental(output_path=output, retention_max=5, run_id=run_id)

        assert first["idempotent_skip"] == "false"
        assert second["idempotent_skip"] == "true"

        wb = load_workbook(output)
        try:
            import_log = wb["IMPORT_LOG"]
            run_ids = [
                str(import_log.cell(row=row, column=2).value or "")
                for row in range(2, import_log.max_row + 1)
            ]
            assert run_ids.count(run_id) == 1
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_snapshot_checksum_is_sha256():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5, run_id="run_checksum_001")
        load_synthetic_incremental(output_path=output, retention_max=5, run_id="run_checksum_002")

        wb = load_workbook(output)
        try:
            ws = wb["SNAPSHOTS"]
            assert ws.max_row >= 2
            checksum = str(ws.cell(row=2, column=7).value or "")
            assert len(checksum) == 64
            int(checksum, 16)
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_rollback_from_missing_snapshot_fails():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5, run_id="run_missing_snapshot")
        with pytest.raises(RuntimeError, match="Snapshot path does not exist"):
            rollback_master_from_snapshot(
                output_path=output,
                snapshot_path=output.parent / "snapshots" / "missing.xlsx",
                retention_max=5,
            )
    finally:
        _cleanup(root)


def test_rollback_from_corrupt_snapshot_fails_and_restores_previous_master():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5, run_id="run_base_ok")
        before_bytes = output.read_bytes()

        bad_snapshot = output.parent / "snapshots" / "bad_snapshot.xlsx"
        bad_snapshot.parent.mkdir(parents=True, exist_ok=True)
        bad_snapshot.write_text("not an excel", encoding="utf-8")

        with pytest.raises(RuntimeError, match="invalid snapshot content"):
            rollback_master_from_snapshot(output_path=output, snapshot_path=bad_snapshot, retention_max=5)

        after_bytes = output.read_bytes()
        assert after_bytes == before_bytes
        validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_rollback_from_schema_invalid_snapshot_fails_and_restores_previous_master():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5, run_id="run_base_schema_ok")
        before_bytes = output.read_bytes()

        broken_snapshot = output.parent / "snapshots" / "broken_snapshot.xlsx"
        shutil.copy2(output, broken_snapshot)
        wb = load_workbook(broken_snapshot)
        wb.remove(wb["CHANGELOG"])
        wb.save(broken_snapshot)
        wb.close()

        with pytest.raises(RuntimeError, match="invalid snapshot content"):
            rollback_master_from_snapshot(output_path=output, snapshot_path=broken_snapshot, retention_max=5)

        after_bytes = output.read_bytes()
        assert after_bytes == before_bytes
        validate_workbook_file(output)
    finally:
        _cleanup(root)
