from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import generate_preview_from_real_xlsx


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"xlsx_review_navigation_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_multi_sheet_input(path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Datos"
    ws1.append(["Codigo", "Descripcion", "Formula", "Importe"])
    ws1.append(["LUV_AP", "EQUIPAMIENTO", "=D2/PEM", 37297.09])
    ws2 = wb.create_sheet("Espacios")
    ws2.append(["Codigo", "Info", "Resumen", "Presupuesto"])
    ws2.append(["BALCON", "", "BALCON", 1000])
    wb.save(path)
    wb.close()


def test_index_and_home_list_adaptive_views():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_multi_sheet_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_nav_001")

        wb = load_workbook(out)
        try:
            index = wb["INDEX"]
            headers = [str(index.cell(row=4, column=i).value or "").strip() for i in range(1, 7)]
            assert headers == ["Vista principal", "Tipo semantico", "Hoja origen", "Estado", "Descripcion", "Abrir"]
            text_blob = " | ".join(str(index.cell(row=r, column=1).value or "") for r in range(5, min(index.max_row, 20) + 1))
            assert "BUDGET_REVIEW_001_Datos" in text_blob
            assert "BUDGET_REVIEW_001_Espacios" in text_blob

            home = wb["BUDGET_REVIEW_001"]
            home_headers = [str(home.cell(row=4, column=i).value or "").strip() for i in range(1, 8)]
            assert home_headers == [
                "Hoja origen",
                "Tipo semantico",
                "Confianza",
                "Vista profesional",
                "Estado",
                "Advertencias",
                "Accion recomendada",
            ]
            assert "Datos" == str(home.cell(row=5, column=1).value)
            assert str(home.cell(row=5, column=7).value or "").strip() != ""
        finally:
            wb.close()
    finally:
        _cleanup(root)
