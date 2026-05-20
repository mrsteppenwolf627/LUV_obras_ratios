from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import (
    OPERATIONAL_PREVIEW_COLUMNS,
    OPERATIONAL_PREVIEW_SHEET,
    generate_preview_from_real_xlsx,
)
from scripts.live_excel_dry_run_evaluator import (
    STATE_OPERATIVE_CANDIDATE,
    STATE_PROMOTION_BLOCKED,
    evaluate_dry_run_workbook,
)


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"xlsx_economic_extraction_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _preview(src: Path, root: Path) -> Path:
    out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
    generate_preview_from_real_xlsx(src, out, source_file_id="sf_synth_xlsx_001")
    return out


def test_separates_description_unit_quantity_price_and_amount():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws.append(["metadata"])
        ws.append(["CÓDIGO", "DESCRIPCIÓN\nCONCEPTO", "Ud.", "Medición", "P. Unitario", "Importe Total"])
        ws.append(["1.1.01", "Partida sintetica", "m2", "2,5", "10,00", "25,00 €"])
        wb.save(src)
        wb.close()

        out = _preview(src, root)
        preview = load_workbook(out)
        try:
            ws_out = preview[OPERATIONAL_PREVIEW_SHEET]
            headers = [ws_out.cell(row=1, column=i).value for i in range(1, len(OPERATIONAL_PREVIEW_COLUMNS) + 1)]
            row = {header: ws_out.cell(row=2, column=idx).value for idx, header in enumerate(headers, start=1)}
            assert row["item_description"] == "Partida sintetica"
            assert "25" not in str(row["item_description"])
            assert row["unit"] == "m2"
            assert str(row["quantity"]) == "2.5"
            assert str(row["unit_price"]) == "10"
            assert str(row["amount"]) == "25"
            assert row["preview_only"] == "TRUE"
            assert preview["RATIO_INPUTS"].max_row == 1
            assert preview["RATIOS_CALCULATED"].max_row == 1
        finally:
            preview.close()

        result = evaluate_dry_run_workbook(out)
        assert result.state == STATE_OPERATIVE_CANDIDATE
    finally:
        _cleanup(root)


def test_partial_header_uses_text_column_instead_of_numeric_code_as_description():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws.append([])
        ws.append([])
        ws.append(["", "", "", "", "Capitulo", "Importe"])
        ws.append(["", "", "", 1, "Partida con cabecera parcial", "1.234,56"])
        wb.save(src)
        wb.close()

        out = _preview(src, root)
        preview = load_workbook(out)
        try:
            ws_out = preview[OPERATIONAL_PREVIEW_SHEET]
            headers = [ws_out.cell(row=1, column=i).value for i in range(1, len(OPERATIONAL_PREVIEW_COLUMNS) + 1)]
            row = {header: ws_out.cell(row=2, column=idx).value for idx, header in enumerate(headers, start=1)}
            assert row["item_description"] == "Partida con cabecera parcial"
            assert str(row["amount"]) == "1234.56"
        finally:
            preview.close()

        result = evaluate_dry_run_workbook(out)
        assert result.state == STATE_OPERATIVE_CANDIDATE
        assert result.metrics["amount_separation_rate"] == 1.0
    finally:
        _cleanup(root)


def test_extracts_clear_trailing_amount_from_description_with_low_confidence_note():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws.append(["Codigo", "Descripcion"])
        ws.append(["1.1.01", "Partida sintetica con importe 1.234,56 €"])
        wb.save(src)
        wb.close()

        out = _preview(src, root)
        preview = load_workbook(out)
        try:
            ws_out = preview[OPERATIONAL_PREVIEW_SHEET]
            headers = [ws_out.cell(row=1, column=i).value for i in range(1, len(OPERATIONAL_PREVIEW_COLUMNS) + 1)]
            row = {header: ws_out.cell(row=2, column=idx).value for idx, header in enumerate(headers, start=1)}
            assert row["item_description"] == "Partida sintetica con importe"
            assert str(row["amount"]) == "1234.56"
            assert "AMOUNT_EXTRACTED_FROM_DESCRIPTION_LOW_CONFIDENCE" in str(row["notes"])
        finally:
            preview.close()
    finally:
        _cleanup(root)


def test_preview_with_amount_mixed_in_description_is_blocked_when_split_fails():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws.append(["Codigo", "Descripcion", "Importe"])
        ws.append(["1.1.01", "Partida 1.234,56 no separable en medio", "1.234,56"])
        wb.save(src)
        wb.close()

        out = _preview(src, root)
        result = evaluate_dry_run_workbook(out)
        assert result.state == STATE_PROMOTION_BLOCKED
        assert "description_amount_split_failed" in result.reasons
    finally:
        _cleanup(root)
