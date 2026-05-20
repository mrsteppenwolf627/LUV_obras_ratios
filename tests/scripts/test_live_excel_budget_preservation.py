from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import generate_preview_from_real_xlsx, validate_workbook_file
from scripts.live_excel_preservation import sanitize_sheet_name


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"live_master_preserved_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_multi_sheet_input(path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Capitulos 2026"
    ws1.append(["Resumen presupuesto"])
    ws1.append(["Codigo", "Descripcion", "Importe"])
    ws1.append(["1.1", "Movimiento de tierras", 1200])
    ws2 = wb.create_sheet("Partidas Detalle")
    ws2.append(["Codigo", "Descripcion", "Unidad", "Cantidad", "Precio Unitario", "Importe"])
    ws2.append(["1.1.01", "Excavacion", "m3", 10, 30, 300])
    wb.save(path)
    wb.close()


def test_preservation_index_and_map_sheets_created():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_multi_sheet_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_preview_pres_001")
        wb = load_workbook(out)
        try:
            assert "PRESERVED_BUDGETS_INDEX" in wb.sheetnames
            assert "PRESERVED_BUDGET_SHEETS" in wb.sheetnames
            assert "PRESERVED_TO_COST_ITEMS_MAP" in wb.sheetnames
            assert wb["PRESERVED_BUDGETS_INDEX"].max_row >= 2
            assert wb["PRESERVED_BUDGET_SHEETS"].max_row >= 3
            assert wb["PRESERVED_TO_COST_ITEMS_MAP"].max_row >= 2
        finally:
            wb.close()
        validate_workbook_file(out)
    finally:
        _cleanup(root)


def test_preserved_sheet_names_are_sanitized_and_unique_across_runs():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_multi_sheet_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_preview_pres_002")
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_preview_pres_003")
        wb = load_workbook(out)
        try:
            preserved_names = [n for n in wb.sheetnames if n.startswith("PRES_")]
            assert len(preserved_names) >= 4
            assert len(set(preserved_names)) == len(preserved_names)
            for name in preserved_names:
                assert len(name) <= 31
                assert all(ch not in name for ch in "[]:*?/\\")
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_preserved_rows_keep_traceability_and_mapping_status():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_multi_sheet_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_preview_pres_004")
        wb = load_workbook(out)
        try:
            first_preserved_name = next(name for name in wb.sheetnames if name.startswith("PRES_"))
            ws = wb[first_preserved_name]
            headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
            assert "__source_sheet_name" in headers
            assert "__source_row_number" in headers
            assert "__source_column_number" in headers

            map_ws = wb["PRESERVED_TO_COST_ITEMS_MAP"]
            statuses = {
                str(map_ws.cell(row=row, column=10).value or "")
                for row in range(2, map_ws.max_row + 1)
            }
            assert "MAPPED" in statuses or "UNMAPPED" in statuses
            assert "UNMAPPED" in statuses or "NOT_COST_ITEM" in statuses
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_sanitize_sheet_name_strips_invalid_characters():
    bad_name = "Presupuesto[Obra]:Fase*1?/\\"
    sanitized = sanitize_sheet_name(bad_name)
    assert all(ch not in sanitized for ch in "[]:*?/\\")
    assert len(sanitized) <= 31
