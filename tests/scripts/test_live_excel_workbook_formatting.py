from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import generate_preview_from_real_xlsx


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"workbook_formatting_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_input(path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Capitulos Presupuesto"
    ws1.append(["Codigo", "Descripcion", "Ud", "Cantidad", "Precio Unitario", "Importe"])
    ws1.append(["1", "CAPITULO MOVIMIENTO DE TIERRAS", "", "", "", ""])
    ws1.append(["1.1.01", "Excavacion", "m3", 10, 30, ""])
    ws1.append(["", "Subtotal capitulo", "", "", "", 300])
    ws1.append(["", "Total presupuesto", "", "", "", 300])
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Campo", "Valor"])
    ws2.append(["Observaciones", "Sin datos sensibles"])
    wb.save(path)
    wb.close()


def _first_sheet_with_prefix(wb: object, prefix: str) -> str:
    return next(name for name in wb.sheetnames if name.startswith(prefix))


def test_index_exists_and_order_is_human_first():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_format_001")

        wb = load_workbook(out)
        try:
            review = _first_sheet_with_prefix(wb, "BUDGET_REVIEW_")
            trace = _first_sheet_with_prefix(wb, "BUDGET_REVIEW_TRACE_")
            assert wb.sheetnames[0] == "INDEX"
            assert wb.sheetnames[1] == review
            assert wb.sheetnames.index(trace) > wb.sheetnames.index(review)

            index_ws = wb["INDEX"]
            assert "PREVIEW_ONLY" in str(index_ws["A2"].value)
            assert "BUDGET_REVIEW_001" in str(index_ws["A3"].value)

            preserved_sheet = _first_sheet_with_prefix(wb, "PRES_")
            assert wb.sheetnames.index("COST_ITEMS") > wb.sheetnames.index(preserved_sheet)
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_preserved_sheets_keep_original_columns_before_technical_columns():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_format_002")

        wb = load_workbook(out)
        try:
            preserved = _first_sheet_with_prefix(wb, "PRES_")
            ws = wb[preserved]
            headers = [str(ws.cell(row=1, column=idx).value or "") for idx in range(1, ws.max_column + 1)]
            assert "__source_sheet_name" in headers
            assert "__source_row_number" in headers
            assert "__source_column_number" in headers

            first_tech = headers.index("__source_sheet_name")
            assert headers[:6] == ["Codigo", "Descripcion", "Ud", "Cantidad", "Precio Unitario", "Importe"]
            assert all(value == "" for value in headers[6:first_tech])
            assert ws.column_dimensions[ws.cell(row=1, column=first_tech + 1).column_letter].hidden is True
            assert ws.freeze_panes == "A2"
            assert ws.auto_filter.ref == f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}1"
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_technical_sheets_are_formatted_and_pending_sheets_are_marked():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_format_003")

        wb = load_workbook(out)
        try:
            for technical_sheet in ["COST_ITEMS", "VALIDATION_RESULTS"]:
                ws = wb[technical_sheet]
                assert ws.freeze_panes == "A2"
                assert ws.auto_filter.ref == f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}1"
                assert ws["A1"].font.bold is True

            ratio_ws = wb["RATIO_INPUTS"]
            assert ratio_ws.max_row == 1
            assert ratio_ws["A1"].comment is not None
            assert "PREVIEW_ONLY" in str(ratio_ws["A1"].comment.text)

            ratios_calc_ws = wb["RATIOS_CALCULATED"]
            assert ratios_calc_ws.max_row == 1
            assert ratios_calc_ws["A1"].comment is not None
        finally:
            wb.close()
    finally:
        _cleanup(root)
