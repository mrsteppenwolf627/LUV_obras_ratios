from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import generate_preview_from_real_xlsx, validate_generated_xlsx_preview


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"xlsx_adaptive_reviews_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_input(path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Datos"
    ws1.append(["Codigo", "Descripcion", "Formula", "Importe"])
    ws1.append(["LUV_AP", "EQUIPAMIENTO", "=D2/PEM", 37297.09])
    ws1.append([None, "REVISION abril", "", 2026])

    ws2 = wb.create_sheet("Hoja1")
    ws2.append(["Cap.", "Nombre del capítulo", "Importe (€)", "Nombre equivalente", "Importe equivalente", "Diferencia"])
    ws2.append([2, "Demoliciones", 687.5, "DEMOLICIONES", 550, "=E2-C2"])
    wb.save(path)
    wb.close()


def test_adaptive_reviews_do_not_mix_semantically_incompatible_sheets():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_adaptive_001")
        validate_generated_xlsx_preview(out)

        wb = load_workbook(out)
        try:
            home = wb["BUDGET_REVIEW_001"]
            entries = {}
            for row in range(5, home.max_row + 1):
                source = str(home.cell(row=row, column=1).value or "").strip()
                view = str(home.cell(row=row, column=4).value or "").strip()
                if source:
                    entries[source] = view
            assert "Datos" in entries
            assert "Hoja1" in entries
            assert entries["Datos"] != entries["Hoja1"]

            datos = wb[entries["Datos"]]
            comp = wb[entries["Hoja1"]]
            datos_headers = [str(datos.cell(row=4, column=i).value or "") for i in range(1, 6)]
            comp_headers = [str(comp.cell(row=4, column=i).value or "") for i in range(1, 8)]
            assert datos_headers[:4] == ["Codigo", "Descripcion", "Formula / Ratio", "Importe"]
            assert comp_headers[:6] == [
                "Cap.",
                "Nombre del capítulo",
                "Importe (€)",
                "Nombre equivalente",
                "Importe equivalente",
                "Diferencia",
            ]
            assert "Cantidad" not in comp_headers
            assert "Precio unitario" not in comp_headers

            # Revision metadata row is excluded from budget summary view.
            desc_values = [str(datos.cell(row=r, column=2).value or "") for r in range(5, datos.max_row + 1)]
            assert all("REVISION" not in value.upper() for value in desc_values)
        finally:
            wb.close()
    finally:
        _cleanup(root)
