from pathlib import Path
import shutil
import uuid

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from scripts.run_xlsx_generalization_dry_run import (
    ALLOWED_INPUT_ROOT,
    ALLOWED_OUTPUT_ROOT,
    PREVIEW_PIPELINE_PHASE,
    SANITIZED_IDS,
    run_xlsx_generalization_dry_run,
)


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"xlsx_generalization_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    ws.append(["Codigo", "Descripcion", "Unidad", "Cantidad", "Precio Unitario", "Importe"])
    ws.append(["1", "Capitulo", "", "", "", ""])
    ws.append(["1.1.01", "Partida de prueba", "ud", 2, 50, 100])
    wb.save(path)
    wb.close()


def test_generalization_rejects_input_outside_allowed_zone():
    root = _make_root()
    try:
        allowed_in = root / ALLOWED_INPUT_ROOT
        allowed_in.mkdir(parents=True, exist_ok=True)
        outside = root / "outside.xlsx"
        _build_xlsx(outside)
        with pytest.raises(ValueError, match="input file must stay under"):
            run_xlsx_generalization_dry_run(
                files=[outside],
                output_dir=ALLOWED_OUTPUT_ROOT,
                repo_root=root,
            )
    finally:
        _cleanup(root)


def test_generalization_rejects_non_xlsx_input():
    root = _make_root()
    try:
        allowed_in = root / ALLOWED_INPUT_ROOT
        allowed_in.mkdir(parents=True, exist_ok=True)
        invalid = allowed_in / "sample.bc3"
        invalid.write_text("~V|FIEBDC-3/2020", encoding="utf-8")
        with pytest.raises(ValueError, match="Only .xlsx is allowed"):
            run_xlsx_generalization_dry_run(
                files=[Path("data/samples/sample.bc3")],
                output_dir=ALLOWED_OUTPUT_ROOT,
                repo_root=root,
            )
    finally:
        _cleanup(root)


def test_generalization_generates_sanitized_results_without_ratio_ingestion():
    root = _make_root()
    try:
        allowed_in = root / ALLOWED_INPUT_ROOT
        allowed_in.mkdir(parents=True, exist_ok=True)
        file1 = allowed_in / "sample1.xlsx"
        file2 = allowed_in / "sample2.xlsx"
        _build_xlsx(file1)
        _build_xlsx(file2)

        report = run_xlsx_generalization_dry_run(
            files=[Path("data/samples/sample1.xlsx"), Path("data/samples/sample2.xlsx")],
            output_dir=ALLOWED_OUTPUT_ROOT,
            repo_root=root,
        )

        assert report["phase"] == PREVIEW_PIPELINE_PHASE
        assert report["auto_promotion_enabled"] is False
        assert len(report["results"]) == 2
        assert report["results"][0]["run_id"] == SANITIZED_IDS[0]
        assert report["results"][1]["run_id"] == SANITIZED_IDS[1]
        for result in report["results"]:
            preview = Path(result["preview_output"])
            assert str((root / ALLOWED_OUTPUT_ROOT).resolve()) in str(preview.resolve())
            assert result["metrics"]["ratio_input_rows"] == 0.0
            assert result["metrics"]["ratio_calculated_rows"] == 0.0

        first_preview = Path(report["results"][0]["preview_output"])
        wb = load_workbook(first_preview)
        try:
            assert "INDEX" in wb.sheetnames
            assert "BUDGET_REVIEW_001" in wb.sheetnames
            assert "BUDGET_REVIEW_TRACE_001" in wb.sheetnames
            assert wb.sheetnames[0] == "INDEX"
            assert wb.sheetnames[1] == "BUDGET_REVIEW_001"
            assert wb.sheetnames[2] == "BUDGET_REVIEW_TRACE_001"
            assert wb.active.title == "INDEX"
            if wb.views:
                assert wb.views[0].activeTab == 0
                assert wb.views[0].firstSheet == 0
            for row_idx in range(2, wb["README_MASTER"].max_row + 1):
                key = str(wb["README_MASTER"].cell(row=row_idx, column=1).value or "").strip().lower()
                if key == "phase":
                    assert str(wb["README_MASTER"].cell(row=row_idx, column=2).value or "") == PREVIEW_PIPELINE_PHASE
                    break
            else:
                raise AssertionError("README_MASTER phase row not found")
            for row in range(5, min(wb["INDEX"].max_row, 60) + 1):
                value = wb["INDEX"].cell(row=row, column=1).value
                if isinstance(value, str):
                    assert not value.startswith("=HYPERLINK(")
                    assert "HYPERLINK is not implemented" not in value

            pres_sheets = [name for name in wb.sheetnames if name.startswith("PRES_")]
            assert pres_sheets
            for sheet_name in pres_sheets:
                ws = wb[sheet_name]
                headers = {
                    str(ws.cell(row=1, column=col_idx).value or ""): ws.column_dimensions[
                        ws.cell(row=1, column=col_idx).column_letter
                    ].hidden
                    for col_idx in range(1, ws.max_column + 1)
                }
                assert headers.get("__source_sheet_name") is True
                assert headers.get("__source_row_number") is True
                assert headers.get("__source_column_number") is True
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_generalization_limit_rejects_more_than_five_files():
    root = _make_root()
    try:
        allowed_in = root / ALLOWED_INPUT_ROOT
        allowed_in.mkdir(parents=True, exist_ok=True)
        files: list[Path] = []
        for idx in range(6):
            path = allowed_in / f"sample{idx + 1}.xlsx"
            _build_xlsx(path)
            files.append(Path(f"data/samples/sample{idx + 1}.xlsx"))
        with pytest.raises(ValueError, match="Maximum supported files"):
            run_xlsx_generalization_dry_run(
                files=files,
                output_dir=ALLOWED_OUTPUT_ROOT,
                repo_root=root,
            )
    finally:
        _cleanup(root)
