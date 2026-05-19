from pathlib import Path
import shutil
import uuid

import pytest
from openpyxl import load_workbook

from scripts.generate_live_excel_master import (
    ReferentialValidationError,
    SchemaValidationError,
    generate_master,
    load_synthetic_incremental,
    rollback_master_from_snapshot,
    validate_workbook_file,
)


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"live_master_integrity_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def test_column_order_change_is_allowed_when_headers_exist():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output_path=output, update=False, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["IMPORT_LOG"]
            first = ws.cell(row=1, column=1).value
            second = ws.cell(row=1, column=2).value
            ws.cell(row=1, column=1, value=second)
            ws.cell(row=1, column=2, value=first)
            wb.save(output)
        finally:
            wb.close()
        validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_misspelled_required_column_fails():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output_path=output, update=False, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["COST_ITEMS"]
            ws.cell(row=1, column=1, value="cost_item_id_typo")
            wb.save(output)
        finally:
            wb.close()
        with pytest.raises(SchemaValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_unknown_validation_status_fails():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["RATIO_INPUTS"]
            ws.cell(row=2, column=7, value="NOT_A_VALID_STATUS")
            wb.save(output)
        finally:
            wb.close()
        with pytest.raises(ReferentialValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_manual_review_required_cannot_promote_to_ratio_inputs():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["RATIO_INPUTS"]
            ws.cell(row=2, column=7, value="MANUAL_REVIEW_REQUIRED")
            wb.save(output)
        finally:
            wb.close()
        with pytest.raises(ReferentialValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_normalized_cost_item_requires_existing_cost_item():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["NORMALIZED_COST_ITEMS"]
            ws.append(
                [
                    "nci_bad",
                    "missing_cost_item",
                    "desc",
                    "ud",
                    "1",
                    "10",
                    "PENDING",
                    "rule",
                    "hash",
                    "VALIDATED",
                ]
            )
            wb.save(output)
        finally:
            wb.close()
        with pytest.raises(ReferentialValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_rollback_from_snapshot_outside_allowed_path_fails():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        external = root / "external_snapshot.xlsx"
        shutil.copy2(output, external)
        with pytest.raises(ValueError):
            rollback_master_from_snapshot(output_path=output, snapshot_path=external, retention_max=5)
    finally:
        _cleanup(root)

