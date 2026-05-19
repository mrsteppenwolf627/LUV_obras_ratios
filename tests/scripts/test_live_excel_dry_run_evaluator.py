from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import generate_preview_from_real_xlsx
from scripts.live_excel_dry_run_evaluator import (
    STATE_MANUAL_REVIEW_REQUIRED,
    STATE_OPERATIVE_CANDIDATE,
    STATE_PRESERVATION_INCOMPLETE,
    STATE_PROMOTION_BLOCKED,
    evaluate_dry_run_workbook,
)


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"live_master_dryrun_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_input(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    ws.append(["Capitulo Codigo", "Capitulo", "Codigo", "Descripcion", "Unidad", "Cantidad", "Precio Unitario", "Importe"])
    ws.append(["2", "Demoliciones", "2.1.01", "Demolicion de tabiqueria", "m2", 10, 68.75, 687.50])
    wb.save(path)
    wb.close()


def _make_preview_output(root: Path) -> Path:
    src = root / "input.xlsx"
    _build_input(src)
    out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
    generate_preview_from_real_xlsx(src, out, source_file_id="sf_preview_eval_001")
    return out


def test_valid_preserved_preview_is_operative_candidate():
    root = _make_root()
    try:
        out = _make_preview_output(root)
        result = evaluate_dry_run_workbook(out, run_id="run_eval_ok")
        assert result.state == STATE_OPERATIVE_CANDIDATE
        assert result.auto_promotion_enabled is False
        assert result.metrics["total_preview_rows"] >= 1
        assert result.metrics["total_preserved_rows"] >= 1
    finally:
        _cleanup(root)


def test_missing_preserved_visible_sheet_is_preservation_incomplete():
    root = _make_root()
    try:
        out = _make_preview_output(root)
        wb = load_workbook(out)
        try:
            to_remove = [name for name in wb.sheetnames if name.startswith("PRES_")]
            for name in to_remove:
                wb.remove(wb[name])
            wb.save(out)
        finally:
            wb.close()
        result = evaluate_dry_run_workbook(out)
        assert result.state == STATE_PRESERVATION_INCOMPLETE
        assert "missing_preserved_sheet" in result.reasons
    finally:
        _cleanup(root)


def test_missing_preserved_index_is_preservation_incomplete():
    root = _make_root()
    try:
        out = _make_preview_output(root)
        wb = load_workbook(out)
        try:
            wb.remove(wb["PRESERVED_BUDGETS_INDEX"])
            wb.save(out)
        finally:
            wb.close()
        result = evaluate_dry_run_workbook(out)
        assert result.state == STATE_PRESERVATION_INCOMPLETE
        assert "missing_preserved_budgets_index" in result.reasons
    finally:
        _cleanup(root)


def test_missing_mapping_sheet_is_preservation_incomplete():
    root = _make_root()
    try:
        out = _make_preview_output(root)
        wb = load_workbook(out)
        try:
            wb.remove(wb["PRESERVED_TO_COST_ITEMS_MAP"])
            wb.save(out)
        finally:
            wb.close()
        result = evaluate_dry_run_workbook(out)
        assert result.state == STATE_PRESERVATION_INCOMPLETE
        assert "missing_preserved_to_cost_item_mapping" in result.reasons
    finally:
        _cleanup(root)


def test_ambiguous_mapping_requires_manual_review():
    root = _make_root()
    try:
        out = _make_preview_output(root)
        wb = load_workbook(out)
        try:
            wb["PRESERVED_TO_COST_ITEMS_MAP"].cell(row=2, column=10, value="AMBIGUOUS")
            wb.save(out)
        finally:
            wb.close()
        result = evaluate_dry_run_workbook(out)
        assert result.state == STATE_MANUAL_REVIEW_REQUIRED
        assert "ambiguous_mapping" in result.reasons
    finally:
        _cleanup(root)


def test_unknown_validation_status_is_promotion_blocked():
    root = _make_root()
    try:
        out = _make_preview_output(root)
        wb = load_workbook(out)
        try:
            wb["IMPORTED_BUDGET_VIEW"].cell(row=2, column=16, value="NOT_A_STATUS")
            wb.save(out)
        finally:
            wb.close()
        result = evaluate_dry_run_workbook(out)
        assert result.state == STATE_PROMOTION_BLOCKED
        assert "unknown_validation_status" in result.reasons
    finally:
        _cleanup(root)


def test_blocked_preview_status_is_promotion_blocked():
    root = _make_root()
    try:
        out = _make_preview_output(root)
        wb = load_workbook(out)
        try:
            wb["IMPORTED_BUDGET_VIEW"].cell(row=2, column=16, value="BLOCKED")
            wb.save(out)
        finally:
            wb.close()
        result = evaluate_dry_run_workbook(out)
        assert result.state == STATE_PROMOTION_BLOCKED
        assert "blocked_validation_status" in result.reasons
    finally:
        _cleanup(root)


def test_ratio_inputs_populated_is_promotion_blocked():
    root = _make_root()
    try:
        out = _make_preview_output(root)
        wb = load_workbook(out)
        try:
            wb["RATIO_INPUTS"].append(["ri_1", "nci_1", "p1", "bv1", "ELIGIBLE", "false", "VALIDATED", "2026-05-19"])
            wb.save(out)
        finally:
            wb.close()
        result = evaluate_dry_run_workbook(out)
        assert result.state == STATE_PROMOTION_BLOCKED
        assert "ratio_inputs_not_allowed" in result.reasons
    finally:
        _cleanup(root)

