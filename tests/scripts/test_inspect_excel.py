from pathlib import Path
import json
import shutil
import uuid

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

from scripts.inspect_excel import inspect_excel_file, inspect_excel_samples, write_reports


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"inspect_excel_{uuid.uuid4().hex}"
    (root / "data" / "samples").mkdir(parents=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def test_header_in_row_1_detected():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["Codigo", "Descripcion", "Unidad", "Cantidad", "Precio", "Importe"])
        ws.append(["IT01", "Excavacion", "m3", 10, 15.5, "=D2*E2"])
        file_path = root / "data" / "samples" / "h1.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/h1.xlsx")
        sheet = result["sheets"][0]
        assert 1 in sheet["candidate_header_rows"]
        assert sheet["candidate_columns"]["importe"]
    finally:
        _cleanup(root)


def test_header_in_row_5_detected():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["", "", ""])
        ws.append(["", "", ""])
        ws.append(["Notas", "", ""])
        ws.append(["", "", ""])
        ws.append(["Cod", "Descripcion", "Total"])
        ws.append(["P01", "Partida", 1200])
        file_path = root / "data" / "samples" / "h5.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/h5.xlsx")
        sheet = result["sheets"][0]
        assert 5 in sheet["candidate_header_rows"]
    finally:
        _cleanup(root)


def test_multiline_header_detected():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["Mediciones", "", "", ""])
        ws.append(["Codigo", "Descripcion", "Cantidad", "Importe"])
        ws.append(["A1", "Texto", 2, 30])
        file_path = root / "data" / "samples" / "multi.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/multi.xlsx")
        sheet = result["sheets"][0]
        assert len(sheet["candidate_header_rows"]) >= 1
    finally:
        _cleanup(root)


def test_numeric_columns_without_headers_detected():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["", "", ""])
        ws.append([2, 10.5, 1200])
        ws.append([4, 11.0, 2000])
        ws.append([3, 12.0, 3000])
        file_path = root / "data" / "samples" / "num_no_header.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/num_no_header.xlsx")
        cols = result["sheets"][0]["candidate_columns"]
        assert cols["cantidad"] or cols["precio"] or cols["importe"]
    finally:
        _cleanup(root)


def test_formulas_detected():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["a", "b", "importe"])
        ws.append([2, 3, "=A2*B2"])
        file_path = root / "data" / "samples" / "formula.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/formula.xlsx")
        assert result["sheets"][0]["formula_cells_count"] == 1
    finally:
        _cleanup(root)


def test_merged_cells_detected():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Cabecera"
        ws.merge_cells("A1:C1")
        ws.append(["", "", ""])
        ws.append(["codigo", "descripcion", "importe"])
        ws.append(["X", "Y", 10])
        file_path = root / "data" / "samples" / "merged.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/merged.xlsx")
        assert result["sheets"][0]["merged_cells_count"] == 1
    finally:
        _cleanup(root)


def test_sparse_sheet_density_profile():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "X"
        ws["Z40"] = "Y"
        file_path = root / "data" / "samples" / "sparse.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/sparse.xlsx")
        sheet = result["sheets"][0]
        assert sheet["density_profile"]["non_empty_pct"] < 5
        assert sheet["used_range"]["min_row"] == 1
        assert sheet["used_range"]["max_row"] == 40
    finally:
        _cleanup(root)


def test_visual_non_tabular_sheet():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "TITULO"
        ws.merge_cells("A1:F1")
        ws["A20"] = "NOTA"
        file_path = root / "data" / "samples" / "visual.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/visual.xlsx")
        sheet = result["sheets"][0]
        assert sheet["is_likely_tabular"] is False
    finally:
        _cleanup(root)


def test_chartsheet_ignored_as_non_worksheet_profile():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Base"
        ws.append(["x", "y"])
        ws.append([1, 2])
        ws.append([2, 3])
        chart = BarChart()
        data = Reference(ws, min_col=2, min_row=1, max_row=3)
        chart.add_data(data, titles_from_data=True)
        ch = wb.create_chartsheet(title="Grafica")
        ch.add_chart(chart)
        file_path = root / "data" / "samples" / "chart.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/chart.xlsx")
        assert result["summary"]["chartsheets_count"] == 1
        chart_sheet = [s for s in result["sheets"] if s["sheet_type"] == "CHARTSHEET"][0]
        assert chart_sheet["candidate_headers"] == []
    finally:
        _cleanup(root)


def test_sanitized_samples_present_and_limited():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["Codigo", "Descripcion", "Importe"])
        ws.append(["A01", "Texto largo de ejemplo para sanitizacion", 12.5])
        ws.append(["A02", "Otro texto", 15])
        file_path = root / "data" / "samples" / "sanitize.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/sanitize.xlsx")
        sample = result["sheets"][0]["sanitized_samples"]
        assert sample["first_non_empty_rows"]
        assert sample["dense_rows"]
    finally:
        _cleanup(root)


def test_simple_table_detected():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["codigo", "descripcion"])
        ws.append(["IT01", "A"])
        ws.append(["IT02", "B"])
        tab = Table(displayName="Tabla1", ref="A1:B3")
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        file_path = root / "data" / "samples" / "table.xlsx"
        wb.save(file_path)

        result = inspect_excel_file(file_path, "data/samples/table.xlsx")
        assert "Tabla1" in result["sheets"][0]["possible_tables"]
    finally:
        _cleanup(root)


def test_inventory_and_reports_generation():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["Codigo", "Descripcion"])
        ws.append(["IT01", "Desc"])
        wb.save(root / "data" / "samples" / "inv.xlsx")

        payload = inspect_excel_samples(root)
        json_path, md_path = write_reports(root, payload)

        assert payload["excel_files_detected"] == 1
        assert json.loads(json_path.read_text(encoding="utf-8"))["excel_files_detected"] == 1
        assert "Excel Diagnostics Inventory" in md_path.read_text(encoding="utf-8")
    finally:
        _cleanup(root)


def test_does_not_modify_input_file():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["Codigo", "Descripcion"])
        ws.append(["IT01", "Desc"])
        file_path = root / "data" / "samples" / "stable.xlsx"
        wb.save(file_path)
        before = file_path.read_bytes()

        _ = inspect_excel_file(file_path, "data/samples/stable.xlsx")
        after = file_path.read_bytes()

        assert before == after
    finally:
        _cleanup(root)
