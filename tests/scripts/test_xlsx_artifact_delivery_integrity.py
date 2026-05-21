"""Phase 9.20: artifact delivery integrity tests.

These tests enforce that the file generated equals the file validated equals the
file delivered for human review: unambiguous folder, unambiguous filenames, a
manifest binding filename <-> SHA-256 <-> phase, and validation performed after
re-opening the artifact from disk (never an in-memory-only workbook).
"""

from __future__ import annotations

import hashlib
from pathlib import Path
import shutil
import subprocess
import uuid

import pytest
from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import (
    PREVIEW_PIPELINE_PHASE,
    validate_generated_xlsx_preview,
)
from scripts.run_xlsx_generalization_dry_run import (
    ALLOWED_INPUT_ROOT,
    ALLOWED_OUTPUT_BOUNDARY,
    run_xlsx_generalization_dry_run,
)

REVIEW_DIR = ALLOWED_OUTPUT_BOUNDARY / "manual_review_phase_9_20"
NAME_TEMPLATE = "phase_9_20_review_{index:03d}"


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"artifact_integrity_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_classic_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(["Codigo", "Descripcion", "Unidad", "Cantidad", "Precio Unitario", "Importe"])
    ws.append(["1", "Capitulo demoliciones", "", "", "", ""])
    ws.append(["1.1.01", "Partida de prueba", "ud", 2, 50, 100])
    ws.append(["1.1.02", "Otra partida", "m2", 4, 25, 100])
    wb.save(path)
    wb.close()


def _build_comparison_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(
        [
            "Cap.",
            "Nombre del capitulo",
            "Importe (€)",
            "Nombre equivalente",
            "Importe equivalente",
            "Diferencia",
        ]
    )
    ws.append(["1", "Actuaciones previas", 1000, "Prev works", 1100, 100])
    ws.append(["2", "Demoliciones", 2000, "Demolitions", 1900, -100])
    ws.append(["3", "Movimientos de tierra", 3000, "Earthworks", 3000, 0])
    wb.save(path)
    wb.close()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(8192), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _run(root: Path, build_funcs: list) -> dict:
    allowed_in = root / ALLOWED_INPUT_ROOT
    allowed_in.mkdir(parents=True, exist_ok=True)
    files = []
    for idx, builder in enumerate(build_funcs, start=1):
        sample = allowed_in / f"sample{idx}.xlsx"
        builder(sample)
        files.append(Path(f"data/samples/sample{idx}.xlsx"))
    return run_xlsx_generalization_dry_run(
        files=files,
        output_dir=REVIEW_DIR,
        repo_root=root,
        name_template=NAME_TEMPLATE,
        manifest_path=REVIEW_DIR / "MANIFEST_phase_9_20.json",
    )


def test_generation_can_write_to_new_review_folder_with_unambiguous_names():
    root = _make_root()
    try:
        report = _run(root, [_build_classic_xlsx, _build_comparison_xlsx])
        out_dir = (root / REVIEW_DIR).resolve()
        assert out_dir.exists()
        f1 = out_dir / "phase_9_20_review_001.xlsx"
        f2 = out_dir / "phase_9_20_review_002.xlsx"
        assert f1.exists() and f2.exists()
        # Legacy generic names must not be produced in this review folder.
        assert not (out_dir / "xlsx_generalization_001_preview.xlsx").exists()
        assert report["phase"] == PREVIEW_PIPELINE_PHASE
    finally:
        _cleanup(root)


def test_manifest_contains_sha256_matching_real_file_on_disk():
    root = _make_root()
    try:
        report = _run(root, [_build_classic_xlsx, _build_comparison_xlsx])
        manifest = report["manifest"]
        assert len(manifest) == 2
        for entry in manifest:
            assert entry["sha256"]
            real_path = Path(entry["output_absolute_path"])
            assert real_path.exists()
            assert entry["sha256"] == _sha256(real_path)
            assert entry["size_bytes"] == real_path.stat().st_size
    finally:
        _cleanup(root)


def test_manifest_reports_phase_9_20_and_index_active():
    root = _make_root()
    try:
        report = _run(root, [_build_classic_xlsx, _build_comparison_xlsx])
        for entry in report["manifest"]:
            assert entry["readme_phase"] == PREVIEW_PIPELINE_PHASE == "9.20"
            assert entry["active_sheet"] == "INDEX"
            assert entry["human_review_start_sheet"] == "INDEX"
            assert entry["required_sheets_present"] is True
            assert entry["validation_status"] == "PASSED"
    finally:
        _cleanup(root)


def test_file_001_has_adaptive_views_in_synthetic_scenario():
    root = _make_root()
    try:
        report = _run(root, [_build_classic_xlsx, _build_comparison_xlsx])
        entry = report["manifest"][0]
        assert entry["adaptive_views_present"] is True
        assert entry["trace_sheets"]
        assert "INDEX" in entry["sheets"]
    finally:
        _cleanup(root)


def test_file_002_has_comparison_view_in_synthetic_scenario():
    root = _make_root()
    try:
        report = _run(root, [_build_classic_xlsx, _build_comparison_xlsx])
        f2 = Path(report["manifest"][1]["output_absolute_path"])
        wb = load_workbook(f2, data_only=False)
        try:
            comparison_views = [
                name
                for name in wb.sheetnames
                if name.startswith("BUDGET_REVIEW_001_")
            ]
            assert comparison_views
            view = wb[comparison_views[0]]
            headers = [str(view.cell(row=4, column=c).value or "") for c in range(1, view.max_column + 1)]
            for expected in ["Cap.", "Importe equivalente", "Diferencia"]:
                assert expected in headers
            assert "Cantidad" not in headers
            assert "Precio unitario" not in headers
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_validation_reads_from_disk_not_memory():
    """If the on-disk file is tampered (phase downgraded), validation must fail.

    This proves the validator reopens the artifact from disk: an in-memory-only
    validation would not notice a post-save mutation.
    """
    root = _make_root()
    try:
        report = _run(root, [_build_classic_xlsx, _build_comparison_xlsx])
        target = Path(report["manifest"][0]["output_absolute_path"])
        # Initially passes from disk.
        validate_generated_xlsx_preview(target, required_phase=PREVIEW_PIPELINE_PHASE)
        # Tamper README phase on disk to 9.9.
        wb = load_workbook(target)
        try:
            ws = wb["README_MASTER"]
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=1).value or "").strip().lower() == "phase":
                    ws.cell(row=r, column=2, value="9.9")
                    break
            wb.save(target)
        finally:
            wb.close()
        with pytest.raises(RuntimeError, match="readme_phase_mismatch"):
            validate_generated_xlsx_preview(target, required_phase=PREVIEW_PIPELINE_PHASE)
    finally:
        _cleanup(root)


def test_missing_index_fails_validation():
    root = _make_root()
    try:
        bad = root / "no_index.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "README_MASTER"
        ws.append(["key", "value", "updated_at", "updated_by"])
        ws.append(["phase", PREVIEW_PIPELINE_PHASE, "", "system"])
        wb.create_sheet("BUDGET_REVIEW_001")
        wb.create_sheet("BUDGET_REVIEW_TRACE_001")
        wb.save(bad)
        wb.close()
        with pytest.raises(RuntimeError, match="missing_INDEX"):
            validate_generated_xlsx_preview(bad, required_phase=PREVIEW_PIPELINE_PHASE)
    finally:
        _cleanup(root)


def test_missing_trace_fails_validation():
    root = _make_root()
    try:
        report = _run(root, [_build_classic_xlsx, _build_comparison_xlsx])
        target = Path(report["manifest"][0]["output_absolute_path"])
        wb = load_workbook(target)
        try:
            for name in list(wb.sheetnames):
                if name.startswith("BUDGET_REVIEW_TRACE_"):
                    del wb[name]
            wb.save(target)
        finally:
            wb.close()
        with pytest.raises(RuntimeError, match="missing_BUDGET_REVIEW_TRACE"):
            validate_generated_xlsx_preview(target, required_phase=PREVIEW_PIPELINE_PHASE)
    finally:
        _cleanup(root)


def test_review_outputs_are_git_ignored():
    """Review artifacts must never be added to git."""
    repo_root = Path(__file__).resolve().parents[2]
    candidate = "outputs/live_excel_master/manual_review_phase_9_20/phase_9_20_review_001.xlsx"
    result = subprocess.run(
        ["git", "check-ignore", candidate],
        cwd=repo_root,
        capture_output=True,
        text=True,
    )
    # check-ignore exits 0 when the path IS ignored.
    assert result.returncode == 0, f"Review output is not git-ignored: {candidate}"


def test_output_dir_outside_boundary_is_rejected():
    root = _make_root()
    try:
        allowed_in = root / ALLOWED_INPUT_ROOT
        allowed_in.mkdir(parents=True, exist_ok=True)
        _build_classic_xlsx(allowed_in / "sample1.xlsx")
        with pytest.raises(ValueError, match="output_dir must stay under"):
            run_xlsx_generalization_dry_run(
                files=[Path("data/samples/sample1.xlsx")],
                output_dir=Path("outputs/elsewhere"),
                repo_root=root,
                name_template=NAME_TEMPLATE,
            )
    finally:
        _cleanup(root)
