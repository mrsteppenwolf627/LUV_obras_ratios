from pathlib import Path
import shutil
import uuid

import pytest
from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import PREVIEW_PIPELINE_PHASE, validate_generated_xlsx_preview
from scripts.run_xlsx_generalization_dry_run import ALLOWED_INPUT_ROOT, ALLOWED_OUTPUT_ROOT, run_xlsx_generalization_dry_run


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"xlsx_output_pipeline_audit_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_semantic_input(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(["Resumen General"])
    ws.append([None, None, None, None])
    ws.append(["Codigo", "Descripcion", "Ratio", "Importe"])
    ws.append(["LUV_MOB", "MOBILIARIO A MEDIDA  Y ACCESORIOS", "=D4 / PEM", 125511.25])
    ws.append(["LUV_AP", "EQUIPAMIENTO", "=D5 / PEM", 37297.09])
    ws.append(["LUV_SA", "APARATOS SANITARIOS", "=D6 / PEM", 14461.06])
    ws.append([None, "HONORARIOS PROYECTO", 0, "=PEM * PorHonPry"])
    ws.append([None, "Deducciones", 0, "=HonPry * PorDedPry"])
    wb.save(path)
    wb.close()


def _review_sheet_name(wb: object) -> str:
    return next(
        name for name in wb.sheetnames if name.startswith("BUDGET_REVIEW_") and not name.startswith("BUDGET_REVIEW_TRACE_")
    )


def test_generalization_pipeline_enforces_professional_output_and_semantics():
    root = _make_root()
    try:
        allowed_in = root / ALLOWED_INPUT_ROOT
        allowed_in.mkdir(parents=True, exist_ok=True)
        src = allowed_in / "sample.xlsx"
        _build_semantic_input(src)

        report = run_xlsx_generalization_dry_run(
            files=[Path("data/samples/sample.xlsx")],
            output_dir=ALLOWED_OUTPUT_ROOT,
            repo_root=root,
        )
        preview = Path(report["results"][0]["preview_output"])
        validate_generated_xlsx_preview(preview, required_phase=PREVIEW_PIPELINE_PHASE)

        wb = load_workbook(preview)
        try:
            assert wb.active.title == "INDEX"
            assert "INDEX" in wb.sheetnames
            assert "BUDGET_REVIEW_001" in wb.sheetnames
            assert "BUDGET_REVIEW_TRACE_001" in wb.sheetnames

            phase_value = ""
            for row_idx in range(2, wb["README_MASTER"].max_row + 1):
                key = str(wb["README_MASTER"].cell(row=row_idx, column=1).value or "").strip().lower()
                if key == "phase":
                    phase_value = str(wb["README_MASTER"].cell(row=row_idx, column=2).value or "").strip()
                    break
            assert phase_value == PREVIEW_PIPELINE_PHASE

            index = wb["INDEX"]
            for row_idx in range(5, min(index.max_row, 60) + 1):
                text = str(index.cell(row=row_idx, column=1).value or "")
                assert not text.startswith("=HYPERLINK(")
                assert "HYPERLINK is not implemented" not in text

            review = wb[_review_sheet_name(wb)]
            luv_ap = []
            for row_idx in range(5, review.max_row + 1):
                code = review.cell(row=row_idx, column=1).value
                desc = review.cell(row=row_idx, column=2).value
                amount = review.cell(row=row_idx, column=6).value
                if code == "LUV_AP":
                    luv_ap.append((desc, amount))
                if isinstance(desc, str):
                    assert not desc.startswith("=")
                    assert "#NAME?" not in desc
                    assert not desc.replace(".", "").replace(",", "").isdigit()
            assert luv_ap, "Expected LUV_AP row in review output"
            assert luv_ap[0][0] == "EQUIPAMIENTO"
            assert str(luv_ap[0][1]) == "37297.09"
            assert bool(review.column_dimensions["G"].hidden)

            cost = wb["COST_ITEMS"]
            descriptions = [str(cost.cell(row=row, column=5).value or "") for row in range(2, cost.max_row + 1)]
            assert all(not value.startswith("=") for value in descriptions)
            assert all("HONPRY" not in value.upper() for value in descriptions)
            assert all("HONDIR" not in value.upper() for value in descriptions)
            assert all("PEM" not in value.upper() for value in descriptions)
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_preview_validator_fails_when_index_uses_formula_hyperlink():
    root = _make_root()
    try:
        out = root / "outputs" / "live_excel_master" / "preview" / "invalid.xlsx"
        out.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "INDEX"
        ws["A1"] = "INDEX"
        ws["A5"] = '=HYPERLINK("#\'BUDGET_REVIEW_001\'!A1","BUDGET_REVIEW_001")'
        ws_review = wb.create_sheet("BUDGET_REVIEW_001")
        ws_review["A4"] = "Codigo"
        ws_review["B4"] = "Descripcion"
        ws_review["F4"] = "Importe"
        ws_review["G4"] = "_review_row_id"
        ws_trace = wb.create_sheet("BUDGET_REVIEW_TRACE_001")
        ws_trace["A1"] = "review_row_id"
        readme = wb.create_sheet("README_MASTER")
        readme.append(["field", "value", "updated_at", "updated_by"])
        readme.append(["phase", PREVIEW_PIPELINE_PHASE, "2026-05-21T00:00:00+00:00", "system"])
        wb.save(out)
        wb.close()

        with pytest.raises(RuntimeError, match="index_formula_hyperlink"):
            validate_generated_xlsx_preview(out, required_phase=PREVIEW_PIPELINE_PHASE)
    finally:
        _cleanup(root)


def test_preview_validator_fails_when_required_review_sheet_is_missing():
    root = _make_root()
    try:
        out = root / "outputs" / "live_excel_master" / "preview" / "invalid_missing_review.xlsx"
        out.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "INDEX"
        ws["A1"] = "INDEX"
        ws_trace = wb.create_sheet("BUDGET_REVIEW_TRACE_001")
        ws_trace["A1"] = "review_row_id"
        readme = wb.create_sheet("README_MASTER")
        readme.append(["field", "value", "updated_at", "updated_by"])
        readme.append(["phase", PREVIEW_PIPELINE_PHASE, "2026-05-21T00:00:00+00:00", "system"])
        wb.save(out)
        wb.close()

        with pytest.raises(RuntimeError, match="missing_BUDGET_REVIEW"):
            validate_generated_xlsx_preview(out, required_phase=PREVIEW_PIPELINE_PHASE)
    finally:
        _cleanup(root)


def test_preview_validator_fails_when_review_has_name_error_or_cost_formula():
    root = _make_root()
    try:
        out = root / "outputs" / "live_excel_master" / "preview" / "invalid_name_and_formula.xlsx"
        out.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        index = wb.active
        index.title = "INDEX"
        review = wb.create_sheet("BUDGET_REVIEW_001")
        trace = wb.create_sheet("BUDGET_REVIEW_TRACE_001")
        readme = wb.create_sheet("README_MASTER")
        cost = wb.create_sheet("COST_ITEMS")

        review["A4"] = "Codigo"
        review["B4"] = "Descripcion"
        review["F4"] = "Importe"
        review["G4"] = "_review_row_id"
        review.column_dimensions["G"].hidden = True
        review["A5"] = "LUV_AP"
        review["B5"] = "#NAME?"
        review["F5"] = 10

        trace["A1"] = "review_row_id"
        readme.append(["field", "value", "updated_at", "updated_by"])
        readme.append(["phase", PREVIEW_PIPELINE_PHASE, "2026-05-21T00:00:00+00:00", "system"])

        cost.append(
            [
                "cost_item_id",
                "budget_version_id",
                "source_file_id",
                "source_row_ref",
                "description_raw",
                "unit_raw",
                "quantity_raw",
                "unit_price_raw",
                "amount_raw",
                "row_hash",
                "validation_status",
            ]
        )
        cost.append(["ci_1", "bv_1", "sf_1", "Datos!1", "=PEM * PorHonPry", "", "", "", "0", "hash_1", "PENDING"])
        wb.save(out)
        wb.close()

        with pytest.raises(RuntimeError, match="review_name_error_rows|cost_items_formula_description_row"):
            validate_generated_xlsx_preview(out, required_phase=PREVIEW_PIPELINE_PHASE)
    finally:
        _cleanup(root)
