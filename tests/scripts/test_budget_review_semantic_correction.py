from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import generate_preview_from_real_xlsx


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"budget_review_semantic_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _review_home_and_details(wb: object) -> tuple[str, list[str]]:
    homes = sorted(
        name
        for name in wb.sheetnames
        if name.startswith("BUDGET_REVIEW_")
        and not name.startswith("BUDGET_REVIEW_TRACE_")
        and len(name.split("_")) == 3
    )
    home = homes[0]
    details = sorted(
        name
        for name in wb.sheetnames
        if name.startswith(home + "_")
    )
    return home, details


def _build_semantic_input(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(["Resumen General"])
    ws.append([None, None, None, None])
    ws.append(["Codigo", "Descripcion", "Ratio", "Importe"])
    ws.append(["LUV_MOB", "MOBILIARIO A MEDIDA  Y ACCESORIOS", "=D4 / PEM", 125511.25])
    ws.append(["LUV_AP", "EQUIPAMIENTO", "=D5 / PEM", 37297.09])
    ws.append(["LUV_SA", "APARATOS SANITARIOS", "=D6 / PEM", 14461.06])
    ws.append([None, "HONORARIOS PROYECTO", 0, "=PEM * PorHonPry"])
    ws.append([None, "Deducciones", 0, "=HonPry * PorDedPry"])
    wb.save(path)
    wb.close()


def test_budget_review_keeps_description_and_amount_semantics():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_semantic_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_semantic_001")

        wb = load_workbook(out)
        try:
            _home, detail_sheets = _review_home_and_details(wb)
            assert detail_sheets
            rows = []
            for sheet_name in detail_sheets:
                review = wb[sheet_name]
                headers = [str(review.cell(row=4, column=idx).value or "").strip() for idx in range(1, review.max_column + 1)]
                if "Codigo" not in headers or "Descripcion" not in headers:
                    continue
                code_col = headers.index("Codigo") + 1
                desc_col = headers.index("Descripcion") + 1
                amount_col = headers.index("Importe") + 1 if "Importe" in headers else None
                for row in range(5, review.max_row + 1):
                    rows.append(
                        (
                            review.cell(row=row, column=code_col).value,
                            review.cell(row=row, column=desc_col).value,
                            review.cell(row=row, column=amount_col).value if amount_col else None,
                        )
                    )
            target = [entry for entry in rows if entry[0] == "LUV_AP"]
            assert target, "Expected LUV_AP row in BUDGET_REVIEW"
            assert target[0][1] == "EQUIPAMIENTO"
            assert str(target[0][2]) == "37297.09"
            assert str(target[0][1]) != "37297.09"
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_index_uses_clean_text_without_hyperlink_formula():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_semantic_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_semantic_002")

        wb = load_workbook(out)
        try:
            index = wb["INDEX"]
            assert "PREVIEW_ONLY" in str(index["A2"].value or "")
            for row in range(5, min(index.max_row, 60) + 1):
                value = index.cell(row=row, column=1).value
                if value is None:
                    continue
                text = str(value)
                assert "HYPERLINK is not implemented" not in text
                assert not text.startswith("=HYPERLINK(")
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_auxiliary_formula_rows_do_not_become_cost_items_or_name_errors():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_semantic_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_semantic_003")

        wb = load_workbook(out)
        try:
            _home, detail_sheets = _review_home_and_details(wb)
            for sheet_name in detail_sheets:
                review = wb[sheet_name]
                headers = [str(review.cell(row=4, column=idx).value or "").strip() for idx in range(1, review.max_column + 1)]
                if "Descripcion" not in headers:
                    continue
                desc_col = headers.index("Descripcion") + 1
                for row in range(5, review.max_row + 1):
                    desc = review.cell(row=row, column=desc_col).value
                    if isinstance(desc, str):
                        assert not desc.startswith("=")
                        assert "#NAME?" not in desc

            cost = wb["COST_ITEMS"]
            descriptions = [
                str(cost.cell(row=row, column=5).value or "")
                for row in range(2, cost.max_row + 1)
            ]
            assert all(not value.startswith("=") for value in descriptions)
            assert "HONORARIOS PROYECTO" not in descriptions
            assert "Deducciones" not in descriptions
        finally:
            wb.close()
    finally:
        _cleanup(root)
