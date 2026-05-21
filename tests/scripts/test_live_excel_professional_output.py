from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import generate_preview_from_real_xlsx, validate_workbook_file


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"live_master_prof_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_multi_semantic_budget(path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Datos"
    ws1.append(["Codigo", "Descripcion", "Formula", "Importe"])
    ws1.append(["LUV_AP", "EQUIPAMIENTO", "=D2/PEM", 37297.09])
    ws1.append(["LUV_FA", "GRIFERIA", "=D3/PEM", 125609.65])
    ws1.append([None, "REVISION abril", "", 2026])

    ws2 = wb.create_sheet("Espacios")
    ws2.append(["Código", "Info", "Resumen", "Pres"])
    ws2.append(["BALCON", "", "BALCON", 1000])
    ws2.append(["COCINA", "", "COCINA", 2000])

    wb.save(path)
    wb.close()


def _build_comparison_budget(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(["Cap.", "Nombre del capítulo", "Importe (€)", "Nombre equivalente", "Importe equivalente", "Diferencia"])
    ws.append([2, "Demoliciones", 687.5, "DEMOLICIONES", 550, "=E2-C2"])
    wb.save(path)
    wb.close()


def _detail_sheets(wb: object, home: str) -> list[str]:
    return sorted(name for name in wb.sheetnames if name.startswith(home + "_"))


def test_professional_output_creates_adaptive_views_without_mixing():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_multi_semantic_budget(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_prof_test_001")

        wb = load_workbook(out)
        try:
            assert wb.sheetnames[0] == "INDEX"
            assert "BUDGET_REVIEW_001" in wb.sheetnames
            assert "BUDGET_REVIEW_TRACE_001" in wb.sheetnames
            details = _detail_sheets(wb, "BUDGET_REVIEW_001")
            assert len(details) == 2

            home = wb["BUDGET_REVIEW_001"]
            assert home["A4"].value == "Hoja origen"
            home_views = {
                str(home.cell(row=row, column=1).value or ""): str(home.cell(row=row, column=4).value or "")
                for row in range(5, home.max_row + 1)
                if str(home.cell(row=row, column=1).value or "").strip()
            }
            assert "Datos" in home_views
            assert "Espacios" in home_views
            assert home_views["Datos"] != home_views["Espacios"]

            datos_ws = wb[home_views["Datos"]]
            espacios_ws = wb[home_views["Espacios"]]
            datos_headers = [str(datos_ws.cell(row=4, column=idx).value or "") for idx in range(1, 8)]
            espacios_headers = [str(espacios_ws.cell(row=4, column=idx).value or "") for idx in range(1, 6)]
            assert datos_headers[:4] == ["Codigo", "Descripcion", "Formula / Ratio", "Importe"]
            assert espacios_headers[:4] == ["Codigo", "Info", "Resumen", "Presupuesto"]

            # No metadata row treated as budget item amount.
            datos_values = [
                str(datos_ws.cell(row=row, column=2).value or "")
                for row in range(5, datos_ws.max_row + 1)
            ]
            assert all("REVISION" not in value.upper() for value in datos_values)

            # Hidden technical id column exists but is hidden.
            assert bool(datos_ws.column_dimensions["E"].hidden)
            assert bool(espacios_ws.column_dimensions["E"].hidden)
        finally:
            wb.close()
        validate_workbook_file(out)
    finally:
        _cleanup(root)


def test_comparison_sheet_keeps_comparison_columns_without_classic_projection():
    root = _make_root()
    try:
        src = root / "comparison.xlsx"
        _build_comparison_budget(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_prof_test_002")

        wb = load_workbook(out)
        try:
            home = wb["BUDGET_REVIEW_001"]
            row = next(
                row_idx
                for row_idx in range(5, home.max_row + 1)
                if str(home.cell(row=row_idx, column=1).value or "").strip() == "Hoja1"
            )
            view_name = str(home.cell(row=row, column=4).value or "").strip()
            view_ws = wb[view_name]
            headers = [str(view_ws.cell(row=4, column=idx).value or "") for idx in range(1, 8)]
            assert headers[:6] == [
                "Cap.",
                "Nombre del capítulo",
                "Importe (€)",
                "Nombre equivalente",
                "Importe equivalente",
                "Diferencia",
            ]
            assert "Cantidad" not in headers
            assert "Precio unitario" not in headers
            assert "Ud" not in headers
            assert str(view_ws.cell(row=5, column=1).value) == "2"
            assert str(view_ws.cell(row=5, column=5).value) == "550"
        finally:
            wb.close()
    finally:
        _cleanup(root)
