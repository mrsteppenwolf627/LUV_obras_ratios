from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook

from scripts.generate_live_excel_master import (
    OPERATIONAL_PREVIEW_COLUMNS,
    OPERATIONAL_PREVIEW_SHEET,
    generate_preview_from_real_xlsx,
    validate_workbook_file,
)


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"live_master_preview_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def _build_synthetic_input(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    ws.append(["Capitulo Codigo", "Capitulo", "Codigo", "Descripcion", "Unidad", "Cantidad", "Precio Unitario", "Importe"])
    ws.append(["2", "Demoliciones", "2.1.01", "Demolicion de tabiqueria", "m2", 10, 68.75, 687.50])
    wb.save(path)
    wb.close()


def test_operational_preview_sheet_created_with_minimum_columns():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_synthetic_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_preview_test_001")
        wb = load_workbook(out)
        try:
            assert OPERATIONAL_PREVIEW_SHEET in wb.sheetnames
            ws = wb[OPERATIONAL_PREVIEW_SHEET]
            actual = [ws.cell(row=1, column=i).value for i in range(1, len(OPERATIONAL_PREVIEW_COLUMNS) + 1)]
            assert actual == OPERATIONAL_PREVIEW_COLUMNS
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_operational_preview_separates_amount_and_keeps_traceability():
    root = _make_root()
    try:
        src = root / "input.xlsx"
        _build_synthetic_input(src)
        out = root / "outputs" / "live_excel_master" / "preview" / "preview.xlsx"
        generate_preview_from_real_xlsx(src, out, source_file_id="sf_preview_test_002")

        wb = load_workbook(out)
        try:
            ws = wb[OPERATIONAL_PREVIEW_SHEET]
            row = [ws.cell(row=2, column=i).value for i in range(1, len(OPERATIONAL_PREVIEW_COLUMNS) + 1)]
            assert row[6] == "2"  # chapter_code
            assert row[7] == "Demoliciones"  # chapter_name
            assert row[8] == "2.1.01"  # item_code
            assert row[9] == "Demolicion de tabiqueria"  # item_description
            assert str(row[13]).startswith("687.5")  # amount
            assert "687.5" not in str(row[9])  # no amount in description
            assert row[4] == "Presupuesto"  # source_sheet_name
            assert str(row[5]) == "2"  # source_row_number
            assert row[16] == "TRUE"  # preview_only
            assert wb["RATIO_INPUTS"].max_row == 1
            assert wb["RATIOS_CALCULATED"].max_row == 1
        finally:
            wb.close()
        validate_workbook_file(out)
    finally:
        _cleanup(root)

