from pathlib import Path
import json
import shutil
import uuid

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo

from scripts.read_excel_full import analyze_excel_sources, write_reports


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"read_excel_full_{uuid.uuid4().hex}"
    (root / "data" / "samples").mkdir(parents=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def test_full_reader_profiles_supported_workbook_and_preserves_traceability():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        ws.append(["Codigo", "Descripcion", "Unidad", "Cantidad", "Precio", "Importe"])
        ws.append([
            "IT-001",
            "Descripcion muy larga para comprobar la sanitizacion de salida del lector integral Excel y forzar el truncado seguro de muestras "
            "sin perder la trazabilidad de la celda de origen en la salida generada",
            "m2",
            3,
            10.5,
            "=D2*E2",
        ])
        ws["B2"].comment = Comment("Comentario sensible de prueba", "Codex")
        tab = Table(displayName="TablaDatos", ref="A1:F2")
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        visual = wb.create_sheet("Visual")
        visual["A1"] = "Seccion"
        visual.merge_cells("A1:C1")
        visual["A20"] = "Nota"
        visual.row_dimensions[2].hidden = True
        visual.column_dimensions["B"].hidden = True

        empty = wb.create_sheet("Vacia")

        hidden = wb.create_sheet("Oculta")
        hidden["A1"] = "Texto oculto"
        hidden.sheet_state = "hidden"

        chart = BarChart()
        data = Reference(ws, min_col=4, min_row=1, max_row=2)
        chart.add_data(data, titles_from_data=True)
        ch = wb.create_chartsheet(title="Grafica")
        ch.add_chart(chart)

        file_path = root / "data" / "samples" / "integral.xlsx"
        wb.save(file_path)

        payload = analyze_excel_sources(root)
        workbook = next(item for item in payload["workbook_summaries"] if item["relative_path_sanitized"] == "data/samples/integral.xlsx")
        sheets = {item["sheet_name_sanitized"]: item for item in payload["sheets"] if item["workbook_ref"] == "data/samples/integral.xlsx"}

        assert payload["global_summary"]["excel_files_detected"] == 1
        assert workbook["readable"] is True
        assert workbook["sheet_count"] == 5
        assert workbook["worksheet_count"] == 4
        assert workbook["chartsheet_count"] == 1

        datos = sheets["Datos"]
        assert datos["candidate_header_rows"][0] == 1
        assert datos["budget_signals"]["signals_by_field"]["importe"]["count"] >= 1
        assert datos["candidate_table_blocks"]
        assert datos["formulas_summary"]["count"] == 1
        assert datos["comments_summary"]["count"] == 1
        assert datos["traceability_map"]
        assert any(item["coordinate"] == "B2" and "COMMENT" in item["flags"] for item in datos["traceability_map"])
        assert any(item["coordinate"] == "B2" and item["sanitized_value"].endswith("...") for item in datos["traceability_map"])
        assert datos["cell_samples_sanitized"]["first_non_empty_rows"]

        visual = sheets["Visual"]
        assert visual["merged_cells_summary"]["count"] == 1
        assert visual["visibility"]["sheet_state"] == "visible"
        assert any(block["block_type"] == "merged_cells" for block in visual["visual_blocks"])
        assert any(block["block_type"] == "hidden_structure" for block in visual["visual_blocks"])

        assert sheets["Vacia"]["is_empty_sheet"] is True
        assert sheets["Vacia"]["manual_review"]
        assert sheets["Oculta"]["visibility"]["sheet_state"] == "hidden"

        chart_sheet = sheets["Grafica"]
        assert chart_sheet["sheet_type"] == "CHARTSHEET"
        assert chart_sheet["visual_blocks"][0]["block_type"] == "chartsheet_context"

        assert "CHARTSHEET_PRESENT" in payload["warnings"]
        assert any(item.startswith("CHARTSHEET_PRESENT") for item in payload["manual_review"])
    finally:
        _cleanup(root)


def test_controlled_exclusions_capture_non_excel_and_legacy_excel():
    root = _make_root()
    try:
        (root / "data" / "samples" / "legacy.xls").write_bytes(b"legacy")
        (root / "data" / "samples" / "legacy.xlsb").write_bytes(b"legacy")
        (root / "data" / "samples" / "note.txt").write_text("not excel", encoding="utf-8")

        payload = analyze_excel_sources(root)

        reasons = {item["reason"] for item in payload["controlled_exclusions"]}
        paths = {item["relative_path_sanitized"] for item in payload["controlled_exclusions"]}

        assert payload["global_summary"]["source_files_total"] == 3
        assert payload["global_summary"]["excel_files_detected"] == 0
        assert payload["global_summary"]["controlled_exclusions_total"] == 3
        assert "UNSUPPORTED_EXCEL_FORMAT" in reasons
        assert "NON_EXCEL_IGNORED" in reasons
        assert "data/samples/legacy.xls" in paths
        assert "data/samples/legacy.xlsb" in paths
        assert "data/samples/note.txt" in paths
        assert len(payload["workbook_summaries"]) == 2
        assert all(not item["readable"] for item in payload["workbook_summaries"])
    finally:
        _cleanup(root)


def test_json_and_markdown_reports_generated_with_contract_keys():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Base"
        ws.append(["Codigo", "Descripcion"])
        ws.append(["IT-01", "Desc"])
        wb.save(root / "data" / "samples" / "simple.xlsx")

        payload = analyze_excel_sources(root)
        json_path, md_path = write_reports(root, payload)

        saved = json.loads(json_path.read_text(encoding="utf-8"))
        md_text = md_path.read_text(encoding="utf-8")

        assert json_path.exists()
        assert md_path.exists()
        assert "reader_metadata" in saved
        assert "workbook_summaries" in saved
        assert "sheets" in saved
        assert "global_summary" in saved
        assert "controlled_exclusions" in saved
        assert "ratios" not in saved
        assert "master" not in saved
        assert "CATEGORY_MAPPING" not in md_text
        assert "Excel Full Reader" in md_text
    finally:
        _cleanup(root)


def test_full_reader_does_not_modify_input_file():
    root = _make_root()
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["Codigo", "Descripcion"])
        ws.append(["IT-01", "Desc"])
        file_path = root / "data" / "samples" / "stable.xlsx"
        wb.save(file_path)
        before = file_path.read_bytes()

        _ = analyze_excel_sources(root)
        after = file_path.read_bytes()

        assert before == after
    finally:
        _cleanup(root)
