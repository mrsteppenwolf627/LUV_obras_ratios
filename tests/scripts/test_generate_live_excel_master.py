from pathlib import Path
import shutil
import uuid

import pytest
from openpyxl import load_workbook

from scripts.generate_live_excel_master import (
    REQUIRED_SHEETS_COLUMNS,
    SchemaValidationError,
    generate_master,
    validate_workbook_file,
)


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"live_master_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def test_create_empty_master_template():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        result = generate_master(output, update=False)
        assert output.exists()
        assert result["updated"] == "false"
        assert result["pre_snapshot"] == ""
        assert result["post_snapshot"] == ""
    finally:
        _cleanup(root)


def test_required_sheets_exist():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output, update=False)
        wb = load_workbook(output)
        try:
            for sheet_name in REQUIRED_SHEETS_COLUMNS:
                assert sheet_name in wb.sheetnames
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_required_columns_exist():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output, update=False)
        wb = load_workbook(output)
        try:
            for sheet_name, expected_columns in REQUIRED_SHEETS_COLUMNS.items():
                ws = wb[sheet_name]
                actual = [ws.cell(row=1, column=i).value for i in range(1, len(expected_columns) + 1)]
                assert actual == expected_columns
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_schema_validation_ok():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output, update=False)
        validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_schema_validation_fails_when_sheet_missing():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output, update=False)
        wb = load_workbook(output)
        wb.remove(wb["CHANGELOG"])
        wb.save(output)
        wb.close()

        with pytest.raises(SchemaValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_schema_validation_fails_when_column_missing():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output, update=False)
        wb = load_workbook(output)
        ws = wb["IMPORT_LOG"]
        ws.cell(row=1, column=1, value="broken_column")
        wb.save(output)
        wb.close()

        with pytest.raises(SchemaValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_update_creates_pre_and_post_snapshots():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output, update=False)
        result = generate_master(output, update=True)

        assert result["pre_snapshot"]
        assert result["post_snapshot"]
        assert Path(result["pre_snapshot"]).exists()
        assert Path(result["post_snapshot"]).exists()

        wb = load_workbook(output)
        try:
            snapshots_ws = wb["SNAPSHOTS"]
            changelog_ws = wb["CHANGELOG"]
            assert snapshots_ws.max_row >= 3
            assert changelog_ws.max_row >= 2
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_existing_without_update_fails():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output, update=False)
        with pytest.raises(RuntimeError):
            generate_master(output, update=False)
    finally:
        _cleanup(root)


def test_only_synthetic_seed_data():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output, update=False)
        wb = load_workbook(output)
        try:
            assert wb["README_MASTER"].max_row >= 2
            assert wb["IMPORT_LOG"].max_row == 1
            assert wb["SOURCE_FILES"].max_row == 1
            assert wb["RATIOS_CALCULATED"].max_row == 1
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_repeatable_update_flow():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output, update=False)
        generate_master(output, update=True)
        generate_master(output, update=True)
        validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_fails_outside_allowed_output_path():
    root = _make_root()
    try:
        output = root / "unsafe" / "master.xlsx"
        with pytest.raises(ValueError):
            generate_master(output, update=False)
    finally:
        _cleanup(root)
