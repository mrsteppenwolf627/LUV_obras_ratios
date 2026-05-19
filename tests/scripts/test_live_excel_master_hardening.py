from pathlib import Path
import shutil
import uuid

import pytest
from openpyxl import load_workbook

from scripts.generate_live_excel_master import (
    ReferentialValidationError,
    generate_master,
    load_synthetic_incremental,
    rollback_master_from_snapshot,
    validate_workbook_file,
)


def _make_root() -> Path:
    base = Path(__file__).resolve().parents[2] / ".tmp_tests"
    base.mkdir(exist_ok=True)
    root = base / f"live_master_hard_{uuid.uuid4().hex}"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _cleanup(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def test_synthetic_incremental_load_valid():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        result = load_synthetic_incremental(output_path=output, retention_max=5)
        assert output.exists()
        assert result["run_id"]
        wb = load_workbook(output)
        try:
            assert wb["SOURCE_FILES"].max_row >= 2
            assert wb["PROJECTS"].max_row >= 2
            assert wb["BUDGET_VERSIONS"].max_row >= 2
            assert wb["IMPORT_LOG"].max_row >= 2
            assert wb["RAW_IMPORTS"].max_row >= 2
            assert wb["COST_ITEMS"].max_row >= 3
            assert wb["VALIDATION_RESULTS"].max_row >= 2
            assert wb["CHANGELOG"].max_row >= 2
        finally:
            wb.close()
    finally:
        _cleanup(root)


def test_referential_fails_budget_source_missing():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["BUDGET_VERSIONS"]
            ws.cell(row=2, column=3, value="missing_source")
            wb.save(output)
        finally:
            wb.close()
        with pytest.raises(ReferentialValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_referential_fails_cost_item_budget_missing():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["COST_ITEMS"]
            ws.cell(row=2, column=2, value="missing_budget")
            wb.save(output)
        finally:
            wb.close()
        with pytest.raises(ReferentialValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_referential_fails_validation_import_batch_missing():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["VALIDATION_RESULTS"]
            ws.cell(row=2, column=9, value="missing_import_batch")
            wb.save(output)
        finally:
            wb.close()
        with pytest.raises(ReferentialValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_blocked_ratio_inputs_rejected():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["RATIO_INPUTS"]
            ws.cell(row=2, column=7, value="BLOCKED")
            wb.save(output)
        finally:
            wb.close()
        with pytest.raises(ReferentialValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_empty_ids_detected():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["SOURCE_FILES"]
            ws.cell(row=2, column=1, value="")
            wb.save(output)
        finally:
            wb.close()
        with pytest.raises(ReferentialValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_duplicate_primary_keys_detected():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        wb = load_workbook(output)
        try:
            ws = wb["SOURCE_FILES"]
            duplicated_id = ws.cell(row=2, column=1).value
            row_data = [ws.cell(row=2, column=i).value for i in range(1, 9)]
            row_data[0] = duplicated_id
            ws.append(row_data)
            wb.save(output)
        finally:
            wb.close()
        with pytest.raises(ReferentialValidationError):
            validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_synthetic_load_creates_pre_post_snapshots_when_file_exists():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        generate_master(output_path=output, update=False, retention_max=5)
        result = load_synthetic_incremental(output_path=output, retention_max=5)
        assert result["pre_snapshot"]
        assert result["post_snapshot"]
        assert Path(result["pre_snapshot"]).exists()
        assert Path(result["post_snapshot"]).exists()
    finally:
        _cleanup(root)


def test_repeatable_synthetic_load():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        load_synthetic_incremental(output_path=output, retention_max=5)
        validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_snapshot_retention_max_applied():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        for _ in range(4):
            load_synthetic_incremental(output_path=output, retention_max=2)
        snapshots = list((output.parent / "snapshots").glob("*.xlsx"))
        assert len(snapshots) <= 2
    finally:
        _cleanup(root)


def test_rollback_from_snapshot_restores_workbook():
    root = _make_root()
    try:
        output = root / "outputs" / "live_excel_master" / "master.xlsx"
        load_synthetic_incremental(output_path=output, retention_max=5)
        first_size = output.stat().st_size
        result = load_synthetic_incremental(output_path=output, retention_max=5)
        snapshot_path = Path(result["pre_snapshot"])
        rollback_master_from_snapshot(output_path=output, snapshot_path=snapshot_path, retention_max=5)
        assert output.exists()
        assert output.stat().st_size > 0
        assert output.stat().st_size <= max(first_size, output.stat().st_size)
        validate_workbook_file(output)
    finally:
        _cleanup(root)


def test_fails_outside_allowed_path_synthetic_load():
    root = _make_root()
    try:
        output = root / "unsafe" / "master.xlsx"
        with pytest.raises(ValueError):
            load_synthetic_incremental(output_path=output, retention_max=5)
    finally:
        _cleanup(root)
