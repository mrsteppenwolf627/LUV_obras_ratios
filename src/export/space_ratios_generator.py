"""Excel generator for space-level ratios: AREAS + RATIOS_ESTANCIA sheets (Hito 3).

Living-document design:
  AREAS       — summary matrix (plant × space) driven by SUMIFS formulas that
                reference a raw-data section. Fill in m² there; summary updates.
  RATIOS_ESTANCIA — one row per space, cross-references AREAS via sheet formulas.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Layout constants
# ---------------------------------------------------------------------------

# Fixed columns in AREAS sheet (1-indexed)
_COL_PLANTA = 1    # A — plant label in summary / plant label in raw data
_COL_ZONA = 2      # B — total m² in summary / zone in raw data
_COL_M2 = 3        # C — (blank in summary) / m² in raw data  ← SUMIFS sum_range
_COL_ESPACIO = 4   # D — (blank in summary) / space name in raw data ← SUMIFS lookup
_COL_COST_RAW = 5  # E — (blank in summary) / coste in raw data

_SPACE_COL_START = 6  # F — first space column in summary section

# Fixed rows in AREAS summary section
_ROW_TITLE = 1
_ROW_ZONES = 2
_ROW_SPACES = 3
_ROW_BLANK1 = 4
_ROW_PS = 5
_ROW_PB = 6
_ROW_PP = 7
_ROW_TOTAL = 8
_ROW_PCT = 9
_ROW_BLANK2 = 10
_ROW_COSTE_LABEL = 11
_ROW_RATIO_LABEL = 12
_ROW_RAW_SECTION = 13   # section header "DATOS BRUTOS"
_ROW_RAW_COLS = 14      # column labels for raw data
_ROW_RAW_DATA = 15      # first actual data row

_SUMIFS_RANGE_END = 200  # generous range for SUMIFS (supports expansion)

# Space catalogue ordered by zone
_ZONES_AND_SPACES: list[tuple[str, list[str]]] = [
    ("NOBLE", [
        "SALA", "COMEDOR", "HABITACION MASTER", "HABITACIONES SECUNDARIAS",
        "BAÑO MASTER", "BAÑOS SECUNDARIOS", "ASEO", "AMENITIES",
    ]),
    ("SERVICIO", [
        "COCINA", "COCINA SERVICIO", "HABITACIONES DE SERVICIO", "ZONAS DE SERVICIOS",
    ]),
    ("EXTERIORES", [
        "BALCONES", "TERRAZAS", "PISCINA", "JARDIN", "COMUNES FACHADA",
    ]),
    ("COMUNES", [
        "PASILLOS", "COMUNES ARQUITECTURA", "INSTALACIONES",
    ]),
]

# Flat ordered list: (zone, space_name)
_FLAT_SPACES: list[tuple[str, str]] = [
    (zone, space)
    for zone, spaces in _ZONES_AND_SPACES
    for space in spaces
]

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

_BLUE_DARK = PatternFill("solid", fgColor="1F3864")
_BLUE_MED = PatternFill("solid", fgColor="2F75B6")
_BLUE_LIGHT = PatternFill("solid", fgColor="D9E1F2")
_GREEN_DARK = PatternFill("solid", fgColor="375623")
_GREEN_LIGHT = PatternFill("solid", fgColor="E2EFDA")
_ORANGE = PatternFill("solid", fgColor="F4B942")
_YELLOW = PatternFill("solid", fgColor="FFFF00")
_GREY = PatternFill("solid", fgColor="D9D9D9")

_FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF", size=10)
_FONT_BOLD = Font(bold=True, size=10)
_FONT_NORMAL = Font(size=10)
_FONT_ITALIC = Font(italic=True, size=9, color="595959")

_THIN = Side(border_style="thin", color="BFBFBF")
_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")


def _style(
    cell: Any,
    fill: Any = None,
    font: Any = None,
    alignment: Any = None,
    border: Any = None,
    number_format: str = "",
) -> None:
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_space_ratios_excel(
    ratios: dict,
    presupuesto: dict,
    output_path: str | Path,
) -> str:
    """
    Generate Excel with AREAS + RATIOS_ESTANCIA sheets.

    Args:
        ratios: output of calculate_space_ratios()
        presupuesto: output of parse_presto() — used for filename / budget_code
        output_path: where to save the .xlsx

    Returns:
        Absolute path of the saved file.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    ws_areas = wb.create_sheet("AREAS")
    _build_areas_sheet(ws_areas, ratios, presupuesto)

    ws_ratios = wb.create_sheet("RATIOS_ESTANCIA")
    _build_ratios_estancia_sheet(ws_ratios, ratios, presupuesto)

    wb.active = ws_areas
    wb.save(str(out))
    return str(out.resolve())


# ---------------------------------------------------------------------------
# AREAS sheet builder
# ---------------------------------------------------------------------------

def _build_areas_sheet(ws: Any, ratios: dict, presupuesto: dict) -> None:
    filename = presupuesto.get("filename", "")
    budget_code = presupuesto.get("budget_code", "")
    total_coste = ratios.get("total_coste", 0.0)

    # Build lookup: space_name → coste (from Presto)
    coste_by_space: dict[str, float] = {}
    for spc in ratios.get("espacios", []):
        coste_by_space[spc["nombre"]] = spc["total"]["coste"]

    n_spaces = len(_FLAT_SPACES)
    last_space_col = _SPACE_COL_START + n_spaces - 1

    # Row heights
    ws.row_dimensions[_ROW_TITLE].height = 20
    ws.row_dimensions[_ROW_ZONES].height = 18
    ws.row_dimensions[_ROW_SPACES].height = 40
    ws.row_dimensions[_ROW_RAW_SECTION].height = 16

    # Column widths
    ws.column_dimensions[get_column_letter(_COL_PLANTA)].width = 12
    ws.column_dimensions[get_column_letter(_COL_ZONA)].width = 14
    ws.column_dimensions[get_column_letter(_COL_M2)].width = 10
    ws.column_dimensions[get_column_letter(_COL_ESPACIO)].width = 28
    ws.column_dimensions[get_column_letter(_COL_COST_RAW)].width = 14
    for col_idx in range(_SPACE_COL_START, last_space_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # --- Row 1: Title ---
    c = ws.cell(_ROW_TITLE, _COL_PLANTA,
                value=f"ÁREAS — {filename}  |  Código: {budget_code}")
    _style(c, fill=_BLUE_DARK, font=_FONT_WHITE_BOLD, alignment=_LEFT)
    ws.merge_cells(
        start_row=_ROW_TITLE, start_column=_COL_PLANTA,
        end_row=_ROW_TITLE, end_column=last_space_col,
    )

    # --- Row 2: Zone group headers (merged) ---
    col_cursor = _SPACE_COL_START
    for zone, spaces in _ZONES_AND_SPACES:
        n = len(spaces)
        c = ws.cell(_ROW_ZONES, col_cursor, value=f"ZONA {zone}")
        _style(c, fill=_BLUE_MED, font=_FONT_WHITE_BOLD, alignment=_CENTER)
        if n > 1:
            ws.merge_cells(
                start_row=_ROW_ZONES, start_column=col_cursor,
                end_row=_ROW_ZONES, end_column=col_cursor + n - 1,
            )
        col_cursor += n

    # Fixed column headers (row 2, cols A-E)
    for col_idx, label in [
        (_COL_PLANTA, "Planta"),
        (_COL_ZONA, "Total m²"),
    ]:
        c = ws.cell(_ROW_ZONES, col_idx, value=label)
        _style(c, fill=_GREY, font=_FONT_BOLD, alignment=_CENTER)

    # --- Row 3: Space name headers ---
    c = ws.cell(_ROW_SPACES, _COL_PLANTA, value="Planta")
    _style(c, fill=_GREY, font=_FONT_BOLD, alignment=_CENTER)
    c = ws.cell(_ROW_SPACES, _COL_ZONA, value="Total m²")
    _style(c, fill=_GREY, font=_FONT_BOLD, alignment=_CENTER)

    for idx, (zone, space_name) in enumerate(_FLAT_SPACES):
        col = _SPACE_COL_START + idx
        c = ws.cell(_ROW_SPACES, col, value=space_name)
        _style(c, fill=_BLUE_LIGHT, font=_FONT_BOLD, alignment=_CENTER, border=_BORDER_THIN)

    # --- Rows 5-9: Summary rows (PS / PB / PP / TOTAL / %) ---
    plant_rows = [
        (_ROW_PS, "PS"),
        (_ROW_PB, "PB"),
        (_ROW_PP, "PP"),
    ]

    for row, plant in plant_rows:
        c = ws.cell(row, _COL_PLANTA, value=plant)
        _style(c, fill=_BLUE_LIGHT, font=_FONT_BOLD, alignment=_CENTER)

        # Total m² for this plant = SUM of all space SUMIFS
        space_cols = [get_column_letter(_SPACE_COL_START + i) for i in range(n_spaces)]
        total_formula = "=" + "+".join(f"{cl}{row}" for cl in space_cols)
        c = ws.cell(row, _COL_ZONA, value=total_formula)
        _style(c, font=_FONT_NORMAL, alignment=_CENTER,
               number_format='#,##0.00" m²"')

        for idx in range(n_spaces):
            col = _SPACE_COL_START + idx
            col_letter = get_column_letter(col)
            formula = (
                f"=SUMIFS($C${_ROW_RAW_DATA}:$C${_SUMIFS_RANGE_END},"
                f"$D${_ROW_RAW_DATA}:$D${_SUMIFS_RANGE_END},{col_letter}${_ROW_SPACES},"
                f"$A${_ROW_RAW_DATA}:$A${_SUMIFS_RANGE_END},$A{row})"
            )
            c = ws.cell(row, col, value=formula)
            _style(c, fill=_GREEN_LIGHT, font=_FONT_NORMAL, alignment=_CENTER,
                   border=_BORDER_THIN, number_format='#,##0.00" m²"')

    # Row 8: TOTAL m²
    c = ws.cell(_ROW_TOTAL, _COL_PLANTA, value="TOTAL")
    _style(c, fill=_GREEN_DARK, font=Font(bold=True, color="FFFFFF", size=10), alignment=_CENTER)

    total_m2_col_b = (
        f"={get_column_letter(_COL_ZONA)}{_ROW_PS}"
        f"+{get_column_letter(_COL_ZONA)}{_ROW_PB}"
        f"+{get_column_letter(_COL_ZONA)}{_ROW_PP}"
    )
    c = ws.cell(_ROW_TOTAL, _COL_ZONA, value=total_m2_col_b)
    _style(c, fill=_GREEN_DARK, font=Font(bold=True, color="FFFFFF", size=10),
           alignment=_CENTER, number_format='#,##0.00" m²"')

    for idx in range(n_spaces):
        col = _SPACE_COL_START + idx
        cl = get_column_letter(col)
        formula = f"={cl}{_ROW_PS}+{cl}{_ROW_PB}+{cl}{_ROW_PP}"
        c = ws.cell(_ROW_TOTAL, col, value=formula)
        _style(c, fill=_GREEN_DARK, font=Font(bold=True, color="FFFFFF", size=10),
               alignment=_CENTER, border=_BORDER_THIN, number_format='#,##0.00" m²"')

    # Row 9: % of total
    c = ws.cell(_ROW_PCT, _COL_PLANTA, value="% m²")
    _style(c, fill=_GREY, font=_FONT_BOLD, alignment=_CENTER)

    b_total_ref = f"${get_column_letter(_COL_ZONA)}${_ROW_TOTAL}"
    for idx in range(n_spaces):
        col = _SPACE_COL_START + idx
        cl = get_column_letter(col)
        formula = f"=IF({b_total_ref}>0,{cl}{_ROW_TOTAL}/{b_total_ref}*100,0)"
        c = ws.cell(_ROW_PCT, col, value=formula)
        _style(c, fill=_GREY, font=_FONT_NORMAL, alignment=_CENTER,
               border=_BORDER_THIN, number_format='0.00"%"')

    # Row 11: Coste (from Presto — static values)
    c = ws.cell(_ROW_COSTE_LABEL, _COL_PLANTA, value="Coste Presto (€)")
    _style(c, fill=_ORANGE, font=_FONT_BOLD, alignment=_CENTER)
    c = ws.cell(_ROW_COSTE_LABEL, _COL_ZONA, value=total_coste)
    _style(c, fill=_ORANGE, font=_FONT_BOLD, alignment=_CENTER,
           number_format='#,##0.00 "€"')

    for idx, (zone, space_name) in enumerate(_FLAT_SPACES):
        col = _SPACE_COL_START + idx
        coste = coste_by_space.get(space_name, 0.0)
        c = ws.cell(_ROW_COSTE_LABEL, col, value=coste if coste else "")
        _style(c, fill=_YELLOW, font=_FONT_NORMAL, alignment=_CENTER,
               border=_BORDER_THIN, number_format='#,##0.00 "€"')

    # Row 12: Ratio €/m²
    c = ws.cell(_ROW_RATIO_LABEL, _COL_PLANTA, value='Ratio €/m²')
    _style(c, fill=_ORANGE, font=_FONT_BOLD, alignment=_CENTER)

    for idx in range(n_spaces):
        col = _SPACE_COL_START + idx
        cl = get_column_letter(col)
        formula = f'=IF({cl}{_ROW_TOTAL}>0,{cl}{_ROW_COSTE_LABEL}/{cl}{_ROW_TOTAL},"")'
        c = ws.cell(_ROW_RATIO_LABEL, col, value=formula)
        _style(c, fill=_YELLOW, font=_FONT_NORMAL, alignment=_CENTER,
               border=_BORDER_THIN, number_format='#,##0.00 "€/m²"')

    # --- Row 13: Raw data section header ---
    c = ws.cell(_ROW_RAW_SECTION, _COL_PLANTA,
                value="DATOS BRUTOS — Rellena la columna m² por espacio y planta")
    _style(c, fill=_BLUE_DARK, font=_FONT_WHITE_BOLD, alignment=_LEFT)
    ws.merge_cells(
        start_row=_ROW_RAW_SECTION, start_column=_COL_PLANTA,
        end_row=_ROW_RAW_SECTION, end_column=last_space_col,
    )

    # --- Row 14: Raw data column headers ---
    for col_idx, label in [
        (_COL_PLANTA, "Planta"),
        (_COL_ZONA, "Zona"),
        (_COL_M2, "m²"),
        (_COL_ESPACIO, "Espacio"),
        (_COL_COST_RAW, "Coste Presto (€)"),
    ]:
        c = ws.cell(_ROW_RAW_COLS, col_idx, value=label)
        _style(c, fill=_BLUE_MED, font=_FONT_WHITE_BOLD, alignment=_CENTER)

    # --- Rows 15+: Pre-populate raw data (3 plants × N spaces) ---
    row = _ROW_RAW_DATA
    for plant in ("PS", "PB", "PP"):
        for zone, space_name in _FLAT_SPACES:
            coste = coste_by_space.get(space_name, 0.0)
            fill = _BLUE_LIGHT if row % 2 == 0 else None
            for col_idx, val in [
                (_COL_PLANTA, plant),
                (_COL_ZONA, zone),
                (_COL_M2, 0.0),       # user fills in m²
                (_COL_ESPACIO, space_name),
                (_COL_COST_RAW, coste if coste else ""),
            ]:
                c = ws.cell(row, col_idx, value=val)
                _style(c, fill=fill, font=_FONT_NORMAL, alignment=_CENTER)
                if col_idx == _COL_M2:
                    c.number_format = '#,##0.00'
                if col_idx == _COL_COST_RAW:
                    c.number_format = '#,##0.00 "€"'
            row += 1

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# RATIOS_ESTANCIA sheet builder
# ---------------------------------------------------------------------------

def _build_ratios_estancia_sheet(ws: Any, ratios: dict, presupuesto: dict) -> None:
    filename = presupuesto.get("filename", "")
    budget_code = presupuesto.get("budget_code", "")

    # Space → column index in AREAS sheet
    space_col_in_areas: dict[str, int] = {
        space: _SPACE_COL_START + idx
        for idx, (_, space) in enumerate(_FLAT_SPACES)
    }

    # Build lookup: name → total from ratios dict
    ratio_by_space: dict[str, dict] = {
        spc["nombre"]: spc for spc in ratios.get("espacios", [])
    }

    # Row 1: Title
    c = ws.cell(1, 1, value=f"RATIOS ESTANCIA — {filename}  |  Código: {budget_code}")
    _style(c, fill=_BLUE_DARK, font=_FONT_WHITE_BOLD, alignment=_LEFT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    # Row 2: note
    c = ws.cell(2, 1, value="Los m² se calculan automáticamente desde la hoja AREAS. "
                "Completa la columna m² en AREAS para ver los ratios actualizados.")
    _style(c, font=_FONT_ITALIC, alignment=_LEFT)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    # Row 4: Column headers
    headers = [
        "Espacio", "Zona",
        "m² PS", "m² PB", "m² PP", "TOTAL m²", "% m²",
        "Presupuesto (€)", "Pres. Prorrateado (€)", "Ratio €/m²",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(4, col, value=h)
        _style(c, fill=_BLUE_MED, font=_FONT_WHITE_BOLD, alignment=_CENTER)
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 2)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14

    # Rows 5+: one row per space (reference AREAS sheet for m² values)
    b_total_ref_areas = (
        f"=AREAS!{get_column_letter(_COL_ZONA)}{_ROW_TOTAL}"
    )
    total_coste = ratios.get("total_coste", 0.0)

    data_row = 5
    for idx, (zone, space_name) in enumerate(_FLAT_SPACES):
        if space_name not in space_col_in_areas:
            continue
        areas_col = get_column_letter(space_col_in_areas[space_name])
        spc_data = ratio_by_space.get(space_name, {})
        coste = spc_data.get("total", {}).get("coste", 0.0)

        fill = _BLUE_LIGHT if data_row % 2 == 0 else None

        # Col 1: Espacio
        c = ws.cell(data_row, 1, value=space_name)
        _style(c, fill=fill, font=_FONT_BOLD, alignment=_LEFT)

        # Col 2: Zona
        c = ws.cell(data_row, 2, value=zone)
        _style(c, fill=fill, font=_FONT_NORMAL, alignment=_CENTER)

        # Cols 3-5: m² per plant (cross-reference to AREAS)
        for col_offset, plant_row in [(3, _ROW_PS), (4, _ROW_PB), (5, _ROW_PP)]:
            formula = f"=AREAS!{areas_col}{plant_row}"
            c = ws.cell(data_row, col_offset, value=formula)
            _style(c, fill=fill, font=_FONT_NORMAL, alignment=_CENTER,
                   number_format='#,##0.00" m²"')

        # Col 6: TOTAL m²
        formula = f"=AREAS!{areas_col}{_ROW_TOTAL}"
        c = ws.cell(data_row, 6, value=formula)
        _style(c, fill=fill, font=_FONT_BOLD, alignment=_CENTER,
               number_format='#,##0.00" m²"')

        # Col 7: % m² (this space / total)
        formula = f"=AREAS!{areas_col}{_ROW_PCT}"
        c = ws.cell(data_row, 7, value=formula)
        _style(c, fill=fill, font=_FONT_NORMAL, alignment=_CENTER,
               number_format='0.00"%"')

        # Col 8: Presupuesto (static coste from Presto)
        c = ws.cell(data_row, 8, value=coste if coste else "")
        _style(c, fill=fill, font=_FONT_NORMAL, alignment=_CENTER,
               number_format='#,##0.00 "€"')

        # Col 9: Presupuesto Prorrateado = coste × (m2_space / total_m2)
        m2_total_ref = f"AREAS!{get_column_letter(_COL_ZONA)}{_ROW_TOTAL}"
        space_m2_ref = f"AREAS!{areas_col}{_ROW_TOTAL}"
        formula = (
            f"=IF({m2_total_ref}>0,"
            f"H{data_row}*{space_m2_ref}/{m2_total_ref},"
            f"H{data_row})"
        )
        c = ws.cell(data_row, 9, value=formula)
        _style(c, fill=fill, font=_FONT_NORMAL, alignment=_CENTER,
               number_format='#,##0.00 "€"')

        # Col 10: Ratio €/m²
        formula = f'=IF(F{data_row}>0,I{data_row}/F{data_row},"")'
        c = ws.cell(data_row, 10, value=formula)
        _style(c, fill=_YELLOW if coste else fill, font=_FONT_BOLD, alignment=_CENTER,
               number_format='#,##0.00 "€/m²"')

        data_row += 1

    # Final row: totals
    data_row += 1
    c = ws.cell(data_row, 1, value="TOTAL")
    _style(c, fill=_GREEN_DARK, font=Font(bold=True, color="FFFFFF", size=10), alignment=_LEFT)

    for col, (formula_template, fmt) in enumerate([
        (f"=AREAS!{get_column_letter(_COL_ZONA)}{_ROW_PS}", '#,##0.00" m²"'),
        (f"=AREAS!{get_column_letter(_COL_ZONA)}{_ROW_PB}", '#,##0.00" m²"'),
        (f"=AREAS!{get_column_letter(_COL_ZONA)}{_ROW_PP}", '#,##0.00" m²"'),
        (f"=AREAS!{get_column_letter(_COL_ZONA)}{_ROW_TOTAL}", '#,##0.00" m²"'),
        ("", ""),
        (str(round(total_coste, 2)) if total_coste else "", '#,##0.00 "€"'),
        ("", ""),
        ("", ""),
    ], start=3):
        c = ws.cell(data_row, col, value=formula_template)
        _style(c, fill=_GREEN_DARK, font=Font(bold=True, color="FFFFFF", size=10),
               alignment=_CENTER, number_format=fmt)

    ws.freeze_panes = "A5"
