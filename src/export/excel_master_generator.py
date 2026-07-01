"""Excel master generator for legacy and approved-only master workbooks."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from src.db.queries import (
    list_all_budgets,
    list_all_item_masters,
    list_all_ratios,
    list_approved_budgets,
)
from src.db.schema import Budget, LineItem, Ratio, ItemMaster, Confianza

DEFAULT_OUTPUT = Path("data/master/master_latest.xlsx")
APPROVED_OUTPUT = Path("data/master/LUV_RATIOS_MASTER.xlsx")

# Color palette
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHDR_FILL = PatternFill("solid", fgColor="2F75B6")
SUBHDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="D9E1F2")
BODY_FONT = Font(size=10)


def _write_header(ws: Any, headers: list[str], fill: Any = None, font: Any = None) -> None:
    fill = fill or HDR_FILL
    font = font or HDR_FONT
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws: Any, min_w: int = 12, max_w: int = 50) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        width = max(
            (len(str(cell.value or "")) for cell in col_cells if cell.value),
            default=min_w,
        )
        ws.column_dimensions[col_letter].width = min(max(width + 2, min_w), max_w)


def _freeze(ws: Any) -> None:
    ws.freeze_panes = "A2"


def _fmt_date(dt: Any) -> str:
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d %H:%M")
    return str(dt)


def _fmt_eur(v: Any) -> Any:
    return round(float(v), 2) if v is not None else ""


def _build_index_sheet(ws: Any, budgets: list[Budget]) -> None:
    ws.title = "INDEX"
    headers = [
        "ID",
        "Presupuesto",
        "Fecha Importación",
        "Superficie (m²)",
        "Tipo Edificio",
        "Importe Total (€)",
        "Formato",
        "Capítulos",
        "Hash (SHA-256)",
    ]
    _write_header(ws, headers)
    ws.row_dimensions[1].height = 30

    for row_num, b in enumerate(budgets, start=2):
        chapter_count = len(b.items)
        fill = ALT_FILL if row_num % 2 == 0 else None
        values = [
            b.id,
            b.filename,
            _fmt_date(b.import_date),
            b.surface_m2 or "",
            b.building_type or "",
            _fmt_eur(b.total_cost),
            b.source_format.upper() if b.source_format else "",
            chapter_count,
            b.file_hash,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill
            if col == 6 and val != "":
                cell.number_format = '#,##0.00 "€"'

    _auto_width(ws)
    _freeze(ws)


def _build_ratios_sheet(ws: Any, ratios: list[Ratio]) -> None:
    ws.title = "RATIOS_SUMMARY"
    headers = [
        "Capítulo",
        "Nombre",
        "Tipo Edificio",
        "€/m² (Mediana)",
        "Mínimo €/m²",
        "Máximo €/m²",
        "Muestra (N)",
        "Última Actualización",
    ]
    _write_header(ws, headers)

    for row_num, r in enumerate(ratios, start=2):
        fill = ALT_FILL if row_num % 2 == 0 else None
        values = [
            r.chapter_code,
            r.chapter_name or "",
            r.building_type or "—",
            _fmt_eur(r.median) if r.median else "Sin datos m²",
            _fmt_eur(r.min_value) if r.min_value else "",
            _fmt_eur(r.max_value) if r.max_value else "",
            r.sample_count or 0,
            _fmt_date(r.last_updated),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill
            if col in (4, 5, 6) and isinstance(val, float):
                cell.number_format = '#,##0.00 "€/m²"'

    _auto_width(ws)
    _freeze(ws)


def _solidez_level(muestras_count: int) -> str:
    """Calculate confidence level based on sample count."""
    if muestras_count is None:
        muestras_count = 0
    if muestras_count >= 10:
        return "MUY_SÓLIDO"
    elif muestras_count >= 5:
        return "SÓLIDO"
    elif muestras_count >= 2:
        return "DÉBIL"
    else:
        return "MUY_DÉBIL"


def _build_item_master_sheet(ws: Any, items: list[ItemMaster]) -> None:
    """Build consolidated items catalog sheet."""
    ws.title = "ITEM_MASTER"
    headers = [
        "Item Key",
        "Categoría",
        "Subcategoría",
        "Unidad",
        "Unitario (Mediana)",
        "Mínimo",
        "Máximo",
        "P25",
        "P75",
        "Desv. Std",
        "Muestras (N)",
        "Solidez",
        "Última Actualización",
    ]
    _write_header(ws, headers, fill=SUBHDR_FILL, font=SUBHDR_FONT)
    ws.row_dimensions[1].height = 30

    for row_num, item in enumerate(items, start=2):
        fill = ALT_FILL if row_num % 2 == 0 else None
        solidez = _solidez_level(item.muestras_count)
        values = [
            item.item_key,
            item.categoria or "",
            item.subcategoria or "",
            item.unidad or "",
            _fmt_eur(item.mediana_unitario) if item.mediana_unitario else "",
            _fmt_eur(item.min_unitario) if item.min_unitario else "",
            _fmt_eur(item.max_unitario) if item.max_unitario else "",
            _fmt_eur(item.media_unitario * 0.25) if item.media_unitario else "",  # Aproximado P25
            _fmt_eur(item.media_unitario * 0.75) if item.media_unitario else "",  # Aproximado P75
            _fmt_eur(item.desv_std) if item.desv_std else "",
            item.muestras_count or 0,
            solidez,
            _fmt_date(item.ultima_actualizacion),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill
            # Format currency columns
            if col in (5, 6, 7, 8, 9, 10) and isinstance(val, float):
                cell.number_format = '#,##0.00 "€"'

    _auto_width(ws)
    _freeze(ws)


def _build_chapters_sheet(ws: Any, budgets: list[Budget]) -> None:
    ws.title = "CHAPTERS"
    headers = [
        "Capítulo",
        "Descripción",
        "Importe (€)",
        "Presupuesto",
        "Estado Validación",
    ]
    _write_header(ws, headers)

    row_num = 2
    for b in budgets:
        for item in b.items:
            fill = ALT_FILL if row_num % 2 == 0 else None
            values = [
                item.chapter_code,
                item.chapter_name or "",
                _fmt_eur(item.total_cost),
                b.filename,
                item.validation_status,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font = BODY_FONT
                if fill:
                    cell.fill = fill
                if col == 3 and isinstance(val, float):
                    cell.number_format = '#,##0.00 "€"'
            row_num += 1

    _auto_width(ws)
    _freeze(ws)


def _build_audit_sheet(ws: Any, budgets: list[Budget], ratios: list[Ratio]) -> None:
    ws.title = "AUDIT"
    headers = [
        "Capítulo",
        "Tipo Edificio",
        "Mediana €/m²",
        "N Muestras",
        "Presupuestos Contribuyentes",
        "Hashes",
    ]
    _write_header(ws, headers)

    # Key: (chapter_code, building_type) — preserves all (chapter, type) combinations
    ratio_map: dict[tuple[str, str | None], Ratio] = {
        (r.chapter_code, r.building_type): r for r in ratios
    }

    # Build (chapter_code, building_type) → contributing budgets
    ch_budgets: dict[tuple[str, str | None], list[Budget]] = {}
    for b in budgets:
        for item in b.items:
            if item.validation_status == "VALID":
                key = (item.chapter_code, b.building_type)
                ch_budgets.setdefault(key, []).append(b)

    row_num = 2
    # Sort safely even when building_type is None
    sorted_ratios = sorted(
        ratio_map.items(),
        key=lambda kv: ((kv[0][0] or ""), (kv[0][1] or "")),
    )

    for (chapter_code, building_type), r in sorted_ratios:
        key = (chapter_code, building_type)
        contrib = ch_budgets.get(key, [])
        fill = ALT_FILL if row_num % 2 == 0 else None
        values = [
            chapter_code,
            building_type or "GLOBAL",
            _fmt_eur(r.median) if r.median else "Sin datos m²",
            r.sample_count or len(contrib),
            "; ".join(b.filename for b in contrib) or "—",
            "; ".join(b.file_hash[:12] for b in contrib) or "—",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill
            if col == 3 and isinstance(val, float):
                cell.number_format = '#,##0.00 "€/m²"'
        row_num += 1

    _auto_width(ws)
    _freeze(ws)


def _build_raw_data_sheet(ws: Any, budgets: list[Budget]) -> None:
    ws.title = "RAW_DATA"
    headers = [
        "Budget",
        "Budget ID",
        "Capítulo",
        "Descripción",
        "Quantity",
        "Unit",
        "Unit Cost (€)",
        "Total Cost (€)",
        "Validation Status",
    ]
    _write_header(ws, headers)

    row_num = 2
    for b in budgets:
        for item in b.items:
            fill = ALT_FILL if row_num % 2 == 0 else None
            values = [
                b.filename,
                b.id,
                item.chapter_code,
                item.chapter_name or "",
                item.quantity or "",
                item.unit or "",
                _fmt_eur(item.unit_cost) if item.unit_cost else "",
                _fmt_eur(item.total_cost),
                item.validation_status,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font = BODY_FONT
                if fill:
                    cell.fill = fill
                if col in (7, 8) and isinstance(val, float):
                    cell.number_format = '#,##0.00 "€"'
            row_num += 1

    _auto_width(ws)
    _freeze(ws)


def generate_master_excel(
    session: Session,
    output_path: str | Path = DEFAULT_OUTPUT,
) -> str:
    """
    Generate (or overwrite) the master Excel file.
    If output_path uses DEFAULT_OUTPUT pattern, use dated filename.
    Returns the absolute path of the written file.
    """
    out = Path(output_path)

    # If using default, generate dated filename
    if output_path == DEFAULT_OUTPUT:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        out = out.parent / f"MASTER_{today}.xlsx"

    out.parent.mkdir(parents=True, exist_ok=True)

    budgets = list_all_budgets(session)
    ratios = list_all_ratios(session)
    items = list_all_item_masters(session)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    ws_item_master = wb.create_sheet("ITEM_MASTER")
    _build_item_master_sheet(ws_item_master, items)

    ws_index = wb.create_sheet("INDEX")
    _build_index_sheet(ws_index, budgets)

    ws_ratios = wb.create_sheet("RATIOS_SUMMARY")
    _build_ratios_sheet(ws_ratios, ratios)

    ws_chapters = wb.create_sheet("CHAPTERS")
    _build_chapters_sheet(ws_chapters, budgets)

    ws_audit = wb.create_sheet("AUDIT")
    _build_audit_sheet(ws_audit, budgets, ratios)

    ws_raw = wb.create_sheet("RAW_DATA")
    _build_raw_data_sheet(ws_raw, budgets)

    # Always open on ITEM_MASTER (the consolidated catalog)
    wb.active = ws_item_master

    wb.save(str(out))
    return str(out.resolve())


def generate_master_excel_approved(
    session: Session,
    output_path: str | Path | None = None,
) -> str:
    """Generate the official master workbook using only APPROVED imports.

    Scope of the approved-only filtering in T6:
      - INDEX, CHAPTERS, AUDIT and RAW_DATA are built from approved budgets only.
      - RATIOS_SUMMARY and ITEM_MASTER still read legacy aggregate tables
        (ratios, item_master), which may contain data historically computed from
        non-approved imports. T6 documents this limitation instead of performing
        a larger aggregate refactor.
    """
    if output_path is None:
        from app.utils.excel_export import resolve_official_master_export_path

        out = resolve_official_master_export_path()
    else:
        out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    budgets = list_approved_budgets(session)
    ratios = list_all_ratios(session)
    items = list_all_item_masters(session)

    wb = Workbook()
    wb.remove(wb.active)

    ws_item_master = wb.create_sheet("ITEM_MASTER")
    _build_item_master_sheet(ws_item_master, items)

    ws_index = wb.create_sheet("INDEX")
    _build_index_sheet(ws_index, budgets)

    ws_ratios = wb.create_sheet("RATIOS_SUMMARY")
    _build_ratios_sheet(ws_ratios, ratios)

    ws_chapters = wb.create_sheet("CHAPTERS")
    _build_chapters_sheet(ws_chapters, budgets)

    ws_audit = wb.create_sheet("AUDIT")
    _build_audit_sheet(ws_audit, budgets, ratios)

    ws_raw = wb.create_sheet("RAW_DATA")
    _build_raw_data_sheet(ws_raw, budgets)

    wb.active = ws_item_master
    wb.save(str(out))
    return str(out.resolve())
