"""Excel reader: extracts chapters and amounts from .xlsx files."""

from __future__ import annotations

import difflib
import re
import unicodedata
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

CHAPTER_CODE_RE = re.compile(r"^C\d{2,}", re.IGNORECASE)
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}

# Keywords that suggest a column contains amounts/costs
AMOUNT_KEYWORDS = ["importe", "total", "coste", "presupuesto", "pres", "cost", "amount"]
CODE_KEYWORDS = ["codigo", "cod", "ref", "code", "cap"]
NAME_KEYWORDS = ["descripcion", "resumen", "concepto", "nombre", "desc", "name"]

# Keywords that indicate a row is a total/summary and should be skipped
TOTAL_ROW_KEYWORDS = [
    "presupuesto de ejecucion", "presupuesto base", "suma", "honorarios",
    "beneficio", "gastos generales", "total general", "iva", "subtotal",
    "deducciones", "incremento",
]


def _norm(text: str) -> str:
    v = unicodedata.normalize("NFKD", text or "")
    v = "".join(ch for ch in v if not unicodedata.combining(ch))
    return v.lower().strip()


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        return v if v > 0 else None
    s = str(value).strip().replace(",", ".").replace(" ", "")
    # Remove currency symbols
    s = re.sub(r"[€$£]", "", s)
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


def _best_match_col(header_row: list[str], keywords: list[str]) -> Optional[int]:
    """Return 0-based column index whose header best matches any keyword."""
    best_idx = None
    best_score = 0.0
    for i, cell in enumerate(header_row):
        n = _norm(cell)
        if not n:
            continue
        for kw in keywords:
            score = difflib.SequenceMatcher(None, n, kw).ratio()
            if score > best_score:
                best_score = score
                best_idx = i
    return best_idx if best_score > 0.4 else None


def _find_header_row(rows: list[list[Any]], max_scan: int = 20) -> Optional[int]:
    """Return 0-based row index of the header row."""
    for i, row in enumerate(rows[:max_scan]):
        texts = [_norm(str(c)) for c in row if c is not None]
        hit = sum(
            1
            for t in texts
            if any(kw in t for kw in AMOUNT_KEYWORDS + CODE_KEYWORDS + NAME_KEYWORDS)
        )
        if hit >= 2:
            return i
    return None


def _detect_col_indices(
    header_row: list[Any],
) -> tuple[Optional[int], Optional[int], Optional[int]]:
    """Return (code_col, name_col, amount_col) as 0-based indices."""
    texts = [str(c) if c is not None else "" for c in header_row]
    code_col = _best_match_col(texts, CODE_KEYWORDS)
    name_col = _best_match_col(texts, NAME_KEYWORDS)
    amount_col = _best_match_col(texts, AMOUNT_KEYWORDS)
    return code_col, name_col, amount_col


def _best_numeric_col(rows: list[list[Any]], start_row: int, num_cols: int) -> Optional[int]:
    """Among numeric columns, return the index of the one with the largest median value."""
    col_values: dict[int, list[float]] = {}
    for row in rows[start_row:]:
        for cidx in range(num_cols):
            if cidx >= len(row):
                continue
            v = _to_float(row[cidx])
            if v is not None:
                col_values.setdefault(cidx, []).append(v)

    if not col_values:
        return None

    import statistics as _stats
    best_col = max(col_values, key=lambda c: _stats.median(col_values[c]))
    return best_col


def _extract_chapters_from_sheet(ws: Any) -> list[dict[str, Any]]:
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    if not rows:
        return []

    num_cols = max((len(r) for r in rows), default=0)
    header_idx = _find_header_row(rows)

    if header_idx is not None:
        header = rows[header_idx]
        code_col, name_col, amount_col = _detect_col_indices(header)
        data_rows = rows[header_idx + 1 :]
    else:
        # Fallback: code=col0, name=col1, amount=col with largest values
        code_col, name_col = 0, 1
        amount_col = _best_numeric_col(rows, 0, num_cols)
        if amount_col is None:
            amount_col = 2
        data_rows = rows

    chapters: list[dict[str, Any]] = []
    for ridx, row in enumerate(data_rows):
        if not any(c is not None for c in row):
            continue

        def _get(col: Optional[int]) -> Any:
            if col is None or col >= len(row):
                return None
            return row[col]

        raw_code = _get(code_col)
        raw_name = _get(name_col)
        raw_amount = _get(amount_col)

        code = str(raw_code).strip() if raw_code is not None else ""
        name = str(raw_name).strip() if raw_name is not None else ""

        # Skip rows where both code and name are empty
        if not code and not name:
            continue

        # Skip total/summary rows
        label = _norm(code + " " + name)
        if any(kw in label for kw in TOTAL_ROW_KEYWORDS):
            continue

        # Skip rows where code is purely numeric (e.g. "1", "3 = 1 + 2")
        if re.match(r"^[\d\s=+\-*/]+$", code):
            continue

        amount = _to_float(raw_amount)

        # Determine validation status
        if amount is None:
            validation_status = "DUBIOUS"
            validation_reason = "importe faltante o invalido"
        elif not code:
            validation_status = "DUBIOUS"
            validation_reason = "codigo de capitulo ausente"
        else:
            validation_status = "VALID"
            validation_reason = None

        chapters.append(
            {
                "chapter_code": code or name[:50],
                "chapter_name": name or code,
                "total_cost": amount,
                "validation_status": validation_status,
                "validation_reason": validation_reason,
                "source_row": header_idx + 1 + ridx if header_idx is not None else ridx,
            }
        )

    return chapters


def read_excel(filepath: str | Path) -> dict[str, Any]:
    """
    Read an Excel file and extract chapter-level cost data.

    Returns a dict with:
      - source_format: 'excel'
      - filename: str
      - sheets_processed: list[str]
      - chapters: list[dict]  (chapter_code, chapter_name, total_cost, confidence)
      - total_cost: float | None
      - warnings: list[str]
      - errors: list[str]
    """
    path = Path(filepath)
    result: dict[str, Any] = {
        "source_format": "excel",
        "filename": path.name,
        "filepath": str(path),
        "sheets_processed": [],
        "chapters": [],
        "total_cost": None,
        "warnings": [],
        "errors": [],
    }

    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        result["errors"].append(f"UNSUPPORTED_EXTENSION: {path.suffix}")
        return result

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        result["errors"].append(f"OPEN_FAILED: {exc}")
        return result

    all_chapters: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            if ws.sheet_state in ("hidden", "veryHidden"):
                continue
            chapters = _extract_chapters_from_sheet(ws)
            if chapters:
                all_chapters.extend(chapters)
                result["sheets_processed"].append(sheet_name)
        except Exception as exc:
            result["warnings"].append(f"SHEET_ERROR sheet={sheet_name!r}: {exc}")

    wb.close()

    # Deduplicate by chapter_code — keep first occurrence (earlier sheets have priority)
    seen_codes: set[str] = set()
    deduped: list[dict[str, Any]] = []
    for ch in all_chapters:
        code = ch["chapter_code"]
        if code not in seen_codes:
            seen_codes.add(code)
            deduped.append(ch)

    if not deduped:
        result["warnings"].append("NO_CHAPTERS_DETECTED")

    result["chapters"] = deduped
    # Sum only chapters with valid amounts for the reported total
    valid_totals = [c["total_cost"] for c in deduped if c["total_cost"] and c.get("validation_status") == "VALID"]
    result["total_cost"] = sum(valid_totals) if valid_totals else None

    return result
