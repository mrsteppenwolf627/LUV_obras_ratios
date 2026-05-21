#!/usr/bin/env python3
"""Create, update, and harden a controlled live Excel master template."""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
import hashlib
from pathlib import Path
import re
import shutil
import sys
from typing import Iterable
from uuid import uuid4

from openpyxl import Workbook, load_workbook  # type: ignore

try:
    from scripts.live_excel_integrity import (
        REQUIRED_SHEETS_COLUMNS,
        ReferentialValidationError,
        SchemaValidationError,
        contains_allowed_anchor,
        ensure_allowed_output_path,
        ensure_allowed_snapshot_path,
        validate_workbook_file,
        validate_workbook_schema,
    )
    from scripts.live_excel_preservation import (
        PRESERVED_BUDGETS_INDEX,
        PRESERVED_BUDGET_SHEETS,
        PRESERVED_TO_COST_ITEMS_MAP,
        build_preserved_visible_sheet_name,
        next_preserved_budget_sequence,
        utc_now_iso as preservation_utc_now_iso,
    )
    from scripts.live_excel_dry_run_evaluator import evaluate_dry_run_workbook
    from scripts.live_excel_professional_output import append_professional_budget_review
    from scripts.live_excel_workbook_formatting import apply_workbook_professional_formatting
    from scripts.xlsx_sheet_semantic_classifier import (
        SHEET_BUDGET_SUMMARY,
        SHEET_COMPARISON_TABLE,
        SHEET_SPACE_BREAKDOWN,
        SheetSemanticClassification,
        classify_worksheet_semantics,
    )
    from scripts.xlsx_budget_detection import (
        MAPPING_AMBIGUOUS,
        MAPPING_MAPPED,
        MAPPING_MANUAL_REVIEW,
        MAPPING_NOT_COST_ITEM,
        MAPPING_UNMAPPED,
        BudgetRowExtraction,
        classify_rows_in_worksheet,
        detect_header_row_and_mapping,
        extract_budget_rows_from_worksheet,
        mapping_status_for_row_class,
        normalize_label,
        parse_budget_number,
    )
except ModuleNotFoundError:
    from live_excel_integrity import (  # type: ignore
        REQUIRED_SHEETS_COLUMNS,
        ReferentialValidationError,
        SchemaValidationError,
        contains_allowed_anchor,
        ensure_allowed_output_path,
        ensure_allowed_snapshot_path,
        validate_workbook_file,
        validate_workbook_schema,
    )
    from live_excel_preservation import (  # type: ignore
        PRESERVED_BUDGETS_INDEX,
        PRESERVED_BUDGET_SHEETS,
        PRESERVED_TO_COST_ITEMS_MAP,
        build_preserved_visible_sheet_name,
        next_preserved_budget_sequence,
        utc_now_iso as preservation_utc_now_iso,
    )
    from live_excel_dry_run_evaluator import evaluate_dry_run_workbook  # type: ignore
    from live_excel_professional_output import append_professional_budget_review  # type: ignore
    from live_excel_workbook_formatting import apply_workbook_professional_formatting  # type: ignore
    from xlsx_sheet_semantic_classifier import (  # type: ignore
        SHEET_BUDGET_SUMMARY,
        SHEET_COMPARISON_TABLE,
        SHEET_SPACE_BREAKDOWN,
        SheetSemanticClassification,
        classify_worksheet_semantics,
    )
    from xlsx_budget_detection import (  # type: ignore
        MAPPING_AMBIGUOUS,
        MAPPING_MAPPED,
        MAPPING_MANUAL_REVIEW,
        MAPPING_NOT_COST_ITEM,
        MAPPING_UNMAPPED,
        BudgetRowExtraction,
        classify_rows_in_worksheet,
        detect_header_row_and_mapping,
        extract_budget_rows_from_worksheet,
        mapping_status_for_row_class,
        normalize_label,
        parse_budget_number,
    )

DEFAULT_OUTPUT = Path("outputs/live_excel_master/live_excel_master.xlsx")
ALLOWED_OUTPUT_ROOT = Path("outputs/live_excel_master")
SNAPSHOTS_DIRNAME = "snapshots"
DEFAULT_RETENTION_MAX = 5
OPERATIONAL_PREVIEW_SHEET = "IMPORTED_BUDGET_VIEW"
PREVIEW_PIPELINE_PHASE = "9.18"
OPERATIONAL_PREVIEW_COLUMNS = [
    "preview_row_id",
    "source_file_id",
    "import_batch_id",
    "budget_version_id",
    "source_sheet_name",
    "source_row_number",
    "chapter_code",
    "chapter_name",
    "item_code",
    "item_description",
    "unit",
    "quantity",
    "unit_price",
    "amount",
    "currency",
    "validation_status",
    "preview_only",
    "notes",
]
PRESERVED_TRACE_COLUMNS = ["__source_sheet_name", "__source_row_number", "__source_column_number"]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()




def _create_workbook_template(phase_value: str = "9.9") -> Workbook:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    for sheet_name, columns in REQUIRED_SHEETS_COLUMNS.items():
        ws = workbook.create_sheet(title=sheet_name)
        ws.append(columns)
        if sheet_name == "README_MASTER":
            ws.append(["master_contract_version", "phase_9_1_v1", utc_now_iso(), "system"])
            ws.append(["phase", phase_value, utc_now_iso(), "system"])
            ws.append(["real_data_allowed", "false", utc_now_iso(), "system"])

    return workbook


def _normalize_header(text: str) -> str:
    return normalize_label(text)


def _to_number_text(value: object) -> str:
    return parse_budget_number(value, field_context="amount").normalized


def _safe_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _upsert_readme_field(readme_ws: object, key: str, value: str, updated_by: str = "system") -> None:
    key_norm = key.strip().lower()
    for row_idx in range(2, readme_ws.max_row + 1):
        current_key = str(readme_ws.cell(row=row_idx, column=1).value or "").strip().lower()
        if current_key == key_norm:
            readme_ws.cell(row=row_idx, column=2, value=value)
            readme_ws.cell(row=row_idx, column=3, value=utc_now_iso())
            readme_ws.cell(row=row_idx, column=4, value=updated_by)
            return
    readme_ws.append([key, value, utc_now_iso(), updated_by])


def validate_generated_xlsx_preview(
    output_path: Path,
    required_phase: str = PREVIEW_PIPELINE_PHASE,
) -> dict[str, str]:
    wb = load_workbook(output_path, data_only=False)
    errors: list[str] = []
    try:
        sheetnames = wb.sheetnames
        if "INDEX" not in sheetnames:
            errors.append("missing_INDEX")
        review_candidates = sorted(
            [
                name
                for name in sheetnames
                if name.startswith("BUDGET_REVIEW_") and not name.startswith("BUDGET_REVIEW_TRACE_")
            ]
        )
        review_home_candidates = [name for name in review_candidates if re.fullmatch(r"BUDGET_REVIEW_\d{3}", name)]
        if not review_home_candidates:
            errors.append("missing_BUDGET_REVIEW_HOME")
            review_home_sheet_name = ""
        else:
            review_home_sheet_name = review_home_candidates[0]
        if not review_candidates:
            errors.append("missing_BUDGET_REVIEW")
        if review_home_sheet_name:
            detail_prefix = review_home_sheet_name + "_"
            review_detail_candidates = [name for name in review_candidates if name.startswith(detail_prefix)]
            if not review_detail_candidates:
                errors.append("missing_adaptive_review_views")
        else:
            review_detail_candidates = []
        trace_candidates = [
            name
            for name in sheetnames
            if name.startswith("BUDGET_REVIEW_TRACE_")
        ]
        if not trace_candidates:
            errors.append("missing_BUDGET_REVIEW_TRACE")

        if wb.active.title != "INDEX":
            errors.append(f"active_sheet_not_INDEX:{wb.active.title}")

        if "README_MASTER" in sheetnames:
            readme_ws = wb["README_MASTER"]
            phase_value = ""
            for row_idx in range(2, readme_ws.max_row + 1):
                key = str(readme_ws.cell(row=row_idx, column=1).value or "").strip().lower()
                if key == "phase":
                    phase_value = str(readme_ws.cell(row=row_idx, column=2).value or "").strip()
                    break
            if phase_value != required_phase:
                errors.append(f"readme_phase_mismatch:{phase_value or 'empty'}")
        else:
            errors.append("missing_README_MASTER")

        if "INDEX" in sheetnames:
            index_ws = wb["INDEX"]
            for row in index_ws.iter_rows(min_row=1, max_row=index_ws.max_row, min_col=1, max_col=min(6, index_ws.max_column)):
                for cell in row:
                    value = cell.value
                    if isinstance(value, str):
                        upper = value.strip().upper()
                        if upper.startswith("=HYPERLINK("):
                            errors.append(f"index_formula_hyperlink:{cell.coordinate}")
                        if "HYPERLINK is not implemented" in value:
                            errors.append(f"index_technical_hyperlink_text:{cell.coordinate}")

        home_entries: list[dict[str, str]] = []
        comparison_views: set[str] = set()
        space_views: set[str] = set()
        if review_home_sheet_name:
            home_ws = wb[review_home_sheet_name]
            headers = [str(home_ws.cell(row=4, column=idx).value or "").strip() for idx in range(1, home_ws.max_column + 1)]
            idx_home = {header: pos + 1 for pos, header in enumerate(headers)}
            if not {"Hoja origen", "Tipo semantico", "Vista profesional"}.issubset(idx_home.keys()):
                errors.append("review_home_missing_expected_columns")
            else:
                for row_idx in range(5, home_ws.max_row + 1):
                    source_sheet_name = str(home_ws.cell(row=row_idx, column=idx_home["Hoja origen"]).value or "").strip()
                    sheet_type = str(home_ws.cell(row=row_idx, column=idx_home["Tipo semantico"]).value or "").strip()
                    view_sheet_name = str(home_ws.cell(row=row_idx, column=idx_home["Vista profesional"]).value or "").strip()
                    if not source_sheet_name and not view_sheet_name:
                        continue
                    home_entries.append(
                        {
                            "source_sheet_name": source_sheet_name,
                            "sheet_type": sheet_type,
                            "view_sheet_name": view_sheet_name,
                        }
                    )
                    if not view_sheet_name or view_sheet_name not in sheetnames:
                        errors.append(f"review_view_missing:{source_sheet_name or 'unknown'}")
                    if sheet_type == SHEET_COMPARISON_TABLE:
                        comparison_views.add(view_sheet_name)
                    if sheet_type == SHEET_SPACE_BREAKDOWN:
                        space_views.add(view_sheet_name)
        if len({entry["sheet_type"] for entry in home_entries if entry["sheet_type"]}) > 1:
            if len({entry["view_sheet_name"] for entry in home_entries if entry["view_sheet_name"]}) < len(home_entries):
                errors.append("review_semantic_views_mixed")

        for review_sheet_name in review_candidates:
            review_ws = wb[review_sheet_name]
            headers = [str(review_ws.cell(row=4, column=idx).value or "").strip() for idx in range(1, review_ws.max_column + 1)]
            idx = {header: pos + 1 for pos, header in enumerate(headers)}
            if "_review_row_id" in idx:
                letter = review_ws.cell(row=4, column=idx["_review_row_id"]).column_letter
                if not bool(review_ws.column_dimensions[letter].hidden):
                    errors.append(f"review_row_id_not_hidden:{review_sheet_name}")
            numeric_desc_rows: list[int] = []
            formula_desc_rows: list[int] = []
            aux_desc_rows: list[int] = []
            name_error_rows: list[int] = []
            aux_pattern = re.compile(r"(?:^=)|(?:\\bPEM\\b)|(?:\\bHONPRY\\b)|(?:\\bHONDIR\\b)|(?:\\bPOR[A-Z])", re.IGNORECASE)
            numeric_pattern = re.compile(r"^-?\d+(?:[.,]\d+)?$")
            desc_col = idx.get("Descripcion")
            amount_col = idx.get("Importe")
            for row_idx in range(5, review_ws.max_row + 1):
                if desc_col:
                    desc = str(review_ws.cell(row=row_idx, column=desc_col).value or "").strip()
                    if numeric_pattern.fullmatch(desc):
                        amount_cell = str(review_ws.cell(row=row_idx, column=amount_col).value or "").strip() if amount_col else ""
                        if amount_cell in {"", "None"}:
                            numeric_desc_rows.append(row_idx)
                    if desc.startswith("="):
                        formula_desc_rows.append(row_idx)
                    if aux_pattern.search(desc):
                        aux_desc_rows.append(row_idx)
                    if "revision" in desc.lower() and amount_col:
                        amount_val = str(review_ws.cell(row=row_idx, column=amount_col).value or "").strip()
                        if amount_val.isdigit() and 1900 <= int(amount_val) <= 2100:
                            errors.append(f"review_metadata_row_misclassified:{review_sheet_name}:{row_idx}")
                for col_idx in range(1, min(review_ws.max_column, 12) + 1):
                    value = review_ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(value, str) and "#NAME?" in value.upper():
                        name_error_rows.append(row_idx)
                        break
            if numeric_desc_rows:
                errors.append(f"review_numeric_description_rows:{review_sheet_name}:{','.join(map(str, numeric_desc_rows[:10]))}")
            if formula_desc_rows:
                errors.append(f"review_formula_description_rows:{review_sheet_name}:{','.join(map(str, formula_desc_rows[:10]))}")
            if aux_desc_rows:
                errors.append(f"review_auxiliary_description_rows:{review_sheet_name}:{','.join(map(str, aux_desc_rows[:10]))}")
            if name_error_rows:
                errors.append(f"review_name_error_rows:{review_sheet_name}:{','.join(map(str, sorted(set(name_error_rows))[:10]))}")

        for comparison_view in comparison_views:
            if not comparison_view or comparison_view not in sheetnames:
                continue
            ws = wb[comparison_view]
            headers = [str(ws.cell(row=4, column=idx).value or "").strip() for idx in range(1, ws.max_column + 1)]
            required = ["Cap.", "Nombre del capítulo", "Importe (€)", "Nombre equivalente", "Importe equivalente", "Diferencia"]
            for item in required:
                if item not in headers:
                    errors.append(f"comparison_view_missing_column:{comparison_view}:{item}")
            if "Cantidad" in headers or "Precio unitario" in headers or "Ud" in headers:
                errors.append(f"comparison_view_unexpected_classic_columns:{comparison_view}")
            if "Diferencia" in headers:
                difference_col = headers.index("Diferencia") + 1
                nonempty_difference = 0
                for row_idx in range(5, ws.max_row + 1):
                    value = str(ws.cell(row=row_idx, column=difference_col).value or "").strip()
                    if value:
                        nonempty_difference += 1
                if ws.max_row >= 5 and nonempty_difference == 0:
                    errors.append(f"comparison_view_difference_lost:{comparison_view}")

        for space_view in space_views:
            if not space_view or space_view not in sheetnames:
                continue
            ws = wb[space_view]
            for row_idx in range(5, ws.max_row + 1):
                desc = str(ws.cell(row=row_idx, column=3).value or "").strip()
                if desc.lower().startswith("subtotal"):
                    errors.append(f"space_view_generated_subtotal:{space_view}:{row_idx}")

        for name in sheetnames:
            if not name.startswith("PRES_"):
                continue
            ws = wb[name]
            for col_idx in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=col_idx).value or "").strip()
                if not header.startswith("__source_"):
                    continue
                letter = ws.cell(row=1, column=col_idx).column_letter
                if not bool(ws.column_dimensions[letter].hidden):
                    errors.append(f"preserved_source_column_visible:{name}:{header}")

        if "COST_ITEMS" in sheetnames:
            cost_ws = wb["COST_ITEMS"]
            headers = [str(cost_ws.cell(row=1, column=idx).value or "").strip() for idx in range(1, cost_ws.max_column + 1)]
            idx = {header: pos + 1 for pos, header in enumerate(headers)}
            desc_col = idx.get("description_raw") or idx.get("description") or idx.get("item_description")
            amount_col = idx.get("amount_raw") or idx.get("amount")
            origin_col = idx.get("origin_record_ref")
            comparison_sources = {
                entry["source_sheet_name"]
                for entry in home_entries
                if entry.get("sheet_type") == SHEET_COMPARISON_TABLE
            }
            space_sources = {
                entry["source_sheet_name"]
                for entry in home_entries
                if entry.get("sheet_type") == SHEET_SPACE_BREAKDOWN
            }
            if desc_col:
                for row_idx in range(2, cost_ws.max_row + 1):
                    desc = str(cost_ws.cell(row=row_idx, column=desc_col).value or "").strip()
                    amount = str(cost_ws.cell(row=row_idx, column=amount_col).value or "").strip() if amount_col else ""
                    origin = str(cost_ws.cell(row=row_idx, column=origin_col).value or "").strip() if origin_col else ""
                    source_name = origin.split("!", 1)[0] if "!" in origin else ""
                    if desc.startswith("="):
                        errors.append(f"cost_items_formula_description_row:{row_idx}")
                        continue
                    if desc == "" and amount in {"", "0", "0.0", "0.00"}:
                        errors.append(f"cost_items_blank_zero_row:{row_idx}")
                        continue
                    if re.search(r"(?:\\bPEM\\b)|(?:\\bHONPRY\\b)|(?:\\bHONDIR\\b)|(?:\\bPOR[A-Z])", desc, flags=re.IGNORECASE):
                        errors.append(f"cost_items_auxiliary_description_row:{row_idx}")
                        continue
                    if source_name and source_name in comparison_sources:
                        errors.append(f"cost_items_from_comparison_sheet:{row_idx}")
                        continue
                    if source_name and source_name in space_sources and amount in {"", "0", "0.0", "0.00"}:
                        errors.append(f"cost_items_space_row_zero_amount:{row_idx}")
                        continue
    finally:
        wb.close()

    if errors:
        raise RuntimeError(
            "Generated XLSX preview failed post-generation validation: " + "; ".join(errors)
        )
    return {"post_generation_validation": "passed", "required_phase": required_phase}


def _detect_header_row_and_mapping(ws: object, max_scan_rows: int = 25) -> tuple[int, dict[str, int]]:
    detection = detect_header_row_and_mapping(ws, max_scan_rows=max_scan_rows)
    return detection.header_row, detection.mapping


def _append_operational_preview_sheet(workbook: Workbook) -> None:
    if OPERATIONAL_PREVIEW_SHEET not in workbook.sheetnames:
        ws = workbook.create_sheet(OPERATIONAL_PREVIEW_SHEET)
        ws.append(OPERATIONAL_PREVIEW_COLUMNS)


def _append_preserved_budget_scaffolding(
    workbook: Workbook,
    source_workbook: Workbook,
    source_file_id: str,
    import_batch_id: str,
    budget_version_id: str,
    cost_item_by_origin: dict[tuple[str, str], str],
) -> dict[str, str]:
    preserved_budget_seq = next_preserved_budget_sequence(workbook)
    preserved_budget_id = f"pb_{preserved_budget_seq:03d}"
    created_at = preservation_utc_now_iso()
    workbook[PRESERVED_BUDGETS_INDEX].append(
        [
            preserved_budget_id,
            source_file_id,
            import_batch_id,
            budget_version_id,
            f"{preserved_budget_seq:03d}",
            f"PRESERVED_{preserved_budget_seq:03d}",
            created_at,
            "PREVIEW_ONLY",
        ]
    )

    created_sheets = 0
    map_rows = 0
    for ws_idx, source_ws in enumerate(source_workbook.worksheets, start=1):
        row_classes = classify_rows_in_worksheet(source_ws)
        preserved_sheet_id = f"pbs_{preserved_budget_seq:03d}_{ws_idx:03d}"
        preserved_sheet_name = build_preserved_visible_sheet_name(
            budget_sequence=preserved_budget_seq,
            sheet_sequence=ws_idx,
            source_sheet_name=source_ws.title,
            existing_names=workbook.sheetnames,
        )
        preserved_ws = workbook.create_sheet(title=preserved_sheet_name)
        workbook[PRESERVED_BUDGET_SHEETS].append(
            [
                preserved_sheet_id,
                preserved_budget_id,
                source_ws.title,
                preserved_sheet_name,
                f"{ws_idx:03d}",
                "TRUE",
                str(source_ws.max_row),
                str(source_ws.max_column),
                created_at,
                "",
            ]
        )

        # Keep source tabular shape, append traceability columns at the end.
        for row_idx in range(1, source_ws.max_row + 1):
            row_values = [
                source_ws.cell(row=row_idx, column=col_idx).value
                for col_idx in range(1, source_ws.max_column + 1)
            ]
            if row_idx == 1:
                header_values = ["" if value is None else str(value) for value in row_values]
                preserved_ws.append(header_values + PRESERVED_TRACE_COLUMNS)
                continue

            text_values = ["" if value is None else str(value) for value in row_values]
            preserved_ws.append(text_values + [source_ws.title, str(row_idx), ""])

            preserved_row_id = f"pr_{preserved_budget_seq:03d}_{ws_idx:03d}_{row_idx:05d}"
            map_key = (source_ws.title, str(row_idx))
            cost_item_id = cost_item_by_origin.get(map_key, "")
            row_class = row_classes.get(row_idx, "")
            mapping_status = MAPPING_MAPPED if cost_item_id else mapping_status_for_row_class(row_class)
            if not cost_item_id and mapping_status == MAPPING_MAPPED:
                mapping_status = MAPPING_UNMAPPED
            mapping_confidence = "1.0" if cost_item_id else ("1.0" if mapping_status == MAPPING_NOT_COST_ITEM else "0.0")
            validation_status = "MANUAL_REVIEW_REQUIRED" if mapping_status in {MAPPING_AMBIGUOUS, MAPPING_MANUAL_REVIEW} else "PENDING"
            note = row_class or "NO_COST_ITEM_MATCH"
            if mapping_status == MAPPING_UNMAPPED:
                note = "NO_COST_ITEM_MATCH"
            workbook[PRESERVED_TO_COST_ITEMS_MAP].append(
                [
                    f"pm_{uuid4().hex[:12]}",
                    source_file_id,
                    import_batch_id,
                    budget_version_id,
                    preserved_sheet_id,
                    preserved_row_id,
                    source_ws.title,
                    str(row_idx),
                    cost_item_id,
                    mapping_status,
                    mapping_confidence,
                    validation_status,
                    "" if cost_item_id else note,
                ]
            )
            map_rows += 1
        created_sheets += 1

    return {
        "preserved_budget_id": preserved_budget_id,
        "preserved_sheets": str(created_sheets),
        "preserved_map_rows": str(map_rows),
    }


def _preview_row_from_extraction(extraction: BudgetRowExtraction, source_file_id: str) -> list[str]:
    return [
        f"pvr_{uuid4().hex[:10]}",
        source_file_id,
        "",
        "",
        extraction.source_sheet_name,
        str(extraction.source_row_number),
        extraction.chapter_code,
        extraction.chapter_name,
        extraction.item_code,
        extraction.item_description,
        extraction.unit,
        extraction.quantity,
        extraction.unit_price,
        extraction.amount,
        "EUR",
        extraction.validation_status,
        "TRUE",
        extraction.notes,
    ]


def _safe_cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _sheet_row_values(ws: object, row_idx: int, max_col: int) -> list[object]:
    return [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]


def _looks_like_metadata_row(description: str, amount_text: str) -> bool:
    desc = normalize_label(description)
    if not desc:
        return False
    if "revision" not in desc and "rev" not in desc and "fecha" not in desc:
        return False
    parsed = parse_budget_number(amount_text, field_context="amount")
    if not parsed.is_valid:
        return True
    try:
        value = int(float(parsed.normalized))
    except ValueError:
        return False
    return 1900 <= value <= 2100


def _semantic_amount_text(value: object) -> str:
    return parse_budget_number(value, field_context="amount").normalized


def _build_semantic_rows_for_nonclassic_sheet(
    ws: object,
    source_file_id: str,
    semantic: SheetSemanticClassification,
) -> list[BudgetRowExtraction]:
    detected = semantic.detected_columns
    max_col = min(getattr(ws, "max_column", 1), 50)
    rows: list[BudgetRowExtraction] = []
    for row_idx in range(semantic.header_row + 1, getattr(ws, "max_row", 0) + 1):
        values = _sheet_row_values(ws, row_idx, max_col)
        if all(_safe_cell_text(value) == "" for value in values):
            continue
        code = _safe_cell_text(values[detected.get("code", detected.get("cap", 0)) - 1]) if detected.get("code", detected.get("cap", 0)) else ""
        description = _safe_cell_text(
            values[detected.get("description", detected.get("chapter_name", 0)) - 1]
        ) if detected.get("description", detected.get("chapter_name", 0)) else ""
        amount = _semantic_amount_text(values[detected.get("amount", 0) - 1]) if detected.get("amount") else ""
        if not description and semantic.sheet_type == SHEET_COMPARISON_TABLE:
            description = _safe_cell_text(values[detected.get("chapter_name", 0) - 1]) if detected.get("chapter_name") else ""
        if not description and not code and not amount:
            continue
        if _looks_like_metadata_row(description, amount):
            rows.append(
                BudgetRowExtraction(
                    source_sheet_name=str(getattr(ws, "title", "")),
                    source_row_number=row_idx,
                    chapter_code="",
                    chapter_name="",
                    item_code=code,
                    item_description=description,
                    unit="",
                    quantity="",
                    unit_price="",
                    amount="",
                    validation_status="PENDING",
                    notes=f"SHEET_TYPE={semantic.sheet_type}|METADATA_ROW",
                    row_class="NON_BUDGET_ROW",
                    mapping_status=MAPPING_NOT_COST_ITEM,
                )
            )
            continue
        notes_parts = [f"SHEET_TYPE={semantic.sheet_type}"]
        if semantic.sheet_type == SHEET_COMPARISON_TABLE:
            eq_name = _safe_cell_text(values[detected.get("equivalent_name", 0) - 1]) if detected.get("equivalent_name") else ""
            eq_amount = _semantic_amount_text(values[detected.get("equivalent_amount", 0) - 1]) if detected.get("equivalent_amount") else ""
            diff_raw = _safe_cell_text(values[detected.get("difference", 0) - 1]) if detected.get("difference") else ""
            if eq_name:
                notes_parts.append(f"EQUIVALENT_NAME={eq_name}")
            if eq_amount:
                notes_parts.append(f"EQUIVALENT_AMOUNT={eq_amount}")
            if diff_raw:
                notes_parts.append(f"DIFFERENCE={diff_raw}")
        elif semantic.sheet_type == SHEET_SPACE_BREAKDOWN:
            info_col = detected.get("info")
            if info_col:
                info_text = _safe_cell_text(values[info_col - 1])
                if info_text:
                    notes_parts.append(f"INFO={info_text}")
        rows.append(
            BudgetRowExtraction(
                source_sheet_name=str(getattr(ws, "title", "")),
                source_row_number=row_idx,
                chapter_code="",
                chapter_name="",
                item_code=code,
                item_description=description,
                unit="",
                quantity="",
                unit_price="",
                amount=amount,
                validation_status="PENDING",
                notes="|".join(notes_parts),
                row_class="COST_ITEM",
                mapping_status=MAPPING_NOT_COST_ITEM,
            )
        )
    return rows


def _iter_operational_extractions_from_workbook(
    source_wb: Workbook,
    source_file_id: str,
    sheet_semantics: dict[str, SheetSemanticClassification] | None = None,
) -> list[BudgetRowExtraction]:
    out_rows: list[BudgetRowExtraction] = []
    for ws in source_wb.worksheets:
        semantic = (sheet_semantics or {}).get(ws.title) or classify_worksheet_semantics(ws)
        if semantic.sheet_type in {SHEET_COMPARISON_TABLE, SHEET_SPACE_BREAKDOWN}:
            out_rows.extend(_build_semantic_rows_for_nonclassic_sheet(ws, source_file_id, semantic))
            continue
        extracted = extract_budget_rows_from_worksheet(ws, source_file_id=source_file_id)
        if semantic.sheet_type == SHEET_BUDGET_SUMMARY:
            adjusted: list[BudgetRowExtraction] = []
            for row in extracted:
                notes = row.notes
                if _looks_like_metadata_row(row.item_description, row.amount):
                    adjusted.append(
                        BudgetRowExtraction(
                            source_sheet_name=row.source_sheet_name,
                            source_row_number=row.source_row_number,
                            chapter_code=row.chapter_code,
                            chapter_name=row.chapter_name,
                            item_code=row.item_code,
                            item_description=row.item_description,
                            unit="",
                            quantity="",
                            unit_price="",
                            amount="",
                            validation_status="PENDING",
                            notes="|".join(part for part in [notes, "METADATA_ROW", "SHEET_TYPE=BUDGET_SUMMARY"] if part),
                            row_class="NON_BUDGET_ROW",
                            mapping_status=MAPPING_NOT_COST_ITEM,
                        )
                    )
                    continue
                adjusted.append(
                    BudgetRowExtraction(
                        source_sheet_name=row.source_sheet_name,
                        source_row_number=row.source_row_number,
                        chapter_code=row.chapter_code,
                        chapter_name=row.chapter_name,
                        item_code=row.item_code,
                        item_description=row.item_description,
                        unit="",
                        quantity="",
                        unit_price="",
                        amount=row.amount,
                        validation_status=row.validation_status,
                        notes="|".join(part for part in [notes, "SHEET_TYPE=BUDGET_SUMMARY"] if part),
                        row_class=row.row_class,
                        mapping_status=row.mapping_status,
                    )
                )
            out_rows.extend(adjusted)
        else:
            out_rows.extend(extracted)
    return out_rows


def _iter_operational_rows_from_xlsx(input_path: Path, source_file_id: str) -> list[list[str]]:
    source_wb = load_workbook(input_path, data_only=False)
    try:
        extractions = _iter_operational_extractions_from_workbook(source_wb, source_file_id)
    finally:
        source_wb.close()
    return [
        _preview_row_from_extraction(extraction, source_file_id)
        for extraction in extractions
    ]


def _checksum_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8192), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _snapshot_path(base_dir: Path, label: str) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return base_dir / f"{timestamp}_{label}_live_excel_master.xlsx"


def _copy_snapshot(src: Path, snapshots_dir: Path, label: str) -> Path:
    snapshots_dir.mkdir(parents=True, exist_ok=True)
    dst = _snapshot_path(snapshots_dir, label)
    shutil.copy2(src, dst)
    return dst


def _append_row(ws: object, row: Iterable[str]) -> None:
    ws.append(list(row))


def _write_snapshot_log(
    workbook: Workbook,
    trigger: str,
    run_id: str,
    storage_ref: str,
    checksum: str,
) -> None:
    snapshots_ws = workbook["SNAPSHOTS"]
    _append_row(
        snapshots_ws,
        [
            f"snp_{trigger}_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}",
            utc_now_iso(),
            "0.1.0",
            trigger,
            run_id,
            storage_ref,
            checksum,
            "system",
        ],
    )


def _apply_snapshot_retention(snapshots_dir: Path, keep_last: int) -> list[str]:
    if keep_last <= 0 or not snapshots_dir.exists():
        return []
    files = sorted([p for p in snapshots_dir.glob("*.xlsx") if p.is_file()], key=lambda p: p.stat().st_mtime)
    to_delete = files[:-keep_last] if len(files) > keep_last else []
    deleted: list[str] = []
    for path in to_delete:
        deleted.append(path.as_posix())
        path.unlink(missing_ok=True)
    return deleted


def _add_synthetic_incremental_rows(workbook: Workbook, run_id: str) -> None:
    ts = utc_now_iso()
    import_batch_id = f"imp_{run_id}"
    source_file_id = f"sf_{run_id}"
    project_id = f"prj_{run_id}"
    budget_version_id = f"bv_{run_id}"
    raw_import_id = f"raw_{run_id}"
    cost_item_id_ok = f"ci_ok_{run_id}"
    cost_item_id_blocked = f"ci_blk_{run_id}"
    validation_result_id = f"vr_{run_id}"
    exclusion_id = f"ex_{run_id}"
    ratio_input_id = f"ri_{run_id}"

    _append_row(
        workbook["IMPORT_LOG"],
        [import_batch_id, run_id, ts, ts, "SYNTHETIC_OK", "2", "1", "", ""],
    )
    _append_row(
        workbook["SOURCE_FILES"],
        [
            source_file_id,
            f"synthetic_{run_id}.xlsx",
            f"synthetic_hash_{run_id}",
            "SYNTHETIC_XLSX",
            "LOW",
            ts,
            import_batch_id,
            "NON_SENSITIVE",
        ],
    )
    _append_row(
        workbook["PROJECTS"],
        [project_id, f"P-{run_id[:8]}", "Synthetic Project", "", "m2", "PENDING", ts, ts],
    )
    _append_row(
        workbook["BUDGET_VERSIONS"],
        [budget_version_id, project_id, source_file_id, "v_synth_1", ts[:10], "true", "VALIDATED", import_batch_id],
    )
    _append_row(
        workbook["RAW_IMPORTS"],
        [raw_import_id, source_file_id, f"raw://synthetic/{run_id}", f"raw_hash_{run_id}", ts, import_batch_id, "DENY"],
    )
    _append_row(
        workbook["COST_ITEMS"],
        [
            cost_item_id_ok,
            budget_version_id,
            source_file_id,
            "A-001",
            "Synthetic cost item OK",
            "ud",
            "1",
            "100.00",
            "100.00",
            f"row_hash_{cost_item_id_ok}",
            "VALIDATED",
        ],
    )
    _append_row(
        workbook["COST_ITEMS"],
        [
            cost_item_id_blocked,
            budget_version_id,
            source_file_id,
            "A-002",
            "Synthetic cost item BLOCKED",
            "ud",
            "1",
            "200.00",
            "200.00",
            f"row_hash_{cost_item_id_blocked}",
            "BLOCKED",
        ],
    )
    _append_row(
        workbook["VALIDATION_RESULTS"],
        [
            validation_result_id,
            "COST_ITEM",
            cost_item_id_ok,
            "RULE_SYNTH_001",
            "INFO",
            "PASS",
            "synthetic validation result",
            ts,
            import_batch_id,
        ],
    )
    _append_row(
        workbook["EXCLUSIONS"],
        [
            exclusion_id,
            "COST_ITEM",
            cost_item_id_blocked,
            "BLOCKED_STATUS",
            "blocked synthetic item",
            "true",
            ts,
            "system",
            import_batch_id,
        ],
    )
    _append_row(
        workbook["RATIO_INPUTS"],
        [
            ratio_input_id,
            f"norm_{cost_item_id_ok}",
            project_id,
            budget_version_id,
            "SYNTHETIC_ELIGIBLE",
            "false",
            "VALIDATED",
            ts,
        ],
    )
    _append_row(
        workbook["CHANGELOG"],
        [
            f"chg_{run_id}",
            ts,
            "synthetic_incremental_load",
            "MULTI",
            "Synthetic incremental load (non-operational ratios placeholders only)",
            "phase_9_3_live_excel_master_hardening",
            "system",
        ],
    )


def _safe_relative_path(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def _collect_sheet_values(workbook: Workbook, sheet_name: str, column_name: str) -> set[str]:
    ws = workbook[sheet_name]
    headers = REQUIRED_SHEETS_COLUMNS[sheet_name]
    col_idx = headers.index(column_name) + 1
    values: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value is None:
            continue
        parsed = str(value).strip()
        if parsed:
            values.add(parsed)
    return values


def _create_pre_snapshot_if_needed(output_path: Path, snapshots_dir: Path, run_id: str) -> str:
    if not output_path.exists():
        return ""
    validate_workbook_file(output_path)
    pre_snapshot = _copy_snapshot(output_path, snapshots_dir, "pre_update")
    return pre_snapshot.as_posix()


def _finalize_post_snapshot(
    output_path: Path,
    snapshots_dir: Path,
    workbook_root: Path,
    retention_max: int,
    run_id: str,
) -> tuple[str, list[str]]:
    post_snapshot = _copy_snapshot(output_path, snapshots_dir, "post_update")
    post_ref = post_snapshot.as_posix()
    deleted = _apply_snapshot_retention(snapshots_dir, retention_max)

    wb = load_workbook(output_path)
    try:
        _write_snapshot_log(
            wb,
            "post_update",
            run_id,
            _safe_relative_path(post_snapshot, workbook_root),
            _checksum_sha256(post_snapshot),
        )
        for deleted_ref in deleted:
            _append_row(
                wb["CHANGELOG"],
                [
                    f"chg_ret_{uuid4().hex[:12]}",
                    utc_now_iso(),
                    "snapshot_retention_delete",
                    "SNAPSHOTS",
                    f"Deleted snapshot due to retention policy: {deleted_ref}",
                    "phase_9_3_live_excel_master_hardening",
                    "system",
                ],
            )
        wb.save(output_path)
    finally:
        wb.close()

    return post_ref, deleted


def generate_master(output_path: Path, update: bool, retention_max: int = DEFAULT_RETENTION_MAX) -> dict[str, str]:
    if not contains_allowed_anchor(output_path):
        raise ValueError("Output path must be inside an 'outputs/live_excel_master' directory.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    snapshots_dir = output_path.parent / SNAPSHOTS_DIRNAME
    run_id = f"run_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"
    pre_snapshot_ref = ""
    post_snapshot_ref = ""

    if output_path.exists() and not update:
        raise RuntimeError("Output workbook already exists. Use --update for controlled update with snapshots.")

    if output_path.exists():
        pre_snapshot_ref = _create_pre_snapshot_if_needed(output_path, snapshots_dir, run_id)

    workbook = _create_workbook_template("9.9")
    if pre_snapshot_ref:
        _write_snapshot_log(
            workbook,
            "pre_update",
            run_id,
            pre_snapshot_ref,
            _checksum_sha256(Path(pre_snapshot_ref)),
        )
    workbook.save(output_path)
    workbook.close()

    validate_workbook_file(output_path)

    if update:
        post_snapshot_ref, _ = _finalize_post_snapshot(
            output_path=output_path,
            snapshots_dir=snapshots_dir,
            workbook_root=output_path.parent.parent.parent,
            retention_max=retention_max,
            run_id=run_id,
        )
        wb_update = load_workbook(output_path)
        try:
            _append_row(
                wb_update["CHANGELOG"],
                [
                    f"chg_{uuid4().hex[:12]}",
                    utc_now_iso(),
                    "template_update",
                    "ALL",
                    "Controlled template update with pre/post snapshots",
                    "phase_9_3_live_excel_master_hardening",
                    "system",
                ],
            )
            wb_update.save(output_path)
        finally:
            wb_update.close()

    return {
        "output": output_path.as_posix(),
        "updated": str(update).lower(),
        "pre_snapshot": pre_snapshot_ref,
        "post_snapshot": post_snapshot_ref,
    }


def load_synthetic_incremental(
    output_path: Path,
    retention_max: int = DEFAULT_RETENTION_MAX,
    run_id: str | None = None,
) -> dict[str, str]:
    if not contains_allowed_anchor(output_path):
        raise ValueError("Output path must be inside an 'outputs/live_excel_master' directory.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    snapshots_dir = output_path.parent / SNAPSHOTS_DIRNAME
    run_id = run_id or f"run_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"

    if output_path.exists():
        wb_existing = load_workbook(output_path)
        try:
            validate_workbook_schema(wb_existing)
            existing_run_ids = _collect_sheet_values(wb_existing, "IMPORT_LOG", "run_id")
            if run_id in existing_run_ids:
                return {
                    "output": output_path.as_posix(),
                    "run_id": run_id,
                    "pre_snapshot": "",
                    "post_snapshot": "",
                    "deleted_snapshots": "0",
                    "idempotent_skip": "true",
                }
        finally:
            wb_existing.close()
        pre_snapshot = _create_pre_snapshot_if_needed(output_path, snapshots_dir, run_id)
    else:
        pre_snapshot = ""
        wb_seed = _create_workbook_template("9.9")
        wb_seed.save(output_path)
        wb_seed.close()

    wb = load_workbook(output_path)
    try:
        validate_workbook_schema(wb)
        _add_synthetic_incremental_rows(wb, run_id)
        if pre_snapshot:
            _write_snapshot_log(
                wb,
                "pre_update",
                run_id,
                pre_snapshot,
                _checksum_sha256(Path(pre_snapshot)),
            )
        wb.save(output_path)
    finally:
        wb.close()

    validate_workbook_file(output_path)

    post_snapshot_ref, deleted = _finalize_post_snapshot(
        output_path=output_path,
        snapshots_dir=snapshots_dir,
        workbook_root=output_path.parent.parent.parent,
        retention_max=retention_max,
        run_id=run_id,
    )

    return {
        "output": output_path.as_posix(),
        "run_id": run_id,
        "pre_snapshot": pre_snapshot,
        "post_snapshot": post_snapshot_ref,
        "deleted_snapshots": str(len(deleted)),
        "idempotent_skip": "false",
    }


def rollback_master_from_snapshot(
    output_path: Path,
    snapshot_path: Path,
    retention_max: int = DEFAULT_RETENTION_MAX,
) -> dict[str, str]:
    if not output_path.exists():
        raise RuntimeError("Cannot rollback non-existing master workbook.")
    if not snapshot_path.exists():
        raise RuntimeError("Snapshot path does not exist.")
    if not contains_allowed_anchor(output_path):
        raise ValueError("Output path must be inside an 'outputs/live_excel_master' directory.")
    ensure_allowed_snapshot_path(snapshot_path, output_path.parent / SNAPSHOTS_DIRNAME)

    snapshots_dir = output_path.parent / SNAPSHOTS_DIRNAME
    run_id = f"run_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"
    pre_snapshot = _create_pre_snapshot_if_needed(output_path, snapshots_dir, run_id)

    try:
        shutil.copy2(snapshot_path, output_path)
        validate_workbook_file(output_path)
    except Exception as exc:
        if pre_snapshot:
            shutil.copy2(Path(pre_snapshot), output_path)
        raise RuntimeError(f"Rollback failed due to invalid snapshot content: {snapshot_path.name}") from exc

    wb = load_workbook(output_path)
    try:
        _write_snapshot_log(
            wb,
            "rollback",
            run_id,
            snapshot_path.as_posix(),
            _checksum_sha256(snapshot_path),
        )
        _append_row(
            wb["CHANGELOG"],
            [
                f"chg_rb_{uuid4().hex[:12]}",
                utc_now_iso(),
                "rollback_apply",
                "ALL",
                f"Rollback applied from snapshot {snapshot_path.name}",
                "phase_9_3_live_excel_master_hardening",
                "system",
            ],
        )
        wb.save(output_path)
    finally:
        wb.close()

    post_snapshot_ref, deleted = _finalize_post_snapshot(
        output_path=output_path,
        snapshots_dir=snapshots_dir,
        workbook_root=output_path.parent.parent.parent,
        retention_max=retention_max,
        run_id=run_id,
    )

    return {
        "output": output_path.as_posix(),
        "run_id": run_id,
        "pre_snapshot": pre_snapshot,
        "post_snapshot": post_snapshot_ref,
        "deleted_snapshots": str(len(deleted)),
    }


def generate_preview_from_real_xlsx(
    input_xlsx_path: Path,
    output_path: Path,
    source_file_id: str = "sf_preview_real_001",
    retention_max: int = DEFAULT_RETENTION_MAX,
) -> dict[str, str]:
    if input_xlsx_path.suffix.lower() != ".xlsx":
        raise ValueError("Preview input must be an .xlsx file.")
    if not input_xlsx_path.exists():
        raise RuntimeError("Preview input file does not exist.")
    if not contains_allowed_anchor(output_path):
        raise ValueError("Output path must be inside an 'outputs/live_excel_master' directory.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not output_path.exists():
        result = generate_master(output_path=output_path, update=False, retention_max=retention_max)
    else:
        validate_workbook_file(output_path)
        result = {
            "output": output_path.as_posix(),
            "updated": "true",
            "pre_snapshot": "",
            "post_snapshot": "",
        }

    source_wb = load_workbook(input_xlsx_path, data_only=False)
    wb = load_workbook(output_path)
    try:
        _append_operational_preview_sheet(wb)
        sheet_semantics = {
            ws.title: classify_worksheet_semantics(ws)
            for ws in source_wb.worksheets
        }
        operational_extractions = _iter_operational_extractions_from_workbook(
            source_wb,
            source_file_id=source_file_id,
            sheet_semantics=sheet_semantics,
        )
        operational_rows = [
            _preview_row_from_extraction(extraction, source_file_id)
            for extraction in operational_extractions
        ]
        ts = utc_now_iso()
        run_id = f"preview_run_{uuid4().hex[:8]}"
        import_batch_id = f"imp_preview_{uuid4().hex[:8]}"
        budget_version_id = f"bv_preview_{uuid4().hex[:8]}"
        project_id = f"prj_preview_{uuid4().hex[:8]}"
        raw_import_id = f"raw_preview_{uuid4().hex[:8]}"

        _upsert_readme_field(wb["README_MASTER"], "phase", PREVIEW_PIPELINE_PHASE)
        _upsert_readme_field(wb["README_MASTER"], "preview_mode", "PREVIEW_ONLY")
        _upsert_readme_field(wb["README_MASTER"], "preview_master_promotion", "false")
        wb["IMPORT_LOG"].append([import_batch_id, run_id, ts, ts, "PREVIEW_ONLY", str(len(operational_rows)), "0", "", ""])
        wb["SOURCE_FILES"].append(
            [source_file_id, "REAL_SAMPLE_SANITIZED.xlsx", "preview_only_hash", "XLSX", "MEDIUM", ts, import_batch_id, "SENSITIVE_LOCAL_ONLY"]
        )
        wb["PROJECTS"].append([project_id, "PREVIEW-001", "PREVIEW_ONLY_PROJECT", "", "m2", "PENDING", ts, ts])
        wb["BUDGET_VERSIONS"].append([budget_version_id, project_id, source_file_id, "preview_v1", ts[:10], "true", "PENDING", import_batch_id])
        wb["RAW_IMPORTS"].append([raw_import_id, source_file_id, "preview://local/sanitized.xlsx", "preview_raw_hash", ts, import_batch_id, "DENY"])
        wb["VALIDATION_RESULTS"].append(
            [
                f"vr_preview_{uuid4().hex[:8]}",
                "WORKBOOK",
                source_file_id,
                "PREVIEW_RULE_001",
                "INFO",
                "PASS",
                "Operational preview generated from isolated local XLSX",
                ts,
                import_batch_id,
            ]
        )
        wb["CHANGELOG"].append(
            [
                f"chg_preview_{uuid4().hex[:8]}",
                ts,
                "preview_real_file_operational_output",
                OPERATIONAL_PREVIEW_SHEET,
                "Operational preview view populated from isolated XLSX (PREVIEW_ONLY)",
                "phase_9_6_preview_fix_operational_output",
                "system",
            ]
        )

        view_ws = wb[OPERATIONAL_PREVIEW_SHEET]
        cost_item_by_origin: dict[tuple[str, str], str] = {}
        for extraction, row in zip(operational_extractions, operational_rows, strict=True):
            row[2] = import_batch_id
            row[3] = budget_version_id
            view_ws.append(row)
            if extraction.should_create_cost_item:
                cost_item_id = f"ci_preview_{uuid4().hex[:8]}"
                wb["COST_ITEMS"].append(
                    [
                        cost_item_id,
                        budget_version_id,
                        source_file_id,
                        f"{row[4]}!{row[5]}",
                        row[9],
                        row[10],
                        row[11],
                        row[12],
                        row[13],
                        f"preview_row_hash_{uuid4().hex[:8]}",
                        "PENDING",
                    ]
                )
                cost_item_by_origin[(row[4], str(row[5]))] = cost_item_id

        preserved_result = _append_preserved_budget_scaffolding(
            workbook=wb,
            source_workbook=source_wb,
            source_file_id=source_file_id,
            import_batch_id=import_batch_id,
            budget_version_id=budget_version_id,
            cost_item_by_origin=cost_item_by_origin,
        )
        professional_result = append_professional_budget_review(
            workbook=wb,
            extractions=operational_extractions,
            source_file_id=source_file_id,
            import_batch_id=import_batch_id,
            budget_version_id=budget_version_id,
            mode_label="PREVIEW_ONLY",
            sheet_semantics=sheet_semantics,
        )
        wb["CHANGELOG"].append(
            [
                f"chg_preview_{uuid4().hex[:8]}",
                ts,
                "professional_budget_review_output",
                professional_result["review_sheet_name"],
                "Professional human-review budget sheet generated with separated traceability.",
                "phase_9_14_professional_budget_review_output",
                "system",
            ]
        )
        formatting_result = apply_workbook_professional_formatting(
            workbook=wb,
            mode_label="PREVIEW_ONLY",
        )
        wb["CHANGELOG"].append(
            [
                f"chg_preview_{uuid4().hex[:8]}",
                ts,
                "workbook_professional_formatting",
                formatting_result["index_sheet"],
                "Workbook-wide formatting applied across human/preserved/trace/technical sheets.",
                "phase_9_15_workbook_wide_professional_formatting",
                "system",
            ]
        )
        wb.save(output_path)
    finally:
        wb.close()
        source_wb.close()

    validate_workbook_file(output_path)
    validation_result = validate_generated_xlsx_preview(output_path, required_phase=PREVIEW_PIPELINE_PHASE)
    result["preview_rows"] = str(len(operational_rows))
    result["preview_sheet"] = OPERATIONAL_PREVIEW_SHEET
    result.update(preserved_result)
    result.update(professional_result)
    result.update(formatting_result)
    result.update(validation_result)
    return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate and harden a controlled live Excel master template (phase 9.9)."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Target .xlsx file path. Must be under outputs/live_excel_master.",
    )
    parser.add_argument(
        "--update",
        action="store_true",
        help="Allow controlled template update when output file already exists.",
    )
    parser.add_argument(
        "--synthetic-load",
        action="store_true",
        help="Append one synthetic incremental batch to the master workbook.",
    )
    parser.add_argument(
        "--rollback-from",
        type=Path,
        default=None,
        help="Restore workbook from a snapshot path, then validate and snapshot again.",
    )
    parser.add_argument(
        "--snapshot-retention-max",
        type=int,
        default=DEFAULT_RETENTION_MAX,
        help="Keep at most N snapshot files in outputs/live_excel_master/snapshots.",
    )
    parser.add_argument(
        "--run-id",
        type=str,
        default=None,
        help="Optional synthetic run_id for idempotent incremental loads.",
    )
    parser.add_argument(
        "--preview-real-file",
        type=Path,
        default=None,
        help="Generate PREVIEW_ONLY operational sheet from one isolated local XLSX file.",
    )
    parser.add_argument(
        "--preview-source-id",
        type=str,
        default="sf_preview_real_001",
        help="Sanitized source_file_id used for preview mode.",
    )
    parser.add_argument(
        "--evaluate-dry-run",
        action="store_true",
        help="Evaluate PREVIEW_ONLY workbook as dry-run candidate without promoting.",
    )
    parser.add_argument(
        "--evaluate-run-id",
        type=str,
        default="dry_run_eval",
        help="run_id label used for dry-run evaluation reporting.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    if args.snapshot_retention_max < 1:
        raise ValueError("--snapshot-retention-max must be >= 1")

    root = Path(__file__).resolve().parents[1]
    output_path = (root / args.output).resolve()
    allowed_root = (root / ALLOWED_OUTPUT_ROOT).resolve()
    ensure_allowed_output_path(output_path, allowed_root)

    if args.evaluate_dry_run:
        evaluation = evaluate_dry_run_workbook(output_path, run_id=args.evaluate_run_id)
        result = {
            "run_id": evaluation.run_id,
            "state": evaluation.state,
            "reasons": ",".join(evaluation.reasons) if evaluation.reasons else "none",
            "auto_promotion_enabled": str(evaluation.auto_promotion_enabled).lower(),
            "total_preview_rows": str(evaluation.metrics["total_preview_rows"]),
            "total_preserved_rows": str(evaluation.metrics["total_preserved_rows"]),
            "mapping_rate": str(evaluation.metrics["mapping_rate"]),
            "mapping_rate_on_candidate_cost_items": str(
                evaluation.metrics["mapping_rate_on_candidate_cost_items"]
            ),
            "traceability_rate": str(evaluation.metrics["traceability_rate"]),
            "manual_review_rate": str(evaluation.metrics["manual_review_rate"]),
            "blocked_rate": str(evaluation.metrics["blocked_rate"]),
            "amount_separation_rate": str(evaluation.metrics["amount_separation_rate"]),
            "ratio_input_rows": str(evaluation.metrics["ratio_input_rows"]),
            "ratio_calculated_rows": str(evaluation.metrics["ratio_calculated_rows"]),
        }
        print("Live Excel dry-run evaluation completed.")
    elif args.preview_real_file is not None:
        input_path = (root / args.preview_real_file).resolve() if not args.preview_real_file.is_absolute() else args.preview_real_file
        result = generate_preview_from_real_xlsx(
            input_xlsx_path=input_path,
            output_path=output_path,
            source_file_id=args.preview_source_id,
            retention_max=args.snapshot_retention_max,
        )
        print("Live Excel PREVIEW_ONLY operational output completed.")
    elif args.rollback_from is not None:
        snapshot_path = (root / args.rollback_from).resolve() if not args.rollback_from.is_absolute() else args.rollback_from
        result = rollback_master_from_snapshot(
            output_path=output_path,
            snapshot_path=snapshot_path,
            retention_max=args.snapshot_retention_max,
        )
        print("Live Excel master rollback completed.")
    elif args.synthetic_load:
        result = load_synthetic_incremental(
            output_path=output_path,
            retention_max=args.snapshot_retention_max,
            run_id=args.run_id,
        )
        print("Live Excel synthetic incremental load completed.")
    else:
        result = generate_master(
            output_path=output_path,
            update=args.update,
            retention_max=args.snapshot_retention_max,
        )
        print("Live Excel master template generation completed.")

    for key, value in result.items():
        print(f"- {key}: {value or 'n/a'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
