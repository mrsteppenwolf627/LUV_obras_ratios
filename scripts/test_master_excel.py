#!/usr/bin/env python3
"""Test master Excel generation."""

import sys
from pathlib import Path
from datetime import datetime, timezone

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.db.models import get_session
from src.export.excel_master_generator import generate_master_excel
from openpyxl import load_workbook


def test_master_excel_generation():
    """Test that master Excel is generated correctly with 36 items."""
    session = get_session()

    try:
        # Generate Excel
        excel_path = generate_master_excel(session)
        excel_file = Path(excel_path)

        print(f"✅ Excel generado: {excel_file.name}")
        assert excel_file.exists(), "Excel file not created"

        # Verify file size and format
        file_size = excel_file.stat().st_size
        print(f"   Tamaño: {file_size:,} bytes")
        assert file_size > 5000, "Excel file too small"

        # Load workbook
        wb = load_workbook(excel_file)
        print(f"✅ Sheets en el workbook: {len(wb.sheetnames)}")
        print(f"   Sheets: {', '.join(wb.sheetnames)}")

        # Verify ITEM_MASTER sheet exists
        assert "ITEM_MASTER" in wb.sheetnames, "ITEM_MASTER sheet not found"
        print("✅ Sheet ITEM_MASTER existe")

        # Check ITEM_MASTER content
        ws_items = wb["ITEM_MASTER"]
        rows = list(ws_items.iter_rows(values_only=True))
        header = rows[0]
        data_rows = rows[1:]

        print(f"✅ Columnas en ITEM_MASTER: {len(header)}")
        expected_columns = [
            "Item Key",
            "Categoría",
            "Subcategoría",
            "Unidad",
            "Unitario (Mediana)",
            "Mínimo",
            "Máximo",
            "P25",
            "P75",
            "Desv. Std",
            "Muestras (N)",
            "Solidez",
            "Última Actualización",
        ]
        print(f"   Expected: {len(expected_columns)} columns")
        assert len(header) == len(expected_columns), f"Expected {len(expected_columns)} columns, got {len(header)}"

        print(f"✅ Items en ITEM_MASTER: {len(data_rows)}")
        assert len(data_rows) == 36, f"Expected 36 items, got {len(data_rows)}"

        # Verify all items have data
        items_with_key = sum(1 for row in data_rows if row[0])  # Item Key in column 0
        print(f"   Items con Item Key: {items_with_key}")
        assert items_with_key == 36, f"Expected 36 items with key, got {items_with_key}"

        # Check solidez calculations
        solidez_col = 11  # Solidez is column 11 (0-indexed)
        solidez_values = [row[solidez_col] for row in data_rows if row[solidez_col]]
        print(f"✅ Niveles de solidez encontrados: {set(solidez_values)}")

        valid_solidez = {"MUY_DÉBIL", "DÉBIL", "SÓLIDO", "MUY_SÓLIDO"}
        for val in solidez_values:
            assert val in valid_solidez, f"Invalid solidez value: {val}"

        # Check file naming
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        expected_name = f"MASTER_{today}.xlsx"
        print(f"✅ Nombre del archivo: {excel_file.name}")
        assert excel_file.name == expected_name, f"Expected {expected_name}, got {excel_file.name}"

        # Verify other sheets still exist
        required_sheets = ["ITEM_MASTER", "INDEX", "RATIOS_SUMMARY", "CHAPTERS", "AUDIT", "RAW_DATA"]
        for sheet_name in required_sheets:
            assert sheet_name in wb.sheetnames, f"Sheet {sheet_name} not found"
        print(f"✅ Todos los sheets requeridos existen")

        print("\n" + "=" * 50)
        print("✅ VALIDACIÓN EXITOSA")
        print("=" * 50)
        print(f"Excel: {expected_name}")
        print(f"Items: 36/36")
        print(f"Sheets: 6")
        print("=" * 50)

    finally:
        session.close()


if __name__ == "__main__":
    try:
        test_master_excel_generation()
    except AssertionError as e:
        print(f"\n❌ TEST FALLIDO: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {e}", file=sys.stderr)
        sys.exit(1)
