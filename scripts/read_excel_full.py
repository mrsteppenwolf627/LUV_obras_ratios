#!/usr/bin/env python3
"""Full, non-destructive Excel reader with cell-level traceability."""

from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
import json
from pathlib import Path
import re
import sys
from typing import Any
import unicodedata

from openpyxl import load_workbook  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

SAMPLES_DIR = Path("data/samples")
REPORT_DIR = Path("reports/excel_full_reader")
JSON_REPORT = REPORT_DIR / "excel_full_reader_inventory.json"
MD_REPORT = REPORT_DIR / "excel_full_reader_inventory_report.md"

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
UNSUPPORTED_EXCEL_EXTENSIONS = {".xls", ".xlsb"}

COLUMN_PATTERNS = {
    "codigo": [r"\bcod(igo)?\b", r"\bref(erencia)?\b", r"\bpartida\b"],
    "descripcion": [r"\bdesc(ripcion)?\b", r"\bconcepto\b", r"\btexto\b"],
    "unidad": [r"\bunidad\b", r"\bud\b", r"\bu\.?\b"],
    "cantidad": [r"\bcantidad\b", r"\bmedicion\b", r"\bqty\b"],
    "precio": [r"\bprecio\b", r"\bp\.?\s*unit\b", r"\bunitario\b"],
    "importe": [r"\bimporte\b", r"\btotal\b", r"\bcoste\b", r"\bpresupuesto\b"],
    "capitulo": [r"\bcap[ii]tulo\b", r"\bcap\b"],
    "partida": [r"\bpartida\b", r"\bitem\b", r"\bitem\b"],
}

HEADER_KEYWORDS = [
    "codigo",
    "cod",
    "partida",
    "descripcion",
    "unidad",
    "ud",
    "cantidad",
    "medicion",
    "precio",
    "importe",
    "total",
    "presupuesto",
    "capitulo",
]

UNIT_TOKEN_RE = re.compile(r"\b(m2|m3|ml|ud|kg|l|cm|mm)\b", re.IGNORECASE)


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_text(text: str) -> str:
    value = unicodedata.normalize("NFKD", text or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return value.lower().strip()


def _truncate(text: str, max_len: int = 120) -> str:
    if len(text) <= max_len:
        return text
    return text[: max_len - 3] + "..."


def _sanitize_path(path: str) -> str:
    return path.replace("\\", "/").strip()


def _sheet_type(sheet: Any) -> str:
    cls = sheet.__class__.__name__.lower()
    if "chartsheet" in cls:
        return "CHARTSHEET"
    if hasattr(sheet, "iter_rows"):
        return "WORKSHEET"
    return "UNKNOWN"


def _sheet_visibility(sheet: Any) -> dict[str, Any]:
    state = getattr(sheet, "sheet_state", "visible") or "visible"
    return {
        "sheet_state": state,
        "is_hidden": state in {"hidden", "veryHidden"},
        "is_very_hidden": state == "veryHidden",
    }


def _detect_excel_file(path: Path) -> dict[str, Any]:
    ext = path.suffix.lower()
    classification = "SUPPORTED_EXCEL" if ext in SUPPORTED_EXTENSIONS else "IGNORED"
    if ext in UNSUPPORTED_EXCEL_EXTENSIONS:
        classification = "UNSUPPORTED_EXCEL"
    elif ext not in SUPPORTED_EXTENSIONS and ext not in UNSUPPORTED_EXCEL_EXTENSIONS:
        classification = "NON_EXCEL_IGNORED"
    return {
        "relative_path_sanitized": _sanitize_path(str(path)),
        "extension": ext,
        "size_bytes": path.stat().st_size,
        "classification": classification,
        "is_excel": ext in SUPPORTED_EXTENSIONS or ext in UNSUPPORTED_EXCEL_EXTENSIONS,
    }


def _scan_header_rows(rows_text: list[list[str]]) -> tuple[list[int], list[list[str]], dict[str, list[str]]]:
    scored_rows: list[tuple[int, int, int, list[str]]] = []
    max_rows = min(50, len(rows_text))

    for ridx in range(max_rows):
        row = rows_text[ridx]
        non_empty = [cell for cell in row if cell]
        if len(non_empty) < 2:
            continue

        text_cells = sum(1 for cell in non_empty if any(ch.isalpha() for ch in cell))
        kw_hits = 0
        for cell in non_empty:
            cell_n = _normalize_text(cell)
            if any(kw in cell_n for kw in HEADER_KEYWORDS):
                kw_hits += 1

        score = kw_hits * 3 + text_cells
        if kw_hits > 0 or text_cells >= 3:
            scored_rows.append((ridx + 1, score, kw_hits, non_empty))

    scored_rows.sort(key=lambda item: (item[1], item[2]), reverse=True)
    candidate_row_numbers = [item[0] for item in scored_rows[:5]]
    candidate_rows_cells = [[_truncate(cell, 80) for cell in item[3][:10]] for item in scored_rows[:5]]

    candidate_columns: dict[str, list[str]] = {k: [] for k in COLUMN_PATTERNS}
    if scored_rows:
        best_row_idx = scored_rows[0][0] - 1
        best_row = rows_text[best_row_idx]
        for cidx, cell in enumerate(best_row, start=1):
            cell_n = _normalize_text(cell)
            if not cell_n:
                continue
            for field, patterns in COLUMN_PATTERNS.items():
                if any(re.search(pattern, cell_n) for pattern in patterns):
                    label = get_column_letter(cidx)
                    if label not in candidate_columns[field]:
                        candidate_columns[field].append(label)

    return candidate_row_numbers, candidate_rows_cells, candidate_columns


def _derive_columns_without_clear_headers(
    rows_text: list[list[str]],
    formula_positions: set[tuple[int, int]],
    number_formats_map: dict[tuple[int, int], str],
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {k: [] for k in COLUMN_PATTERNS}
    if min_row == 0 or min_col == 0:
        return out

    for cidx in range(min_col, max_col + 1):
        text_count = 0
        numeric_count = 0
        formula_count = 0
        unit_hits = 0
        large_numeric_count = 0

        for ridx in range(min_row, max_row + 1):
            if ridx - 1 >= len(rows_text) or cidx - 1 >= len(rows_text[ridx - 1]):
                continue
            cell = rows_text[ridx - 1][cidx - 1]
            if not cell:
                continue
            if (ridx, cidx) in formula_positions:
                formula_count += 1

            cell_n = _normalize_text(cell)
            if UNIT_TOKEN_RE.search(cell_n):
                unit_hits += 1

            num_candidate = cell_n.replace(".", "").replace(",", "")
            if num_candidate.isdigit():
                numeric_count += 1
                try:
                    parsed = float(cell_n.replace(".", "").replace(",", ".")) if "," in cell_n else float(cell_n)
                except Exception:
                    parsed = 0.0
                if abs(parsed) >= 100:
                    large_numeric_count += 1
            else:
                text_count += 1

        label = get_column_letter(cidx)
        active_cells = text_count + numeric_count
        if active_cells == 0:
            continue

        if unit_hits > 0 and label not in out["unidad"]:
            out["unidad"].append(label)

        if formula_count > 0 and (numeric_count >= 2 or large_numeric_count >= 1) and label not in out["importe"]:
            out["importe"].append(label)

        if numeric_count >= max(2, text_count) and large_numeric_count >= 1 and label not in out["importe"]:
            out["importe"].append(label)

        fmt_hits = 0
        for ridx in range(min_row, max_row + 1):
            fmt = _normalize_text(number_formats_map.get((ridx, cidx), ""))
            if any(token in fmt for token in ["#", "0.00", "$", "€", "_-"]):
                fmt_hits += 1
        if fmt_hits >= 2 and label not in out["precio"]:
            out["precio"].append(label)

        if numeric_count >= 2 and formula_count == 0 and label not in out["cantidad"]:
            out["cantidad"].append(label)

        if text_count >= 3 and numeric_count <= 1 and label not in out["descripcion"]:
            out["descripcion"].append(label)

        if text_count >= 2 and numeric_count <= 2 and label not in out["codigo"]:
            out["codigo"].append(label)

        if text_count >= 2 and any("part" in _normalize_text(rows_text[ridx - 1][cidx - 1]) for ridx in range(min_row, max_row + 1) if ridx - 1 < len(rows_text) and cidx - 1 < len(rows_text[ridx - 1])) and label not in out["partida"]:
            out["partida"].append(label)

        if any("cap" in _normalize_text(rows_text[ridx - 1][cidx - 1]) for ridx in range(min_row, max_row + 1) if ridx - 1 < len(rows_text) and cidx - 1 < len(rows_text[ridx - 1])) and label not in out["capitulo"]:
            out["capitulo"].append(label)

    return out


def _merge_candidate_columns(primary: dict[str, list[str]], secondary: dict[str, list[str]]) -> dict[str, list[str]]:
    merged: dict[str, list[str]] = {k: [] for k in COLUMN_PATTERNS}
    for key in merged:
        for value in primary.get(key, []) + secondary.get(key, []):
            if value and value not in merged[key]:
                merged[key].append(value)
    return merged


def _build_traceability_entry(
    cell: Any,
    row_idx: int,
    col_idx: int,
    row_hidden: bool,
    col_hidden: bool,
    style_id: int | None,
) -> dict[str, Any]:
    value_text = _as_text(cell.value)
    flags: list[str] = []
    if value_text.startswith("="):
        flags.append("FORMULA")
    if cell.comment is not None:
        flags.append("COMMENT")
    if style_id not in (None, 0):
        flags.append("STYLED")
    if row_hidden:
        flags.append("ROW_HIDDEN")
    if col_hidden:
        flags.append("COLUMN_HIDDEN")

    return {
        "row": row_idx,
        "column": col_idx,
        "coordinate": cell.coordinate,
        "data_type": getattr(cell, "data_type", None),
        "value_type": type(cell.value).__name__,
        "sanitized_value": _truncate(value_text, 140),
        "formula": value_text if value_text.startswith("=") else None,
        "flags": flags,
    }


def _profile_worksheet(sheet: Any, workbook_ref: str) -> dict[str, Any]:
    max_row = getattr(sheet, "max_row", 0) or 0
    max_column = getattr(sheet, "max_column", 0) or 0
    sheet_name = sheet.title
    visibility = _sheet_visibility(sheet)

    rows_text: list[list[str]] = []
    traceability_map: list[dict[str, Any]] = []
    non_empty_rows = 0
    non_empty_cols_set: set[int] = set()
    formula_cells_count = 0
    comment_cells_count = 0
    numeric_formats: set[str] = set()
    formula_positions: set[tuple[int, int]] = set()
    number_formats_map: dict[tuple[int, int], str] = {}
    non_empty_cells = 0
    styled_cells = 0
    style_counter: Counter[int] = Counter()
    comment_samples: list[dict[str, Any]] = []
    min_used_row = 0
    max_used_row = 0
    min_used_col = 0
    max_used_col = 0
    row_counts: dict[int, int] = {}
    col_counts: dict[int, int] = {}
    hidden_rows: list[int] = []
    hidden_columns: list[str] = []

    if max_row and max_column:
        for ridx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column), start=1):
            row_values: list[str] = []
            has_data = False
            row_hidden = bool(getattr(sheet.row_dimensions.get(ridx), "hidden", False))
            if row_hidden and ridx not in hidden_rows:
                hidden_rows.append(ridx)
            for cidx, cell in enumerate(row, start=1):
                text = _as_text(cell.value)
                row_values.append(text)
                number_formats_map[(ridx, cidx)] = _as_text(cell.number_format)
                style_id = getattr(cell, "style_id", None)
                if style_id not in (None, 0):
                    styled_cells += 1
                    style_counter[int(style_id)] += 1

                col_letter = get_column_letter(cidx)
                col_hidden = bool(getattr(sheet.column_dimensions.get(col_letter), "hidden", False))
                if col_hidden and col_letter not in hidden_columns:
                    hidden_columns.append(col_letter)

                if text:
                    has_data = True
                    non_empty_cells += 1
                    non_empty_cols_set.add(cidx)
                    row_counts[ridx] = row_counts.get(ridx, 0) + 1
                    col_counts[cidx] = col_counts.get(cidx, 0) + 1
                    if min_used_row == 0 or ridx < min_used_row:
                        min_used_row = ridx
                    if ridx > max_used_row:
                        max_used_row = ridx
                    if min_used_col == 0 or cidx < min_used_col:
                        min_used_col = cidx
                    if cidx > max_used_col:
                        max_used_col = cidx

                    traceability_map.append(
                        _build_traceability_entry(
                            cell,
                            ridx,
                            cidx,
                            row_hidden=row_hidden,
                            col_hidden=col_hidden,
                            style_id=style_id,
                        )
                    )

                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells_count += 1
                    formula_positions.add((ridx, cidx))

                if cell.comment is not None:
                    comment_cells_count += 1
                    if len(comment_samples) < 10:
                        comment_samples.append(
                            {
                                "coordinate": cell.coordinate,
                                "author": _truncate(_as_text(cell.comment.author), 80),
                                "text": _truncate(_as_text(cell.comment.text), 140),
                            }
                        )

                fmt = _as_text(cell.number_format)
                if fmt and fmt not in {"General", "@"} and len(numeric_formats) < 30:
                    numeric_formats.add(fmt)

            if has_data:
                non_empty_rows += 1
            rows_text.append(row_values)
    else:
        rows_text = []

    used_cells_total = 0
    if min_used_row and min_used_col:
        used_cells_total = (max_used_row - min_used_row + 1) * (max_used_col - min_used_col + 1)
    density = (non_empty_cells / used_cells_total) if used_cells_total else 0.0

    candidate_header_rows, candidate_header_samples, headers_based_columns = _scan_header_rows(rows_text)
    inferred_columns = _derive_columns_without_clear_headers(
        rows_text,
        formula_positions,
        number_formats_map,
        min_used_row,
        max_used_row,
        min_used_col,
        max_used_col,
    )
    candidate_columns = _merge_candidate_columns(headers_based_columns, inferred_columns)

    possible_tables = []
    for table_name, table in getattr(sheet, "tables", {}).items():
        possible_tables.append(
            {
                "block_type": "named_table",
                "name": table_name,
                "ref": getattr(table, "ref", None),
                "source": "worksheet_table",
            }
        )

    candidate_table_blocks: list[dict[str, Any]] = []
    for table in possible_tables:
        candidate_table_blocks.append(table)

    if candidate_header_rows:
        for row_no in candidate_header_rows[:3]:
            candidate_table_blocks.append(
                {
                    "block_type": "heuristic_table_block",
                    "header_row": row_no,
                    "range": {
                        "min_row": row_no,
                        "max_row": max_used_row or max_row or row_no,
                        "min_column": min_used_col or 1,
                        "max_column": max_used_col or max_column or 1,
                    },
                    "confidence": "heuristic",
                    "source": "candidate_header_rows",
                }
            )
    elif non_empty_rows > 1 and len(non_empty_cols_set) > 1 and density >= 0.03 and min_used_row:
        candidate_table_blocks.append(
            {
                "block_type": "heuristic_table_block",
                "header_row": None,
                "range": {
                    "min_row": min_used_row,
                    "max_row": max_used_row or max_row or min_used_row,
                    "min_column": min_used_col or 1,
                    "max_column": max_used_col or max_column or 1,
                },
                "confidence": "low",
                "source": "used_range_density",
            }
        )

    merged_ranges = [str(rng) for rng in getattr(sheet.merged_cells, "ranges", [])]
    merged_cells_count = len(merged_ranges)
    width_samples: dict[str, float] = {}
    for cidx in range(min_used_col or 1, min((max_used_col or 1), (min_used_col or 1) + 7) + 1):
        letter = get_column_letter(cidx)
        width = getattr(sheet.column_dimensions.get(letter), "width", None)
        if isinstance(width, (int, float)):
            width_samples[letter] = float(width)

    top_rows = sorted(row_counts.items(), key=lambda item: item[1], reverse=True)[:5]
    top_cols = sorted(col_counts.items(), key=lambda item: item[1], reverse=True)[:5]
    first_non_empty_rows = []
    for ridx in sorted(row_counts.keys())[:3]:
        row = rows_text[ridx - 1]
        row_vals = [cell for cell in row if cell]
        first_non_empty_rows.append({"row": ridx, "cells": [_truncate(cell, 80) for cell in row_vals[:6]]})

    dense_rows = []
    for ridx, cnt in top_rows[:3]:
        row = rows_text[ridx - 1]
        row_vals = [cell for cell in row if cell]
        dense_rows.append(
            {
                "row": ridx,
                "non_empty_cells": cnt,
                "cells": [_truncate(cell, 80) for cell in row_vals[:6]],
            }
        )

    is_empty_sheet = non_empty_rows == 0
    is_likely_tabular = non_empty_rows > 1 and len(non_empty_cols_set) > 1 and density >= 0.03
    style_ratio = (styled_cells / non_empty_cells) if non_empty_cells else 0.0

    visual_blocks: list[dict[str, Any]] = []
    if merged_ranges:
        visual_blocks.append(
            {
                "block_type": "merged_cells",
                "count": len(merged_ranges),
                "ranges": merged_ranges[:10],
            }
        )
    if hidden_rows or hidden_columns:
        visual_blocks.append(
            {
                "block_type": "hidden_structure",
                "hidden_rows": hidden_rows[:20],
                "hidden_columns": hidden_columns[:20],
            }
        )
    if styled_cells > 0 and non_empty_cells > 0 and style_ratio > 3 and density < 0.1:
        visual_blocks.append(
            {
                "block_type": "high_format_low_data_density",
                "styled_cells": styled_cells,
                "style_vs_data_ratio": round(style_ratio, 2),
                "density_pct": round(density * 100, 2),
            }
        )

    warnings: list[str] = []
    manual_review: list[str] = []
    if formula_cells_count > 0:
        warnings.append("FORMULAS_PRESENT")
    if merged_cells_count > 0:
        warnings.append("MERGED_CELLS_PRESENT")
    if hidden_rows or hidden_columns:
        warnings.append("HIDDEN_STRUCTURE_PRESENT")
    if not candidate_header_rows:
        manual_review.append(f"NO_CLEAR_HEADERS:{sheet_name}")
    if is_empty_sheet:
        warnings.append("EMPTY_WORKSHEET_PRESENT")
        manual_review.append(f"EMPTY_WORKSHEET:{sheet_name}")
    if not is_likely_tabular:
        warnings.append("NON_TABULAR_WORKSHEET")
        manual_review.append(f"NON_TABULAR_WORKSHEET:{sheet_name}")

    budget_signals = {
        "candidate_header_rows": candidate_header_rows,
        "candidate_columns": candidate_columns,
        "signals_by_field": {
            field: {
                "columns": cols,
                "count": len(cols),
            }
            for field, cols in candidate_columns.items()
        },
    }

    comments_summary = {
        "count": comment_cells_count,
        "samples": comment_samples,
    }

    styles_summary = {
        "styled_cells": styled_cells,
        "unique_style_ids": sorted(style_counter.keys()),
        "unique_style_ids_count": len(style_counter),
        "top_style_ids": [
            {"style_id": style_id, "count": count}
            for style_id, count in style_counter.most_common(5)
        ],
        "style_vs_data_ratio": round(style_ratio, 2),
    }

    return {
        "sheet_ref": f"{workbook_ref}::{_sanitize_path(sheet_name)}",
        "workbook_ref": workbook_ref,
        "sheet_name_sanitized": _sanitize_path(sheet_name),
        "sheet_type": "WORKSHEET",
        "used_range": {
            "min_row": min_used_row or None,
            "max_row": max_used_row or None,
            "min_column": min_used_col or None,
            "max_column": max_used_col or None,
        },
        "dimensions": {
            "max_row": max_row,
            "max_column": max_column,
        },
        "visibility": visibility,
        "merged_cells_summary": {
            "count": merged_cells_count,
            "ranges": merged_ranges[:10],
        },
        "formulas_summary": {
            "count": formula_cells_count,
            "sample_cells": [
                item["coordinate"]
                for item in traceability_map
                if item["formula"] is not None
            ][:10],
        },
        "comments_summary": comments_summary,
        "styles_summary": styles_summary,
        "density_profile": {
            "non_empty_cells": non_empty_cells,
            "used_cells_total": used_cells_total,
            "non_empty_pct": round(density * 100, 2),
            "top_non_empty_rows": [{"row": row, "count": count} for row, count in top_rows],
            "top_non_empty_columns": [{"column": get_column_letter(col), "count": count} for col, count in top_cols],
        },
        "candidate_header_rows": candidate_header_rows,
        "candidate_columns": candidate_columns,
        "candidate_table_blocks": candidate_table_blocks,
        "visual_blocks": visual_blocks,
        "budget_signals": budget_signals,
        "traceability_map": traceability_map,
        "cell_samples_sanitized": {
            "first_non_empty_rows": first_non_empty_rows,
            "dense_rows": dense_rows,
            "possible_header_rows": [
                {"row": row_no, "cells": cells}
                for row_no, cells in zip(candidate_header_rows, candidate_header_samples)
            ],
        },
        "warnings": sorted(set(warnings)),
        "manual_review": sorted(set(manual_review)),
        "is_empty_sheet": is_empty_sheet,
        "is_likely_tabular": is_likely_tabular,
        "unknown_or_unsupported": [],
    }


def _profile_chartsheet(sheet: Any, workbook_ref: str) -> dict[str, Any]:
    sheet_name = _sanitize_path(sheet.title)
    warnings = ["CHARTSHEET_PRESENT"]
    manual_review = [f"CHARTSHEET_PRESENT:{sheet_name}"]
    return {
        "sheet_ref": f"{workbook_ref}::{sheet_name}",
        "workbook_ref": workbook_ref,
        "sheet_name_sanitized": sheet_name,
        "sheet_type": "CHARTSHEET",
        "used_range": {
            "min_row": None,
            "max_row": None,
            "min_column": None,
            "max_column": None,
        },
        "dimensions": {
            "max_row": None,
            "max_column": None,
        },
        "visibility": _sheet_visibility(sheet),
        "merged_cells_summary": {
            "count": 0,
            "ranges": [],
        },
        "formulas_summary": {
            "count": 0,
            "sample_cells": [],
        },
        "comments_summary": {
            "count": 0,
            "samples": [],
        },
        "styles_summary": {
            "styled_cells": 0,
            "unique_style_ids": [],
            "unique_style_ids_count": 0,
            "top_style_ids": [],
            "style_vs_data_ratio": 0.0,
        },
        "density_profile": {
            "non_empty_cells": 0,
            "used_cells_total": 0,
            "non_empty_pct": 0.0,
            "top_non_empty_rows": [],
            "top_non_empty_columns": [],
        },
        "candidate_header_rows": [],
        "candidate_columns": {k: [] for k in COLUMN_PATTERNS},
        "candidate_table_blocks": [],
        "visual_blocks": [
            {
                "block_type": "chartsheet_context",
                "note": "non_tabular_context",
            }
        ],
        "budget_signals": {
            "candidate_header_rows": [],
            "candidate_columns": {k: [] for k in COLUMN_PATTERNS},
            "signals_by_field": {k: {"columns": [], "count": 0} for k in COLUMN_PATTERNS},
        },
        "traceability_map": [],
        "cell_samples_sanitized": {
            "first_non_empty_rows": [],
            "dense_rows": [],
            "possible_header_rows": [],
        },
        "warnings": warnings,
        "manual_review": manual_review,
        "is_empty_sheet": True,
        "is_likely_tabular": False,
        "unknown_or_unsupported": [],
    }


def _build_error_workbook_summary(workbook_ref: str, relative_path: str, ext: str, error: str) -> dict[str, Any]:
    return {
        "workbook_ref": workbook_ref,
        "relative_path_sanitized": relative_path,
        "extension": ext,
        "sheet_count": 0,
        "worksheet_count": 0,
        "chartsheet_count": 0,
        "readable": False,
        "errors": [error],
        "warnings": [],
        "manual_review": [error],
        "risks": [error],
    }


def _build_unsupported_workbook_summary(workbook_ref: str, relative_path: str, ext: str, reason: str) -> dict[str, Any]:
    return {
        "workbook_ref": workbook_ref,
        "relative_path_sanitized": relative_path,
        "extension": ext,
        "sheet_count": 0,
        "worksheet_count": 0,
        "chartsheet_count": 0,
        "readable": False,
        "errors": [reason],
        "warnings": [reason],
        "manual_review": [reason],
        "risks": [reason],
    }


def analyze_excel_sources(root: Path) -> dict[str, Any]:
    samples_path = root / SAMPLES_DIR
    generated_at = datetime.now(timezone.utc).isoformat()

    if not samples_path.exists():
        return {
            "reader_metadata": {
                "generated_at": generated_at,
                "reader": "read_excel_full",
                "samples_dir": _sanitize_path(str(SAMPLES_DIR)),
                "supported_extensions": sorted(SUPPORTED_EXTENSIONS),
                "ignored_extensions": sorted(UNSUPPORTED_EXCEL_EXTENSIONS),
            },
            "source_files": [],
            "workbook_summaries": [],
            "sheets": [],
            "global_summary": {
                "files_total": 0,
                "source_files_total": 0,
                "excel_files_detected": 0,
                "workbooks_readable": 0,
                "workbooks_unreadable": 0,
                "worksheet_total": 0,
                "chartsheet_total": 0,
                "empty_sheets_total": 0,
                "non_tabular_sheets_total": 0,
                "traced_cells_total": 0,
                "controlled_exclusions_total": 0,
            },
            "risks": [],
            "warnings": [],
            "manual_review": [],
            "controlled_exclusions": [],
        }

    all_files = sorted([p for p in samples_path.rglob("*") if p.is_file()])
    source_files: list[dict[str, Any]] = []
    controlled_exclusions: list[dict[str, Any]] = []
    workbook_summaries: list[dict[str, Any]] = []
    sheet_entries: list[dict[str, Any]] = []
    risks: set[str] = set()
    warnings: set[str] = set()
    manual_review: set[str] = set()
    worksheet_total = 0
    chartsheet_total = 0
    empty_sheets_total = 0
    non_tabular_sheets_total = 0
    traced_cells_total = 0
    readable_workbooks = 0
    unreadable_workbooks = 0

    for path in all_files:
        ext = path.suffix.lower()
        rel = _sanitize_path(str(path.relative_to(root)))
        file_entry = _detect_excel_file(path)
        file_entry["relative_path_sanitized"] = rel
        source_files.append(file_entry)

        if ext not in SUPPORTED_EXTENSIONS and ext not in UNSUPPORTED_EXCEL_EXTENSIONS:
            exclusion = {
                "relative_path_sanitized": rel,
                "extension": ext,
                "reason": "NON_EXCEL_IGNORED",
            }
            controlled_exclusions.append(exclusion)
            continue

        workbook_ref = rel
        if ext in UNSUPPORTED_EXCEL_EXTENSIONS:
            reason = "UNSUPPORTED_EXCEL_FORMAT"
            summary = _build_unsupported_workbook_summary(workbook_ref, rel, ext, reason)
            workbook_summaries.append(summary)
            controlled_exclusions.append(
                {
                    "relative_path_sanitized": rel,
                    "extension": ext,
                    "reason": reason,
                }
            )
            warnings.add(reason)
            manual_review.add(reason)
            continue

        try:
            wb = load_workbook(path, read_only=False, data_only=False)
        except Exception as exc:
            unreadable_workbooks += 1
            error = f"WORKBOOK_OPEN_ERROR:{exc}"
            summary = _build_error_workbook_summary(workbook_ref, rel, ext, error)
            workbook_summaries.append(summary)
            risks.add("WORKBOOK_OPEN_ERROR")
            warnings.add("WORKBOOK_OPEN_ERROR")
            manual_review.add(error)
            continue

        sheet_objects = list(getattr(wb, "worksheets", [])) + list(getattr(wb, "chartsheets", []))
        sheet_count = len(getattr(wb, "sheetnames", sheet_objects))
        worksheet_count = len(getattr(wb, "worksheets", []))
        chartsheet_count = len(getattr(wb, "chartsheets", []))
        workbook_readable = True
        workbook_errors: list[str] = []
        workbook_warnings: set[str] = set()
        workbook_manual_review: set[str] = set()
        workbook_risks: set[str] = set()

        for sheet in sheet_objects:
            stype = _sheet_type(sheet)
            if stype == "WORKSHEET":
                sheet_entry = _profile_worksheet(sheet, workbook_ref)
                worksheet_total += 1
                traced_cells_total += len(sheet_entry.get("traceability_map", []))
                if sheet_entry.get("is_empty_sheet"):
                    empty_sheets_total += 1
                if not sheet_entry.get("is_likely_tabular"):
                    non_tabular_sheets_total += 1
            elif stype == "CHARTSHEET":
                sheet_entry = _profile_chartsheet(sheet, workbook_ref)
                chartsheet_total += 1
                non_tabular_sheets_total += 1
            else:
                sheet_entry = {
                    "sheet_ref": f"{workbook_ref}::{_sanitize_path(getattr(sheet, 'title', 'UNKNOWN'))}",
                    "workbook_ref": workbook_ref,
                    "sheet_name_sanitized": _sanitize_path(getattr(sheet, "title", "UNKNOWN")),
                    "sheet_type": stype,
                    "used_range": {
                        "min_row": None,
                        "max_row": None,
                        "min_column": None,
                        "max_column": None,
                    },
                    "dimensions": {
                        "max_row": None,
                        "max_column": None,
                    },
                    "visibility": {"sheet_state": "unknown", "is_hidden": False, "is_very_hidden": False},
                    "merged_cells_summary": {"count": 0, "ranges": []},
                    "formulas_summary": {"count": 0, "sample_cells": []},
                    "comments_summary": {"count": 0, "samples": []},
                    "styles_summary": {
                        "styled_cells": 0,
                        "unique_style_ids": [],
                        "unique_style_ids_count": 0,
                        "top_style_ids": [],
                        "style_vs_data_ratio": 0.0,
                    },
                    "density_profile": {
                        "non_empty_cells": 0,
                        "used_cells_total": 0,
                        "non_empty_pct": 0.0,
                        "top_non_empty_rows": [],
                        "top_non_empty_columns": [],
                    },
                    "candidate_header_rows": [],
                    "candidate_columns": {k: [] for k in COLUMN_PATTERNS},
                    "candidate_table_blocks": [],
                    "visual_blocks": [],
                    "budget_signals": {
                        "candidate_header_rows": [],
                        "candidate_columns": {k: [] for k in COLUMN_PATTERNS},
                        "signals_by_field": {k: {"columns": [], "count": 0} for k in COLUMN_PATTERNS},
                    },
                    "traceability_map": [],
                    "cell_samples_sanitized": {
                        "first_non_empty_rows": [],
                        "dense_rows": [],
                        "possible_header_rows": [],
                    },
                    "warnings": ["UNKNOWN_SHEET_TYPE"],
                    "manual_review": [f"UNKNOWN_SHEET_TYPE:{getattr(sheet, 'title', 'UNKNOWN')}"],
                    "is_empty_sheet": True,
                    "is_likely_tabular": False,
                    "unknown_or_unsupported": [],
                }
                non_tabular_sheets_total += 1

            sheet_entries.append(sheet_entry)
            workbook_warnings.update(sheet_entry.get("warnings", []))
            workbook_manual_review.update(sheet_entry.get("manual_review", []))
            workbook_risks.update(sheet_entry.get("warnings", []))
            risks.update(sheet_entry.get("warnings", []))
            warnings.update(sheet_entry.get("warnings", []))
            manual_review.update(sheet_entry.get("manual_review", []))

        wb.close()

        workbook_summary = {
            "workbook_ref": workbook_ref,
            "relative_path_sanitized": rel,
            "extension": ext,
            "sheet_count": sheet_count,
            "worksheet_count": worksheet_count,
            "chartsheet_count": chartsheet_count,
            "readable": workbook_readable,
            "errors": workbook_errors,
            "warnings": sorted(workbook_warnings),
            "manual_review": sorted(workbook_manual_review),
            "risks": sorted(workbook_risks),
        }
        workbook_summaries.append(workbook_summary)
        readable_workbooks += 1
        warnings.update(workbook_warnings)
        manual_review.update(workbook_manual_review)
        risks.update(workbook_risks)

    payload = {
        "reader_metadata": {
            "generated_at": generated_at,
            "reader": "read_excel_full",
            "samples_dir": _sanitize_path(str(SAMPLES_DIR)),
            "supported_extensions": sorted(SUPPORTED_EXTENSIONS),
            "ignored_extensions": sorted(UNSUPPORTED_EXCEL_EXTENSIONS),
        },
        "source_files": source_files,
        "workbook_summaries": workbook_summaries,
        "sheets": sheet_entries,
        "global_summary": {
            "files_total": len(all_files),
            "source_files_total": len(source_files),
            "excel_files_detected": sum(1 for item in source_files if item.get("classification") == "SUPPORTED_EXCEL"),
            "workbooks_readable": readable_workbooks,
            "workbooks_unreadable": unreadable_workbooks,
            "worksheet_total": worksheet_total,
            "chartsheet_total": chartsheet_total,
            "empty_sheets_total": empty_sheets_total,
            "non_tabular_sheets_total": non_tabular_sheets_total,
            "traced_cells_total": traced_cells_total,
            "controlled_exclusions_total": len(controlled_exclusions),
        },
        "risks": sorted(risks),
        "warnings": sorted(warnings),
        "manual_review": sorted(manual_review),
        "controlled_exclusions": controlled_exclusions,
    }
    return payload


def write_reports(root: Path, payload: dict[str, Any]) -> tuple[Path, Path]:
    json_path = root / JSON_REPORT
    md_path = root / MD_REPORT
    json_path.parent.mkdir(parents=True, exist_ok=True)

    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    lines = [
        "# Excel Full Reader",
        "",
        "> Local report. Real-data output may be sensitive and must remain outside Git.",
        "",
        f"- Generated at (UTC): {payload.get('reader_metadata', {}).get('generated_at')}",
        f"- Samples dir: {payload.get('reader_metadata', {}).get('samples_dir')}",
        f"- Files total: {payload.get('global_summary', {}).get('files_total', 0)}",
        f"- Excel files detected: {payload.get('global_summary', {}).get('excel_files_detected', 0)}",
        f"- Workbooks readable: {payload.get('global_summary', {}).get('workbooks_readable', 0)}",
        f"- Worksheets total: {payload.get('global_summary', {}).get('worksheet_total', 0)}",
        f"- Chartsheets total: {payload.get('global_summary', {}).get('chartsheet_total', 0)}",
        "",
        "## Global warnings and review",
        "",
        f"- Risks: {', '.join(payload.get('risks', [])) or 'none'}",
        f"- Warnings: {', '.join(payload.get('warnings', [])) or 'none'}",
        f"- Manual review: {', '.join(payload.get('manual_review', [])) or 'none'}",
        "",
        "## Workbooks",
        "",
    ]

    for workbook in payload.get("workbook_summaries", []):
        lines.append(f"### {workbook.get('relative_path_sanitized')}")
        lines.append(f"- Readable: {workbook.get('readable')}")
        lines.append(f"- Sheets: {workbook.get('sheet_count')}")
        lines.append(f"- Worksheets: {workbook.get('worksheet_count')}")
        lines.append(f"- Chartsheets: {workbook.get('chartsheet_count')}")
        lines.append(f"- Warnings: {', '.join(workbook.get('warnings', [])) or 'none'}")
        lines.append(f"- Manual review: {', '.join(workbook.get('manual_review', [])) or 'none'}")
        lines.append("")

    lines.extend(
        [
            "## Sheets",
            "",
        ]
    )

    for sheet in payload.get("sheets", []):
        lines.append(f"### {sheet.get('sheet_ref')}")
        lines.append(f"- Type: {sheet.get('sheet_type')}")
        lines.append(f"- Visibility: {sheet.get('visibility', {}).get('sheet_state')}")
        lines.append(f"- Used range: {sheet.get('used_range')}")
        lines.append(f"- Candidate headers: {sheet.get('candidate_header_rows', [])}")
        lines.append(f"- Candidate columns: {sheet.get('candidate_columns', {})}")
        lines.append(f"- Warnings: {', '.join(sheet.get('warnings', [])) or 'none'}")
        lines.append(f"- Manual review: {', '.join(sheet.get('manual_review', [])) or 'none'}")
        lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")
    return json_path, md_path


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    payload = analyze_excel_sources(root)
    json_path, md_path = write_reports(root, payload)

    print("Excel full reader summary")
    print(f"- Samples dir: {payload.get('reader_metadata', {}).get('samples_dir')}")
    print(f"- Files total: {payload.get('global_summary', {}).get('files_total', 0)}")
    print(f"- Excel files detected: {payload.get('global_summary', {}).get('excel_files_detected', 0)}")
    print(f"- Workbooks readable: {payload.get('global_summary', {}).get('workbooks_readable', 0)}")
    print(f"- Worksheets total: {payload.get('global_summary', {}).get('worksheet_total', 0)}")
    print(f"- Chartsheets total: {payload.get('global_summary', {}).get('chartsheet_total', 0)}")
    print(f"- JSON report: {json_path.relative_to(root).as_posix()}")
    print(f"- Markdown report: {md_path.relative_to(root).as_posix()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
