"""Workbook-wide professional formatting for PREVIEW_ONLY outputs."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class SheetMeta:
    name: str
    category: str
    description: str


THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
HEADER_FONT = Font(name="Calibri", size=10, bold=True)
BODY_FONT = Font(name="Calibri", size=10, bold=False)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="EDEDED")
INDEX_TITLE_FONT = Font(name="Calibri", size=14, bold=True)
CATEGORY_TAB_COLORS = {
    "INDEX": "1F4E78",
    "REVIEW": "2F75B5",
    "PRESERVED": "548235",
    "TRACE": "BF9000",
    "TECHNICAL": "7F7F7F",
}
_PENDING_TECH_SHEETS = {
    "NORMALIZED_COST_ITEMS",
    "CATEGORY_MAPPING",
    "RATIO_INPUTS",
    "RATIOS_CALCULATED",
}


def _utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")


def _starts_with_any(name: str, prefixes: tuple[str, ...]) -> bool:
    return any(name.startswith(prefix) for prefix in prefixes)


def _sheet_category(name: str) -> str:
    if name == "INDEX":
        return "INDEX"
    if _starts_with_any(name, ("BUDGET_REVIEW_",)):
        if _starts_with_any(name, ("BUDGET_REVIEW_TRACE_",)):
            return "TRACE"
        return "REVIEW"
    if _starts_with_any(name, ("PRES_",)):
        return "PRESERVED"
    if name in {"PRESERVED_BUDGETS_INDEX", "PRESERVED_BUDGET_SHEETS", "IMPORTED_BUDGET_VIEW"}:
        return "PRESERVED"
    if name in {"PRESERVED_TO_COST_ITEMS_MAP", "VALIDATION_RESULTS"}:
        return "TRACE"
    return "TECHNICAL"


def _sheet_description(name: str, category: str) -> str:
    if name == "INDEX":
        return "Navegacion principal del workbook."
    if category == "REVIEW":
        return "Hoja de revision humana del presupuesto."
    if category == "PRESERVED":
        return "Capa preservada equivalente al input."
    if category == "TRACE":
        return "Trazabilidad y control tecnico de mapeo/validacion."
    return "Hoja tecnica interna del sistema."


def _reorder_sheets(workbook: object) -> None:
    def rank(sheet_name: str) -> tuple[int, str]:
        if sheet_name == "INDEX":
            return (0, sheet_name)
        if sheet_name.startswith("BUDGET_REVIEW_") and not sheet_name.startswith("BUDGET_REVIEW_TRACE_"):
            return (1, sheet_name)
        if sheet_name.startswith("BUDGET_REVIEW_TRACE_"):
            return (2, sheet_name)
        if sheet_name.startswith("PRES_"):
            return (3, sheet_name)
        if sheet_name == "IMPORTED_BUDGET_VIEW":
            return (4, sheet_name)
        if sheet_name in {"PRESERVED_BUDGETS_INDEX", "PRESERVED_BUDGET_SHEETS"}:
            return (5, sheet_name)
        if sheet_name == "PRESERVED_TO_COST_ITEMS_MAP":
            return (6, sheet_name)
        if sheet_name == "COST_ITEMS":
            return (7, sheet_name)
        if sheet_name == "VALIDATION_RESULTS":
            return (8, sheet_name)
        return (9, sheet_name)

    ordered = sorted(workbook.worksheets, key=lambda ws: rank(ws.title))
    workbook._sheets = ordered


def _set_initial_open_sheet(workbook: object, preferred_name: str = "INDEX") -> None:
    if preferred_name not in workbook.sheetnames:
        return
    target_index = workbook.sheetnames.index(preferred_name)
    workbook.active = target_index
    for ws in workbook.worksheets:
        ws.sheet_view.tabSelected = False
    workbook[preferred_name].sheet_view.tabSelected = True
    if workbook.views:
        workbook.views[0].activeTab = target_index
        workbook.views[0].firstSheet = 0


def _autowidth(ws: object, max_rows: int = 120) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(ws.max_row, max_rows) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            text = "" if value is None else str(value)
            if len(text) > max_len:
                max_len = len(text)
        width = min(max(10, max_len + 2), 56)
        ws.column_dimensions[letter].width = width


def _style_header(ws: object, header_row: int) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_body(ws: object, start_row: int) -> None:
    for row_idx in range(start_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.font is None or not cell.font.bold:
                cell.font = BODY_FONT
            if cell.alignment is None:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = THIN_BORDER


def _apply_tabular_defaults(ws: object, header_row: int = 1) -> None:
    _style_header(ws, header_row=header_row)
    _style_body(ws, start_row=header_row + 1)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{header_row}"
    _autowidth(ws)


def _hide_preserved_trace_columns(ws: object) -> None:
    technical_prefixes = ("__source_",)
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col_idx).value or "").strip()
        if any(header.startswith(prefix) for prefix in technical_prefixes):
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True


def _mark_pending_sheet_note(ws: object) -> None:
    if ws.max_row > 1:
        return
    header = ws.cell(row=1, column=1)
    if header.comment is None:
        header.comment = Comment("PREVIEW_ONLY: hoja pendiente de alimentacion en fase posterior.", "system")


def _apply_review_sheet_style(ws: object) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:4"
    ws.sheet_properties.tabColor = CATEGORY_TAB_COLORS["REVIEW"]


def _apply_index_sheet(workbook: object, mode_label: str) -> None:
    if "INDEX" in workbook.sheetnames:
        ws = workbook["INDEX"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = workbook.create_sheet("INDEX", 0)
    ws["A1"] = "LUV Obras Ratios - Navegacion del Workbook"
    ws["A1"].font = INDEX_TITLE_FONT
    ws["A2"] = (
        f"Modo: {mode_label} | Estado: no operativo | Ratios finales: no calculados | "
        f"Actualizado: {_utc_now()}"
    )
    ws["A3"] = "Empiece por BUDGET_REVIEW_001. No edite hojas tecnicas manualmente."
    ws["A4"] = "Hoja"
    ws["B4"] = "Categoria"
    ws["C4"] = "Descripcion"
    _style_header(ws, header_row=4)

    row = 5
    for name in workbook.sheetnames:
        if name == "INDEX":
            continue
        category = _sheet_category(name)
        link_cell = ws.cell(row=row, column=1, value=name)
        safe_name = name.replace("'", "''")
        link_cell.hyperlink = f"#'{safe_name}'!A1"
        link_cell.style = "Hyperlink"
        ws.cell(row=row, column=2, value=category.lower())
        ws.cell(row=row, column=3, value=_sheet_description(name, category))
        row += 1
    _style_body(ws, start_row=5)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:C4"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = CATEGORY_TAB_COLORS["INDEX"]


def _collect_sheet_meta(workbook: object) -> list[SheetMeta]:
    metas: list[SheetMeta] = []
    for name in workbook.sheetnames:
        category = _sheet_category(name)
        metas.append(SheetMeta(name=name, category=category, description=_sheet_description(name, category)))
    return metas


def apply_workbook_professional_formatting(workbook: object, mode_label: str = "PREVIEW_ONLY") -> dict[str, str]:
    """Apply global workbook presentation without mutating business data."""
    _apply_index_sheet(workbook, mode_label=mode_label)
    _reorder_sheets(workbook)
    _set_initial_open_sheet(workbook, preferred_name="INDEX")
    metas = _collect_sheet_meta(workbook)

    for meta in metas:
        ws = workbook[meta.name]
        if meta.category == "INDEX":
            continue
        if meta.category == "REVIEW":
            _apply_review_sheet_style(ws)
            # Review sheet already has custom header on row 4.
            _style_header(ws, header_row=4)
            _style_body(ws, start_row=5)
            continue
        if meta.category == "PRESERVED":
            _apply_tabular_defaults(ws, header_row=1)
            if meta.name.startswith("PRES_"):
                _hide_preserved_trace_columns(ws)
            ws.sheet_properties.tabColor = CATEGORY_TAB_COLORS["PRESERVED"]
            continue
        if meta.category == "TRACE":
            _apply_tabular_defaults(ws, header_row=1)
            ws.sheet_properties.tabColor = CATEGORY_TAB_COLORS["TRACE"]
            continue

        _apply_tabular_defaults(ws, header_row=1)
        if meta.name in _PENDING_TECH_SHEETS:
            _mark_pending_sheet_note(ws)
        ws.sheet_properties.tabColor = CATEGORY_TAB_COLORS["TECHNICAL"]

    return {
        "index_sheet": "INDEX",
        "sheet_count": str(len(workbook.sheetnames)),
    }
