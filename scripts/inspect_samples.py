#!/usr/bin/env python3
"""Inspect sample files in data/samples without importing or modifying data."""

from __future__ import annotations

from datetime import datetime, timezone
import json
from pathlib import Path
import re
import sys
from typing import Any

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - availability depends on environment
    openpyxl = None

SAMPLES_DIR = Path("data/samples")
JSON_REPORT = Path("reports/sample_inspections/sample_inventory.json")
MD_REPORT = Path("reports/sample_inspections/sample_inventory_report.md")
SANITIZED_JSON_REPORT = Path("reports/sample_inspections/sample_inventory_sanitized.json")
SANITIZED_MD_REPORT = Path("reports/sample_inspections/sample_inventory_sanitized_report.md")

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
BACKUP_HINTS = ("_bkp", "backup", "copia", "old", "antiguo")
VERSION_HINTS = ("final", "actualizado", "contrato", "mediciones", "fase")
VERSION_PATTERN = re.compile(r"\bv[1-9]\d*\b", re.IGNORECASE)


CLASS_BY_EXTENSION = {
    ".xlsx": "EXCEL",
    ".xlsm": "EXCEL",
    ".xls": "EXCEL",
    ".bc3": "BC3",
    ".pdf": "PDF",
    ".presto": "PRESTO",
    ".pzh": "PZH",
}

IGNORED_FILENAMES = {".gitkeep"}


def classify_extension(ext: str) -> str:
    return CLASS_BY_EXTENSION.get(ext.lower(), "OTHER")


def has_any_hint(text: str, hints: tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(hint in lowered for hint in hints)


def detect_encoding(raw: bytes) -> str:
    candidates = ("utf-8", "cp1252", "latin-1")
    for encoding in candidates:
        try:
            raw.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "binary_or_unknown"


def inspect_excel(path: Path) -> dict[str, Any]:
    if openpyxl is None:
        return {
            "status": "EXCEL_INSPECTION_SKIPPED_OPENPYXL_NOT_AVAILABLE",
            "sheet_names": [],
            "sheet_dimensions": [],
            "preview_headers": {},
            "notes": "openpyxl is not available in current environment.",
        }

    if path.suffix.lower() == ".xls":
        return {
            "status": "EXCEL_INSPECTION_SKIPPED_UNSUPPORTED_LEGACY_XLS",
            "sheet_names": [],
            "sheet_dimensions": [],
            "preview_headers": {},
            "notes": "openpyxl does not support legacy .xls files.",
        }

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return {
            "status": "EXCEL_INSPECTION_FAILED",
            "sheet_names": [],
            "sheet_dimensions": [],
            "preview_headers": {},
            "notes": f"Could not open workbook: {exc}",
        }

    sheet_dimensions: list[dict[str, Any]] = []
    sheet_types: list[dict[str, str]] = []
    headers: dict[str, list[str]] = {}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if hasattr(ws, "iter_rows"):
            sheet_type = "WORKSHEET"
        elif ws.__class__.__name__.lower() == "readonlychartsheet" or ws.__class__.__name__.lower().endswith("chartsheet"):
            sheet_type = "CHARTSHEET"
        else:
            sheet_type = "UNKNOWN_SHEET_TYPE"
        max_row = getattr(ws, "max_row", None)
        max_column = getattr(ws, "max_column", None)
        sheet_dimensions.append(
            {
                "sheet": sheet_name,
                "max_row": max_row,
                "max_column": max_column,
            }
        )
        sheet_types.append({"sheet": sheet_name, "sheet_type": sheet_type})

        preview: list[str] = []
        if sheet_type == "WORKSHEET":
            rows = ws.iter_rows(min_row=1, max_row=3, values_only=True)
            for row in rows:
                for cell in row:
                    if isinstance(cell, str):
                        stripped = cell.strip()
                        if stripped:
                            preview.append(stripped)
                    elif cell is not None:
                        preview.append(str(cell))
                    if len(preview) >= 10:
                        break
                if len(preview) >= 10:
                    break
        headers[sheet_name] = preview

    workbook.close()
    return {
        "status": "EXCEL_INSPECTED_BASIC",
        "sheet_names": workbook.sheetnames,
        "sheet_types": sheet_types,
        "sheet_dimensions": sheet_dimensions,
        "preview_headers": headers,
        "notes": "Basic non-destructive inspection completed.",
    }


def inspect_bc3(path: Path) -> dict[str, Any]:
    try:
        raw = path.read_bytes()
    except Exception as exc:
        return {
            "status": "BC3_READ_FAILED",
            "is_text_like": False,
            "detected_encoding": "unknown",
            "line_sample": [],
            "notes": f"Could not read file: {exc}",
        }

    sample = raw[:8192]
    encoding = detect_encoding(sample)
    is_text_like = encoding != "binary_or_unknown"

    lines: list[str] = []
    if is_text_like:
        decoded = sample.decode(encoding, errors="replace")
        for line in decoded.splitlines()[:20]:
            lines.append(line[:240])

    return {
        "status": "BC3_INSPECTED_SUPERFICIAL",
        "is_text_like": is_text_like,
        "detected_encoding": encoding,
        "line_sample": lines,
        "notes": "Sampled initial bytes only. No parser logic applied.",
    }


def inspect_pdf() -> dict[str, Any]:
    return {
        "status": "REFERENCE_ONLY_DIAGNOSTIC",
        "notes": "PDF registered only as reference. No OCR or economic extraction.",
    }


def inspect_presto_or_pzh() -> dict[str, Any]:
    return {
        "status": "NEEDS_FORMAT_RESEARCH",
        "notes": "Format retained for traceability. Interpretation deferred.",
    }


def inspect_other() -> dict[str, Any]:
    return {
        "status": "OTHER_FORMAT_RECORDED",
        "notes": "Unrecognized extension captured for inventory.",
    }


def inspect_samples(root: Path) -> dict[str, Any]:
    samples_path = root / SAMPLES_DIR
    generated_at = datetime.now(timezone.utc).isoformat()

    if not samples_path.exists():
        return {
            "generated_at": generated_at,
            "samples_dir": str(SAMPLES_DIR).replace("\\", "/"),
            "exists": False,
            "files_count_total": 0,
            "sample_files_count": 0,
            "ignored_files_count": 0,
            "files_count": 0,
            "message": "Samples directory does not exist.",
            "files": [],
            "counts_by_class_total": {},
            "counts_by_class": {},
        }

    files = sorted([p for p in samples_path.rglob("*") if p.is_file()])
    entries: list[dict[str, Any]] = []
    counts_total: dict[str, int] = {}
    counts_sample: dict[str, int] = {}
    ignored_count = 0

    for file_path in files:
        rel = file_path.relative_to(root)
        rel_str = str(rel).replace("\\", "/")
        rel_lower = rel_str.lower()
        ext = file_path.suffix.lower()
        is_ignored = file_path.name.lower() in IGNORED_FILENAMES
        file_class = classify_extension(ext)
        counts_total[file_class] = counts_total.get(file_class, 0) + 1

        stat = file_path.stat()
        backup_hint = has_any_hint(rel_lower, BACKUP_HINTS)
        version_hint = has_any_hint(rel_lower, VERSION_HINTS) or bool(VERSION_PATTERN.search(rel_lower))

        entry: dict[str, Any] = {
            "relative_path": rel_str,
            "filename": file_path.name,
            "extension": ext,
            "size_bytes": stat.st_size,
            "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
            "classification": file_class,
            "backup_hint": backup_hint,
            "version_or_phase_hint": version_hint,
            "is_ignored": is_ignored,
        }

        if is_ignored:
            ignored_count += 1
            entry["ignored_reason"] = "SUPPORT_FILE"
            entry["inspection"] = {
                "status": "IGNORED_SUPPORT_FILE",
                "notes": "Support file excluded from operational sample counts.",
            }
        elif file_class == "EXCEL":
            counts_sample[file_class] = counts_sample.get(file_class, 0) + 1
            entry["inspection"] = inspect_excel(file_path)
        elif file_class == "BC3":
            counts_sample[file_class] = counts_sample.get(file_class, 0) + 1
            entry["inspection"] = inspect_bc3(file_path)
        elif file_class == "PDF":
            counts_sample[file_class] = counts_sample.get(file_class, 0) + 1
            entry["inspection"] = inspect_pdf()
        elif file_class in {"PRESTO", "PZH"}:
            counts_sample[file_class] = counts_sample.get(file_class, 0) + 1
            entry["inspection"] = inspect_presto_or_pzh()
        else:
            counts_sample[file_class] = counts_sample.get(file_class, 0) + 1
            entry["inspection"] = inspect_other()

        entries.append(entry)

    sample_files_count = len(entries) - ignored_count
    message = "OK" if sample_files_count > 0 else "No sample files found in data/samples."
    return {
        "generated_at": generated_at,
        "samples_dir": str(SAMPLES_DIR).replace("\\", "/"),
        "exists": True,
        "files_count_total": len(entries),
        "sample_files_count": sample_files_count,
        "ignored_files_count": ignored_count,
        "files_count": sample_files_count,
        "message": message,
        "files": entries,
        "counts_by_class_total": counts_total,
        "counts_by_class": counts_sample,
    }


def load_duplicates_by_hash(root: Path) -> list[dict[str, Any]]:
    hash_report = root / Path("reports/sample_inspections/file_hashes.json")
    if not hash_report.exists():
        return []
    try:
        payload = json.loads(hash_report.read_text(encoding="utf-8"))
    except Exception:
        return []
    return payload.get("duplicates_by_hash", [])


def build_sanitized_inventory(inventory: dict[str, Any], duplicates_by_hash: list[dict[str, Any]]) -> dict[str, Any]:
    sanitized_files: list[dict[str, Any]] = []
    for entry in inventory.get("files", []):
        inspection = entry.get("inspection", {})
        sanitized_inspection: dict[str, Any] = {
            "status": inspection.get("status"),
            "notes": inspection.get("notes"),
        }
        if entry.get("classification") == "EXCEL":
            sanitized_inspection["sheet_names"] = inspection.get("sheet_names", [])
            sanitized_inspection["sheet_types"] = inspection.get("sheet_types", [])
            sanitized_inspection["sheet_count"] = len(inspection.get("sheet_names", []))
            sanitized_inspection["sheet_dimensions"] = inspection.get("sheet_dimensions", [])
        elif entry.get("classification") == "BC3":
            sanitized_inspection["detected_encoding"] = inspection.get("detected_encoding")
            sanitized_inspection["is_text_like"] = inspection.get("is_text_like")

        sanitized_files.append(
            {
                "relative_path": entry.get("relative_path"),
                "extension": entry.get("extension"),
                "size_bytes": entry.get("size_bytes"),
                "classification": entry.get("classification"),
                "backup_hint": entry.get("backup_hint"),
                "version_or_phase_hint": entry.get("version_or_phase_hint"),
                "is_ignored": entry.get("is_ignored"),
                "ignored_reason": entry.get("ignored_reason"),
                "inspection": sanitized_inspection,
            }
        )

    return {
        "generated_at": inventory.get("generated_at"),
        "samples_dir": inventory.get("samples_dir"),
        "exists": inventory.get("exists"),
        "files_count_total": inventory.get("files_count_total"),
        "sample_files_count": inventory.get("sample_files_count"),
        "ignored_files_count": inventory.get("ignored_files_count"),
        "message": inventory.get("message"),
        "counts_by_class_total": inventory.get("counts_by_class_total", {}),
        "counts_by_class": inventory.get("counts_by_class", {}),
        "duplicates_by_hash": duplicates_by_hash,
        "files": sanitized_files,
    }


def write_reports(root: Path, inventory: dict[str, Any], sanitized: dict[str, Any]) -> tuple[Path, Path, Path, Path]:
    json_path = root / JSON_REPORT
    md_path = root / MD_REPORT
    sanitized_json_path = root / SANITIZED_JSON_REPORT
    sanitized_md_path = root / SANITIZED_MD_REPORT
    json_path.parent.mkdir(parents=True, exist_ok=True)

    json_path.write_text(
        json.dumps(inventory, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    sanitized_json_path.write_text(
        json.dumps(sanitized, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    lines: list[str] = []
    lines.append("# Sample Inventory Report")
    lines.append("")
    lines.append("> Local report: may include sensitive previews. Keep out of Git.")
    lines.append("")
    lines.append(f"- Generated at (UTC): {inventory['generated_at']}")
    lines.append(f"- Samples directory: {inventory['samples_dir']}")
    lines.append(f"- Exists: {inventory['exists']}")
    lines.append(f"- Files count total: {inventory['files_count_total']}")
    lines.append(f"- Sample files count: {inventory['sample_files_count']}")
    lines.append(f"- Ignored files count: {inventory['ignored_files_count']}")
    lines.append(f"- Message: {inventory['message']}")
    lines.append("")

    lines.append("## Counts by class")
    lines.append("")
    if inventory.get("counts_by_class"):
        for key in sorted(inventory["counts_by_class"]):
            lines.append(f"- {key}: {inventory['counts_by_class'][key]}")
    else:
        lines.append("- No files classified.")
    lines.append("")

    lines.append("## Files")
    lines.append("")
    if inventory.get("files"):
        for entry in inventory["files"]:
            lines.append(f"### {entry['relative_path']}")
            lines.append(f"- Classification: {entry['classification']}")
            lines.append(f"- Backup hint: {entry['backup_hint']}")
            lines.append(f"- Version/phase hint: {entry['version_or_phase_hint']}")
            lines.append(f"- Inspection status: {entry['inspection']['status']}")
            notes = entry["inspection"].get("notes")
            if notes:
                lines.append(f"- Notes: {notes}")
            lines.append("")
    else:
        lines.append("No sample files found.")
        lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")

    safe_lines: list[str] = []
    safe_lines.append("# Sample Inventory Sanitized Report")
    safe_lines.append("")
    safe_lines.append("- Shareable report without file content previews.")
    safe_lines.append(f"- Generated at (UTC): {sanitized['generated_at']}")
    safe_lines.append(f"- Files count total: {sanitized['files_count_total']}")
    safe_lines.append(f"- Sample files count: {sanitized['sample_files_count']}")
    safe_lines.append(f"- Ignored files count: {sanitized['ignored_files_count']}")
    safe_lines.append("")
    safe_lines.append("## Counts by class (operational)")
    safe_lines.append("")
    if sanitized.get("counts_by_class"):
        for key in sorted(sanitized["counts_by_class"]):
            safe_lines.append(f"- {key}: {sanitized['counts_by_class'][key]}")
    else:
        safe_lines.append("- No files classified.")
    safe_lines.append("")
    safe_lines.append(f"- Duplicate hash groups: {len(sanitized.get('duplicates_by_hash', []))}")
    safe_lines.append("")

    sanitized_md_path.write_text("\n".join(safe_lines), encoding="utf-8")
    return json_path, md_path, sanitized_json_path, sanitized_md_path


def print_summary(
    inventory: dict[str, Any],
    json_path: Path,
    md_path: Path,
    sanitized_json_path: Path,
    sanitized_md_path: Path,
    root: Path,
) -> None:
    print("Sample inspection summary")
    print(f"- Samples dir: {inventory['samples_dir']}")
    print(f"- Exists: {inventory['exists']}")
    print(f"- Files total: {inventory['files_count_total']}")
    print(f"- Sample files: {inventory['sample_files_count']}")
    print(f"- Ignored files: {inventory['ignored_files_count']}")
    print(f"- JSON report: {json_path.relative_to(root).as_posix()}")
    print(f"- Markdown report: {md_path.relative_to(root).as_posix()}")
    print(f"- Sanitized JSON report: {sanitized_json_path.relative_to(root).as_posix()}")
    print(f"- Sanitized Markdown report: {sanitized_md_path.relative_to(root).as_posix()}")


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    inventory = inspect_samples(root)
    duplicates_by_hash = load_duplicates_by_hash(root)
    sanitized = build_sanitized_inventory(inventory, duplicates_by_hash)
    json_path, md_path, sanitized_json_path, sanitized_md_path = write_reports(root, inventory, sanitized)
    print_summary(inventory, json_path, md_path, sanitized_json_path, sanitized_md_path, root)

    if not inventory["exists"]:
        print("WARNING: data/samples directory does not exist.")
    elif inventory["sample_files_count"] == 0:
        print("INFO: No sample files found in data/samples.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
