#!/usr/bin/env python
# Generar master Excel con ratios consolidados

import sqlite3
import pandas as pd
from datetime import datetime
from pathlib import Path
import os

def generar_master_excel(db_path="data/master/ratios.db", output_path="data/exports"):
    """Genera Excel master con ratios consolidados"""

    os.makedirs(output_path, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d")
    excel_file = os.path.join(output_path, f"MASTER_{timestamp}.xlsx")

    print("Generando Master Excel...")

    conn = sqlite3.connect(db_path)

    # Query 1: Items con ratios
    query_items = """
    SELECT
        COALESCE(categoria, 'SIN_CATEGORIA') as categoria,
        COALESCE(subcategoria, '') as subcategoria,
        item_key as descripcion,
        muestras_count as N,
        ROUND(COALESCE(min_unitario, 0), 2) as precio_min,
        ROUND(COALESCE(max_unitario, 0), 2) as precio_max,
        ROUND(COALESCE(media_unitario, 0), 2) as precio_promedio,
        ROUND(COALESCE(desv_std, 0), 2) as desv_std,
        CASE
            WHEN muestras_count >= 10 THEN 'MUY_SÓLIDO'
            WHEN muestras_count >= 5 THEN 'SÓLIDO'
            WHEN muestras_count >= 2 THEN 'DÉBIL'
            ELSE 'MUY_DÉBIL'
        END as confianza,
        COALESCE(categoria_asignada, 'PENDIENTE') as estado
    FROM item_master
    ORDER BY categoria, descripcion
    """

    df_items = pd.read_sql_query(query_items, conn)

    print(f"  Items extraídos: {len(df_items)}")

    # Query 2: Resumen por categoría
    query_categorias = """
    SELECT
        COALESCE(categoria, 'SIN_CATEGORIA') as categoria,
        COUNT(*) as items,
        ROUND(AVG(muestras_count), 1) as N_promedio,
        ROUND(AVG(media_unitario), 2) as precio_promedio,
        CASE
            WHEN AVG(muestras_count) >= 10 THEN 'MUY_SÓLIDO'
            WHEN AVG(muestras_count) >= 5 THEN 'SÓLIDO'
            WHEN AVG(muestras_count) >= 2 THEN 'DÉBIL'
            ELSE 'MUY_DÉBIL'
        END as confianza_promedio
    FROM item_master
    GROUP BY categoria
    ORDER BY N_promedio DESC
    """

    df_categorias = pd.read_sql_query(query_categorias, conn)

    print(f"  Categorías: {len(df_categorias)}")

    # Contar items por confianza
    confianza_dist = df_items['confianza'].value_counts()

    # Crear Excel con múltiples sheets
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df_items.to_excel(writer, sheet_name='ITEMS', index=False)
        df_categorias.to_excel(writer, sheet_name='RESUMEN_CATEGORIAS', index=False)

        # Sheet metadata
        metadata_data = {
            'Campo': [
                'Fecha generación',
                'Total items únicos',
                'Items SÓLIDO (N≥5)',
                'Items DÉBIL (N 2-4)',
                'Items MUY_DÉBIL (N=1)',
                'Presupuestos importados',
                'Deduplicación (%)',
                'Última actualización'
            ],
            'Valor': [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                len(df_items),
                len(df_items[df_items['N'] >= 5]),
                len(df_items[(df_items['N'] >= 2) & (df_items['N'] < 5)]),
                len(df_items[df_items['N'] == 1]),
                5,
                60,
                datetime.now().strftime("%Y-%m-%d")
            ]
        }
        df_metadata = pd.DataFrame(metadata_data)
        df_metadata.to_excel(writer, sheet_name='METADATA', index=False)

    # Aplicar formato
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = load_workbook(excel_file)

        # Formato: Headers
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet ITEMS: colorear por confianza
        ws_items = wb['ITEMS']

        # Header
        for cell in ws_items[1]:
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Datos
        confianza_colors = {
            'MUY_SÓLIDO': 'C6EFCE',  # Verde claro
            'SÓLIDO': 'FFEB9C',      # Amarillo
            'DÉBIL': 'FFC7CE',       # Rojo claro
            'MUY_DÉBIL': 'E8E8E8'    # Gris
        }

        for row_idx, row in enumerate(ws_items.iter_rows(min_row=2, max_row=len(df_items)+1), start=2):
            confianza_cell = row[7]  # Columna confianza
            confianza = confianza_cell.value

            color = confianza_colors.get(confianza, 'FFFFFF')
            for cell in row:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-ajustar ancho columnas
        for column in ws_items.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_items.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(excel_file)

    except Exception as e:
        print(f"  Warning: No se pudo aplicar formato Excel: {e}")

    conn.close()

    print(f"OK: Master Excel generado: {excel_file}")
    print("\nResumen:")
    print(f"  Items unicos: {len(df_items)}")
    print(f"  SOLIDO (N>=5): {len(df_items[df_items['N'] >= 5])}")
    print(f"  DEBIL (N 2-4): {len(df_items[(df_items['N'] >= 2) & (df_items['N'] < 5)])}")
    print(f"  MUY_DEBIL (N=1): {len(df_items[df_items['N'] == 1])}")
    print(f"  Categorias: {len(df_categorias)}")

    return excel_file

if __name__ == "__main__":
    generar_master_excel()
