"""Integrity validation helpers for the live Excel master workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook  # type: ignore

BLOCKED_STATUSES = {"BLOCKED", "ERROR", "VALIDATION_BLOCKED", "MANUAL_REVIEW_REQUIRED"}
ALLOWED_VALIDATION_STATUSES = {
    "VALIDATED",
    "PENDING",
    "WARNING",
    "BLOCKED",
    "ERROR",
    "VALIDATION_BLOCKED",
    "MANUAL_REVIEW_REQUIRED",
}

REQUIRED_SHEETS_COLUMNS: dict[str, list[str]] = {
    "README_MASTER": ["field", "value", "updated_at", "updated_by"],
    "IMPORT_LOG": [
        "import_batch_id",
        "run_id",
        "started_at",
        "finished_at",
        "status",
        "records_written",
        "records_skipped",
        "error_code",
        "error_message",
    ],
    "SOURCE_FILES": [
        "source_file_id",
        "original_filename",
        "file_hash",
        "file_type_detected",
        "source_priority",
        "ingested_at",
        "import_batch_id",
        "sensitivity_flag",
    ],
    "PROJECTS": [
        "project_id",
        "project_code",
        "project_name",
        "surface_base_value",
        "surface_base_unit",
        "surface_base_status",
        "created_at",
        "updated_at",
    ],
    "BUDGET_VERSIONS": [
        "budget_version_id",
        "project_id",
        "source_file_id",
        "version_name",
        "version_date",
        "is_latest_version",
        "validation_status",
        "import_batch_id",
    ],
    "RAW_IMPORTS": [
        "raw_import_id",
        "source_file_id",
        "raw_path_ref",
        "raw_hash",
        "ingested_at",
        "import_batch_id",
        "raw_access_policy",
    ],
    "COST_ITEMS": [
        "cost_item_id",
        "budget_version_id",
        "source_file_id",
        "origin_record_ref",
        "description_raw",
        "unit_raw",
        "quantity_raw",
        "unit_price_raw",
        "amount_raw",
        "row_hash",
        "validation_status",
    ],
    "NORMALIZED_COST_ITEMS": [
        "normalized_cost_item_id",
        "cost_item_id",
        "normalized_description",
        "normalized_unit",
        "normalized_quantity",
        "normalized_amount",
        "normalization_status",
        "normalization_rule_ref",
        "row_hash",
        "validation_status",
    ],
    "CATEGORY_MAPPING": [
        "mapping_id",
        "mapping_key",
        "target_category",
        "mapping_confidence",
        "mapping_status",
        "decision_source",
        "approved_by",
        "approved_at",
    ],
    "VALIDATION_RESULTS": [
        "validation_result_id",
        "entity_type",
        "entity_id",
        "rule_id",
        "severity",
        "status",
        "message",
        "validated_at",
        "import_batch_id",
    ],
    "RATIO_INPUTS": [
        "ratio_input_id",
        "normalized_cost_item_id",
        "project_id",
        "budget_version_id",
        "eligibility_status",
        "exclusion_flag",
        "validation_status",
        "effective_at",
    ],
    "RATIOS_CALCULATED": [
        "ratio_calc_id",
        "ratio_code",
        "project_scope",
        "input_version_ref",
        "calculated_value",
        "calculation_status",
        "calculated_at",
    ],
    "EXCLUSIONS": [
        "exclusion_id",
        "entity_type",
        "entity_id",
        "reason_code",
        "reason_detail",
        "is_reversible",
        "excluded_at",
        "excluded_by",
        "import_batch_id",
    ],
    "SNAPSHOTS": [
        "snapshot_id",
        "snapshot_ts",
        "master_version",
        "trigger_reason",
        "source_run_id",
        "storage_ref",
        "checksum",
        "created_by",
    ],
    "CHANGELOG": [
        "change_id",
        "change_ts",
        "change_type",
        "affected_sheet",
        "change_summary",
        "decision_ref",
        "applied_by",
    ],
}

SYNTHETIC_PK_COLUMNS = {
    "SOURCE_FILES": "source_file_id",
    "PROJECTS": "project_id",
    "BUDGET_VERSIONS": "budget_version_id",
    "IMPORT_LOG": "import_batch_id",
    "RAW_IMPORTS": "raw_import_id",
    "COST_ITEMS": "cost_item_id",
    "VALIDATION_RESULTS": "validation_result_id",
    "EXCLUSIONS": "exclusion_id",
    "RATIO_INPUTS": "ratio_input_id",
}

REQUIRED_NON_EMPTY_COLUMNS = {
    "BUDGET_VERSIONS": ["budget_version_id", "source_file_id", "validation_status"],
    "COST_ITEMS": ["cost_item_id", "budget_version_id", "source_file_id", "validation_status"],
    "NORMALIZED_COST_ITEMS": ["normalized_cost_item_id", "cost_item_id", "validation_status"],
    "RATIO_INPUTS": ["ratio_input_id", "budget_version_id", "validation_status"],
}

VALIDATION_STATUS_SHEETS = {"BUDGET_VERSIONS", "COST_ITEMS", "NORMALIZED_COST_ITEMS", "RATIO_INPUTS"}


class SchemaValidationError(RuntimeError):
    """Raised when workbook schema does not satisfy required contract."""


class ReferentialValidationError(RuntimeError):
    """Raised when workbook data does not satisfy referential contract."""


def _header_values(ws: object, expected_len: int) -> list[str]:
    return [
        str(getattr(ws.cell(row=1, column=idx), "value") or "").strip()
        for idx in range(1, expected_len + 1)
    ]


def _sheet_dict_rows(workbook: object, sheet_name: str) -> list[dict[str, str]]:
    ws = workbook[sheet_name]
    headers = REQUIRED_SHEETS_COLUMNS[sheet_name]
    rows: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=idx).value for idx in range(1, len(headers) + 1)]
        row_map = {headers[i]: ("" if values[i] is None else str(values[i]).strip()) for i in range(len(headers))}
        if any(row_map.values()):
            rows.append(row_map)
    return rows


def ensure_allowed_output_path(path: Path, root: Path) -> None:
    try:
        path.resolve().relative_to(root.resolve())
    except ValueError as exc:
        raise ValueError(f"Output path must be inside '{root.as_posix()}': {path.as_posix()}") from exc


def ensure_allowed_snapshot_path(path: Path, root: Path) -> None:
    try:
        path.resolve().relative_to(root.resolve())
    except ValueError as exc:
        raise ValueError(f"Snapshot path must be inside '{root.as_posix()}': {path.as_posix()}") from exc


def contains_allowed_anchor(path: Path) -> bool:
    parts = [p.lower() for p in path.parts]
    return any(parts[idx] == "outputs" and parts[idx + 1] == "live_excel_master" for idx in range(len(parts) - 1))


def validate_workbook_schema(workbook: object) -> None:
    missing_sheets = [name for name in REQUIRED_SHEETS_COLUMNS if name not in workbook.sheetnames]
    if missing_sheets:
        raise SchemaValidationError(f"Missing required sheets: {', '.join(sorted(missing_sheets))}")

    for sheet_name, columns in REQUIRED_SHEETS_COLUMNS.items():
        ws = workbook[sheet_name]
        actual = _header_values(ws, len(columns))
        missing_columns = [col for col in columns if col not in actual]
        if missing_columns:
            raise SchemaValidationError(
                f"Sheet '{sheet_name}' missing required columns: {', '.join(missing_columns)}"
            )


def validate_referential_integrity(workbook: object) -> None:
    source_rows = _sheet_dict_rows(workbook, "SOURCE_FILES")
    budget_rows = _sheet_dict_rows(workbook, "BUDGET_VERSIONS")
    raw_rows = _sheet_dict_rows(workbook, "RAW_IMPORTS")
    cost_rows = _sheet_dict_rows(workbook, "COST_ITEMS")
    validation_rows = _sheet_dict_rows(workbook, "VALIDATION_RESULTS")
    import_rows = _sheet_dict_rows(workbook, "IMPORT_LOG")
    exclusion_rows = _sheet_dict_rows(workbook, "EXCLUSIONS")
    ratio_input_rows = _sheet_dict_rows(workbook, "RATIO_INPUTS")
    normalized_rows = _sheet_dict_rows(workbook, "NORMALIZED_COST_ITEMS")

    source_ids = {row["source_file_id"] for row in source_rows if row["source_file_id"]}
    budget_ids = {row["budget_version_id"] for row in budget_rows if row["budget_version_id"]}
    import_batch_ids = {row["import_batch_id"] for row in import_rows if row["import_batch_id"]}
    cost_ids = {row["cost_item_id"] for row in cost_rows if row["cost_item_id"]}
    blocked_cost_ids = {
        row["cost_item_id"]
        for row in cost_rows
        if row.get("cost_item_id")
        and row.get("validation_status", "").upper() in BLOCKED_STATUSES
    }
    excluded_cost_ids = {
        row["entity_id"]
        for row in exclusion_rows
        if row.get("entity_type", "").upper() == "COST_ITEM" and row.get("entity_id")
    }

    for sheet_name, pk_col in SYNTHETIC_PK_COLUMNS.items():
        seen: set[str] = set()
        for row in _sheet_dict_rows(workbook, sheet_name):
            pk = row.get(pk_col, "")
            if not pk:
                raise ReferentialValidationError(f"{sheet_name}.{pk_col} cannot be empty.")
            if pk in seen:
                raise ReferentialValidationError(f"Duplicate key found in {sheet_name}.{pk_col}: {pk}")
            seen.add(pk)

    for sheet_name, required_cols in REQUIRED_NON_EMPTY_COLUMNS.items():
        for row in _sheet_dict_rows(workbook, sheet_name):
            for col in required_cols:
                if not row.get(col):
                    raise ReferentialValidationError(f"{sheet_name}.{col} cannot be empty.")

    for sheet_name in VALIDATION_STATUS_SHEETS:
        for row in _sheet_dict_rows(workbook, sheet_name):
            status = row.get("validation_status", "").strip().upper()
            if not status:
                raise ReferentialValidationError(f"{sheet_name}.validation_status cannot be empty.")
            if status not in ALLOWED_VALIDATION_STATUSES:
                raise ReferentialValidationError(
                    f"{sheet_name}.validation_status is unknown: {row.get('validation_status', '')}"
                )

    for row in budget_rows:
        ref = row.get("source_file_id", "")
        if ref and ref not in source_ids:
            raise ReferentialValidationError(
                f"BUDGET_VERSIONS.source_file_id not found in SOURCE_FILES: {ref}"
            )

    for row in cost_rows:
        budget_ref = row.get("budget_version_id", "")
        source_ref = row.get("source_file_id", "")
        if budget_ref and budget_ref not in budget_ids:
            raise ReferentialValidationError(
                f"COST_ITEMS.budget_version_id not found in BUDGET_VERSIONS: {budget_ref}"
            )
        if source_ref and source_ref not in source_ids:
            raise ReferentialValidationError(
                f"COST_ITEMS.source_file_id not found in SOURCE_FILES: {source_ref}"
            )

    for row in normalized_rows:
        cost_ref = row.get("cost_item_id", "")
        if cost_ref and cost_ref not in cost_ids:
            raise ReferentialValidationError(
                f"NORMALIZED_COST_ITEMS.cost_item_id not found in COST_ITEMS: {cost_ref}"
            )
        if cost_ref and cost_ref in blocked_cost_ids:
            raise ReferentialValidationError(
                "NORMALIZED_COST_ITEMS contains blocked/error cost item reference."
            )

    for row in raw_rows:
        ref = row.get("source_file_id", "")
        if ref and ref not in source_ids:
            raise ReferentialValidationError(f"RAW_IMPORTS.source_file_id not found in SOURCE_FILES: {ref}")

    for row in validation_rows:
        ref = row.get("import_batch_id", "")
        if ref and ref not in import_batch_ids:
            raise ReferentialValidationError(
                f"VALIDATION_RESULTS.import_batch_id not found in IMPORT_LOG: {ref}"
            )

    for row in exclusion_rows:
        if row.get("entity_type", "").upper() == "SOURCE_FILE":
            source_entity = row.get("entity_id", "")
            if source_entity and source_entity not in source_ids:
                raise ReferentialValidationError(
                    f"EXCLUSIONS SOURCE_FILE entity_id not found in SOURCE_FILES: {source_entity}"
                )

    for row in ratio_input_rows:
        status = row.get("validation_status", "").upper()
        if status in BLOCKED_STATUSES:
            raise ReferentialValidationError(
                "RATIO_INPUTS contains blocked/error/manual-review validation_status; promotion must be blocked."
            )
        budget_ref = row.get("budget_version_id", "")
        if budget_ref and budget_ref not in budget_ids:
            raise ReferentialValidationError(
                f"RATIO_INPUTS.budget_version_id not found in BUDGET_VERSIONS: {budget_ref}"
            )
        if row.get("normalized_cost_item_id", "") in excluded_cost_ids:
            raise ReferentialValidationError("RATIO_INPUTS contains excluded item reference.")


def validate_workbook_file(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        validate_workbook_schema(workbook)
        validate_referential_integrity(workbook)
    finally:
        workbook.close()
