#!/usr/bin/env python3
"""Generate synthetic test budgets for validation."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

# Ensure output directory exists
OUTPUT_DIR = Path("data/samples/PRESUPUESTOS")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Color palette
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)


def create_presupuesto_1():
    """Create TEST_PRESUPUESTO_EXTRA_1.xlsx (5 capítulos, 15 items)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    # Header
    headers = ["Capítulo", "Código", "Descripción", "Unitario (€)", "Cantidad", "Total (€)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data
    data = [
        # Chapter 01: Excavación
        ("01", "01.01", "Excavación a máquina", 55.00, 120, 6600.00),
        ("01", "01.02", "Excavación manual", 45.00, 30, 1350.00),
        ("01", "01.03", "Retirada de tierras", 8.00, 200, 1600.00),
        # Chapter 02: Cimentación
        ("02", "02.01", "Hormigón armado pilares", 480.00, 8, 3840.00),
        ("02", "02.02", "Acero de refuerzo", 1250.00, 2.5, 3125.00),
        ("02", "02.03", "Encofrado de hormigón", 35.00, 150, 5250.00),
        # Chapter 03: Estructura
        ("03", "03.01", "Acero laminado estructural", 1200.00, 3, 3600.00),
        ("03", "03.02", "Soldadura estructural", 125.00, 50, 6250.00),
        ("03", "03.03", "Elementos de unión", 0.80, 1000, 800.00),
        # Chapter 04: Envolvente
        ("04", "04.01", "Ladrillo cara vista", 0.12, 8500, 1020.00),
        ("04", "04.02", "Aislamiento térmico", 25.00, 600, 15000.00),
        ("04", "04.03", "Mortero de unión", 500.00, 4, 2000.00),
        # Chapter 05: Cubierta
        ("05", "05.01", "Teja cerámica", 0.15, 5000, 750.00),
        ("05", "05.02", "Estructura cubierta madera", 45.00, 120, 5400.00),
        ("05", "05.03", "Impermeabilización", 30.00, 200, 6000.00),
    ]

    for row_num, (chapter, code, desc, unitario, qty, total) in enumerate(data, start=2):
        ws.cell(row=row_num, column=1, value=chapter).font = BODY_FONT
        ws.cell(row=row_num, column=2, value=code).font = BODY_FONT
        ws.cell(row=row_num, column=3, value=desc).font = BODY_FONT

        cell_unit = ws.cell(row=row_num, column=4, value=unitario)
        cell_unit.font = BODY_FONT
        cell_unit.number_format = '#,##0.00 "€"'

        ws.cell(row=row_num, column=5, value=qty).font = BODY_FONT

        cell_total = ws.cell(row=row_num, column=6, value=total)
        cell_total.font = BODY_FONT
        cell_total.number_format = '#,##0.00 "€"'

    # Auto width
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 18

    output_file = OUTPUT_DIR / "TEST_PRESUPUESTO_EXTRA_1.xlsx"
    wb.save(output_file)
    print(f"✅ Creado: {output_file.name}")
    return output_file


def create_presupuesto_2():
    """Create TEST_PRESUPUESTO_EXTRA_2.xlsx (5 capítulos, 12 items, datos superpuestos)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    # Header
    headers = ["Capítulo", "Código", "Descripción", "Unitario (€)", "Cantidad", "Total (€)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data (some items overlap with presupuesto_1)
    data = [
        # Chapter 01: Excavación (overlapping)
        ("01", "01.01", "Excavación a máquina", 60.00, 100, 6000.00),  # Different quantity/unit price
        ("01", "01.04", "Compactación de tierras", 12.00, 150, 1800.00),  # New item
        # Chapter 02: Cimentación (overlapping)
        ("02", "02.01", "Hormigón armado pilares", 500.00, 6, 3000.00),  # Different values
        ("02", "02.04", "Drenaje perimetral", 28.00, 100, 2800.00),  # New item
        # Chapter 03: Estructura (overlapping)
        ("03", "03.02", "Soldadura estructural", 130.00, 45, 5850.00),  # Different values
        ("03", "03.04", "Protección de estructura", 18.00, 200, 3600.00),  # New item
        # Chapter 04: Envolvente (overlapping)
        ("04", "04.02", "Aislamiento térmico", 26.00, 550, 14300.00),  # Different values
        ("04", "04.04", "Ventilación fachada", 15.00, 300, 4500.00),  # New item
        # Chapter 05: Cubierta (overlapping)
        ("05", "05.01", "Teja cerámica", 0.16, 4500, 720.00),  # Different values
        ("05", "05.04", "Canalización", 22.00, 180, 3960.00),  # New item
        ("05", "05.05", "Ganchos de seguridad", 8.50, 50, 425.00),  # New item
        ("05", "05.06", "Sellado de encuentros", 5.00, 400, 2000.00),  # New item
    ]

    for row_num, (chapter, code, desc, unitario, qty, total) in enumerate(data, start=2):
        ws.cell(row=row_num, column=1, value=chapter).font = BODY_FONT
        ws.cell(row=row_num, column=2, value=code).font = BODY_FONT
        ws.cell(row=row_num, column=3, value=desc).font = BODY_FONT

        cell_unit = ws.cell(row=row_num, column=4, value=unitario)
        cell_unit.font = BODY_FONT
        cell_unit.number_format = '#,##0.00 "€"'

        ws.cell(row=row_num, column=5, value=qty).font = BODY_FONT

        cell_total = ws.cell(row=row_num, column=6, value=total)
        cell_total.font = BODY_FONT
        cell_total.number_format = '#,##0.00 "€"'

    # Auto width
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 18

    output_file = OUTPUT_DIR / "TEST_PRESUPUESTO_EXTRA_2.xlsx"
    wb.save(output_file)
    print(f"✅ Creado: {output_file.name}")
    return output_file


if __name__ == "__main__":
    print("📝 Generando presupuestos de prueba...\n")
    file1 = create_presupuesto_1()
    file2 = create_presupuesto_2()

    print(f"\n📂 Guardados en: {OUTPUT_DIR}")
    print(f"\nARCHIVOS CREADOS:")
    print(f"  1. {file1.name}")
    print(f"     - 5 capítulos")
    print(f"     - 15 items")
    print(f"  2. {file2.name}")
    print(f"     - 5 capítulos")
    print(f"     - 12 items")
    print(f"     - Datos superpuestos con presupuesto 1")
