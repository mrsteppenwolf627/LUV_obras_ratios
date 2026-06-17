#!/usr/bin/env python3
"""Auditable converter: 260318_PEC Presto-export XLSX -> JSON for /api/import/budgets.

SCOPE: this converter is specific to the Presto-export layout of
`260318_PEC- Presupuesto_CONTRATO.xlsx` (single sheet, header row with
columns Codigo/Nat/Ud/Resumen/.../CanPres/Pres/ImpPres). It is NOT generic
and must not be reused on the other budgets without their own inspection.

It only READS the Excel and WRITES a local JSON. It performs no network call,
no POST, and no DB access. Prices are never invented or defaulted: rows
without a real unit price (Pres) or quantity (CanPres) are discarded.

Authorized mapping (this file only):
  descripcion     <- Resumen
  unidad          <- Ud
  cantidad        <- CanPres
  precio_unitario <- Pres
  importe (validation only) <- ImpPres
  capitulo        <- top-level code prefix of Codigo (split on '.', [0])
Only rows with Nat == 'Partida' are processed.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_INPUT = Path("data/temp_import_xlsx_only/260318_PEC- Presupuesto_CONTRATO.xlsx")
DEFAULT_OUTPUT = Path("data/temp_import_xlsx_only/260318_PEC.import.json")
# Source has no building_type field; use an explicit neutral placeholder
# (NOT inventing residential/commercial). Documented in CONTEXT.md.
BUILDING_TYPE = "sin_especificar"

# Schema bounds mirrored from app/schemas/import_budgets.py to pre-filter
MAX_DESC = 500
MAX_PRICE = 1_000_000
MAX_QTY = 1_000_000


def _norm(s: str) -> str:
    """Lowercase + strip accents for robust header matching."""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return s.strip().lower()


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest().lower()


def find_header(ws):
    """Locate the header row and map required column names to 1-based indices."""
    wanted = {
        "codigo": ("codigo",),
        "nat": ("nat",),
        "ud": ("ud",),
        "resumen": ("resumen",),
        "canpres": ("canpres",),
        "pres": ("pres",),
        "imppres": ("imppres",),
    }
    for r in range(1, min(30, ws.max_row) + 1):
        labels = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val is None:
                continue
            labels[_norm(val)] = c
        # Header row must expose at least Nat, Resumen, CanPres, Pres
        if {"nat", "resumen", "canpres", "pres"}.issubset(labels.keys()):
            colmap = {}
            for key, aliases in wanted.items():
                for a in aliases:
                    if a in labels:
                        colmap[key] = labels[a]
                        break
            return r, colmap
    raise RuntimeError("No se encontro la fila de cabecera esperada (Nat/Resumen/CanPres/Pres)")


def convert(input_path: Path, output_path: Path) -> dict:
    wb = load_workbook(input_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, col = find_header(ws)

    file_hash = sha256_of(input_path)

    lineas = []
    partidas = 0
    discarded = {
        "sin_precio": 0,      # Pres vacio o <= 0
        "sin_cantidad": 0,    # CanPres vacio o <= 0
        "sin_descripcion": 0, # Resumen vacio
        "fuera_schema": 0,    # precio/cantidad fuera de limites
    }
    discarded_samples = []
    coherence_ok = 0
    coherence_bad = 0

    for r in range(header_row + 1, ws.max_row + 1):
        nat = ws.cell(r, col["nat"]).value
        if nat is None or _norm(nat) != "partida":
            continue
        partidas += 1

        desc_raw = ws.cell(r, col["resumen"]).value
        ud_raw = ws.cell(r, col["ud"]).value if "ud" in col else None
        canpres = _num(ws.cell(r, col["canpres"]).value)
        pres = _num(ws.cell(r, col["pres"]).value)
        imppres = _num(ws.cell(r, col["imppres"]).value) if "imppres" in col else None
        codigo = ws.cell(r, col["codigo"]).value if "codigo" in col else None

        desc = str(desc_raw).strip() if desc_raw is not None else ""
        reason = None
        if not desc:
            discarded["sin_descripcion"] += 1
            reason = "sin_descripcion"
        elif pres is None or pres <= 0:
            discarded["sin_precio"] += 1
            reason = "sin_precio"
        elif canpres is None or canpres <= 0:
            discarded["sin_cantidad"] += 1
            reason = "sin_cantidad"
        elif pres > MAX_PRICE or canpres > MAX_QTY or len(desc) > MAX_DESC:
            discarded["fuera_schema"] += 1
            reason = "fuera_schema"

        if reason is not None:
            if len(discarded_samples) < 10:
                discarded_samples.append(
                    {"row": r, "codigo": str(codigo), "desc": desc[:40],
                     "CanPres": canpres, "Pres": pres, "ImpPres": imppres, "motivo": reason}
                )
            continue

        # Coherence validation (does not alter data; only counts)
        if imppres is not None:
            if abs(canpres * pres - imppres) <= max(0.5, 0.01 * abs(imppres)):
                coherence_ok += 1
            else:
                coherence_bad += 1

        capitulo = str(codigo).split(".")[0] if codigo else ""
        unidad = str(ud_raw).strip() if ud_raw is not None and str(ud_raw).strip() else "ud"

        lineas.append({
            "numero": r,                      # source Excel row (traceability)
            "capitulo": capitulo[:100],
            "descripcion": desc,
            "cantidad": canpres,
            "unidad": unidad[:50],
            "precio_unitario": pres,
        })

    wb.close()

    payload = {
        "filename": input_path.name,
        "file_hash": file_hash,
        "building_type": BUILDING_TYPE,
        "lineas": lineas,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)

    return {
        "payload": payload,
        "partidas": partidas,
        "exportadas": len(lineas),
        "discarded": discarded,
        "discarded_samples": discarded_samples,
        "coherence_ok": coherence_ok,
        "coherence_bad": coherence_bad,
        "header_row": header_row,
        "colmap": col,
        "file_hash": file_hash,
        "output": str(output_path),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = ap.parse_args()

    if not args.input.exists():
        print(f"ERROR: no existe el fichero {args.input}", file=sys.stderr)
        return 1

    res = convert(args.input, args.output)

    print("=== CONVERSION 260318_PEC ===")
    print("input        :", args.input)
    print("output JSON  :", res["output"])
    print("file_hash    :", res["file_hash"])
    print("header_row   :", res["header_row"], "colmap:", res["colmap"])
    print("building_type:", BUILDING_TYPE, "(placeholder neutro, no inventado)")
    print()
    print("partidas detectadas :", res["partidas"])
    print("lineas exportadas   :", res["exportadas"])
    d = res["discarded"]
    print("descartadas         :", sum(d.values()),
          f"(sin_precio={d['sin_precio']}, sin_cantidad={d['sin_cantidad']}, "
          f"sin_descripcion={d['sin_descripcion']}, fuera_schema={d['fuera_schema']})")
    print(f"coherencia cant*precio==importe: OK={res['coherence_ok']} BAD={res['coherence_bad']}")
    print()
    print("=== Muestra 10 lineas EXPORTADAS ===")
    for ln in res["payload"]["lineas"][:10]:
        print(f"  R{ln['numero']:<5} [{ln['capitulo']:<8}] {ln['unidad']:<4} "
              f"cant={ln['cantidad']:<9} pu={ln['precio_unitario']:<9} {ln['descripcion'][:40]}")
    print()
    print("=== Muestra 10 lineas DESCARTADAS ===")
    for s in res["discarded_samples"]:
        print(f"  R{s['row']:<5} {s['motivo']:<15} CanPres={s['CanPres']} Pres={s['Pres']} "
              f"ImpPres={s['ImpPres']} {s['desc']}")

    # Local schema validation (no network, no DB)
    print()
    print("=== Validacion contra schema BudgetImportRequest (sin red/BD) ===")
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
        from app.schemas.import_budgets import BudgetImportRequest
        model = BudgetImportRequest(**res["payload"])
        print(f"  OK: schema valido. lineas={len(model.lineas)}, "
              f"file_hash valido, building_type={model.building_type!r}")
    except Exception as e:
        print("  FALLO de schema:", repr(e))
        return 2

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
