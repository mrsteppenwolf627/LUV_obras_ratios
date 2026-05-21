"""Adaptive professional review sheets for PREVIEW_ONLY XLSX outputs."""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
import re
from typing import Any
from uuid import uuid4

from openpyxl.styles import Alignment, Font, PatternFill

try:
    from scripts.xlsx_sheet_semantic_classifier import (
        SHEET_AUXILIARY,
        SHEET_BUDGET_CLASSIC,
        SHEET_BUDGET_SUMMARY,
        SHEET_COMPARISON_TABLE,
        SHEET_FORMULA_CALCULATION,
        SHEET_METADATA,
        SHEET_SPACE_BREAKDOWN,
        SHEET_UNKNOWN,
        SheetSemanticClassification,
    )
except ModuleNotFoundError:
    from xlsx_sheet_semantic_classifier import (  # type: ignore
        SHEET_AUXILIARY,
        SHEET_BUDGET_CLASSIC,
        SHEET_BUDGET_SUMMARY,
        SHEET_COMPARISON_TABLE,
        SHEET_FORMULA_CALCULATION,
        SHEET_METADATA,
        SHEET_SPACE_BREAKDOWN,
        SHEET_UNKNOWN,
        SheetSemanticClassification,
    )


TRACE_HEADERS = [
    "review_row_id",
    "review_sheet_name",
    "review_row_number",
    "sheet_type",
    "mapping_status",
    "validation_status",
    "source_file_id",
    "import_batch_id",
    "budget_version_id",
    "source_sheet_name",
    "source_row_number",
    "preserved_row_id",
    "cost_item_id",
    "trace_notes",
]


TYPE_HEADERS: dict[str, list[str]] = {
    SHEET_BUDGET_CLASSIC: ["Codigo", "Descripcion", "Ud", "Cantidad", "Precio unitario", "Importe"],
    SHEET_BUDGET_SUMMARY: ["Codigo", "Descripcion", "Formula / Ratio", "Importe"],
    SHEET_SPACE_BREAKDOWN: ["Codigo", "Info", "Resumen", "Presupuesto"],
    SHEET_COMPARISON_TABLE: ["Cap.", "Nombre del capítulo", "Importe (€)", "Nombre equivalente", "Importe equivalente", "Diferencia"],
    SHEET_FORMULA_CALCULATION: ["Concepto", "Formula", "Valor"],
    SHEET_AUXILIARY: ["Contenido", "Notas"],
    SHEET_METADATA: ["Contenido", "Notas"],
    SHEET_UNKNOWN: ["Contenido", "Notas"],
}


def _next_sequence(workbook: Any, prefix: str) -> int:
    pattern = re.compile(rf"^{re.escape(prefix)}_(\d{{3}})$")
    max_seq = 0
    for name in workbook.sheetnames:
        match = pattern.match(name)
        if match:
            max_seq = max(max_seq, int(match.group(1)))
    return max_seq + 1


def _safe_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _to_decimal(value: object) -> Decimal | None:
    text = _safe_text(value)
    if not text:
        return None
    text = text.replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _looks_like_year(text: str) -> bool:
    if not text.isdigit():
        return False
    year = int(text)
    return 1900 <= year <= 2100


def _looks_like_metadata_row(description: str, amount: str) -> bool:
    low = description.lower()
    if not low:
        return False
    month_pattern = re.compile(
        r"\b(?:enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre)\b",
        flags=re.IGNORECASE,
    )
    if month_pattern.search(low) and re.search(r"\b(19|20)\d{2}\b", low):
        return True
    if "revision" not in low and "fecha" not in low and "abril" not in low:
        return False
    return _looks_like_year(amount)


def _source_to_preserved_sheet(workbook: Any) -> dict[str, str]:
    if "PRESERVED_BUDGET_SHEETS" not in workbook.sheetnames:
        return {}
    ws = workbook["PRESERVED_BUDGET_SHEETS"]
    headers = [str(ws.cell(row=1, column=idx).value or "").strip() for idx in range(1, ws.max_column + 1)]
    idx = {name: pos + 1 for pos, name in enumerate(headers)}
    if "source_sheet_name" not in idx or "preserved_sheet_name" not in idx:
        return {}
    mapping: dict[str, str] = {}
    for row_idx in range(2, ws.max_row + 1):
        source = str(ws.cell(row=row_idx, column=idx["source_sheet_name"]).value or "").strip()
        preserved = str(ws.cell(row=row_idx, column=idx["preserved_sheet_name"]).value or "").strip()
        if source and preserved:
            mapping[source] = preserved
    return mapping


def _build_preserved_trace_index(workbook: Any, import_batch_id: str) -> dict[tuple[str, str], dict[str, str]]:
    if "PRESERVED_TO_COST_ITEMS_MAP" not in workbook.sheetnames:
        return {}
    ws = workbook["PRESERVED_TO_COST_ITEMS_MAP"]
    headers = [str(ws.cell(row=1, column=idx).value or "").strip() for idx in range(1, ws.max_column + 1)]
    idx = {name: pos + 1 for pos, name in enumerate(headers)}
    required = {"source_sheet_name", "source_row_number", "preserved_row_id", "cost_item_id", "mapping_status", "notes", "import_batch_id"}
    if not required.issubset(idx.keys()):
        return {}
    trace: dict[tuple[str, str], dict[str, str]] = {}
    for row_idx in range(2, ws.max_row + 1):
        row_import_batch_id = str(ws.cell(row=row_idx, column=idx["import_batch_id"]).value or "").strip()
        if row_import_batch_id != import_batch_id:
            continue
        key = (
            str(ws.cell(row=row_idx, column=idx["source_sheet_name"]).value or "").strip(),
            str(ws.cell(row=row_idx, column=idx["source_row_number"]).value or "").strip(),
        )
        if not key[0] or not key[1]:
            continue
        trace[key] = {
            "preserved_row_id": str(ws.cell(row=row_idx, column=idx["preserved_row_id"]).value or "").strip(),
            "cost_item_id": str(ws.cell(row=row_idx, column=idx["cost_item_id"]).value or "").strip(),
            "mapping_status": str(ws.cell(row=row_idx, column=idx["mapping_status"]).value or "").strip(),
            "map_notes": str(ws.cell(row=row_idx, column=idx["notes"]).value or "").strip(),
        }
    return trace


def _sanitize_suffix(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    return (cleaned or "SHEET")[:18]


def _cell_text(ws: Any, row_idx: int, col_idx: int | None) -> str:
    if not col_idx:
        return ""
    return _safe_text(ws.cell(row=row_idx, column=col_idx).value)


def _extract_view_rows(
    preserved_ws: Any,
    semantic: SheetSemanticClassification,
) -> list[dict[str, str]]:
    fields = semantic.detected_columns
    rows: list[dict[str, str]] = []
    for row_idx in range(semantic.header_row + 1, preserved_ws.max_row + 1):
        code = _cell_text(preserved_ws, row_idx, fields.get("code") or fields.get("cap"))
        description = _cell_text(preserved_ws, row_idx, fields.get("description") or fields.get("chapter_name"))
        amount = _cell_text(preserved_ws, row_idx, fields.get("amount"))
        formula = _cell_text(preserved_ws, row_idx, fields.get("formula"))
        if semantic.sheet_type == SHEET_COMPARISON_TABLE:
            description = _cell_text(preserved_ws, row_idx, fields.get("chapter_name"))
        if not any([code, description, amount, formula]):
            continue
        if _looks_like_metadata_row(description, amount):
            continue
        rows.append(
            {
                "row_number": str(row_idx),
                "code": code,
                "description": description,
                "unit": _cell_text(preserved_ws, row_idx, fields.get("unit")),
                "quantity": _cell_text(preserved_ws, row_idx, fields.get("quantity")),
                "unit_price": _cell_text(preserved_ws, row_idx, fields.get("unit_price")),
                "amount": amount,
                "formula": formula,
                "info": _cell_text(preserved_ws, row_idx, fields.get("info")),
                "equivalent_name": _cell_text(preserved_ws, row_idx, fields.get("equivalent_name")),
                "equivalent_amount": _cell_text(preserved_ws, row_idx, fields.get("equivalent_amount")),
                "difference": _cell_text(preserved_ws, row_idx, fields.get("difference")),
            }
        )
    return rows


def _style_review_sheet(ws: Any, visible_columns: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(visible_columns, 2))
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A5"
    for col_idx in range(1, visible_columns + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.auto_filter.ref = f"A4:{chr(64 + visible_columns)}4"
    for row_idx in range(5, ws.max_row + 1):
        for col_idx in range(1, visible_columns + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            header = str(ws.cell(row=4, column=col_idx).value or "").lower()
            if header in {"descripcion", "nombre del capítulo", "resumen", "contenido", "notas", "formula / ratio", "concepto"}:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif header in {"ud", "cap."}:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if header in {"cantidad", "precio unitario", "importe", "importe (€)", "importe equivalente", "presupuesto", "valor"}:
                parsed = _to_decimal(cell.value)
                if parsed is not None:
                    cell.value = float(parsed)
                    cell.number_format = '#,##0.00 "EUR"' if "importe" in header or header == "presupuesto" else "#,##0.00"
    ws.sheet_view.showGridLines = False
    for col_idx in range(1, visible_columns + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 20 if col_idx != 2 else 52
    if ws.max_column >= visible_columns + 1:
        ws.column_dimensions[chr(64 + visible_columns + 1)].hidden = True


def _append_trace_row(
    trace_ws: Any,
    review_row_id: str,
    review_sheet_name: str,
    review_row_number: int,
    sheet_type: str,
    source_file_id: str,
    import_batch_id: str,
    budget_version_id: str,
    source_sheet_name: str,
    source_row_number: str,
    map_row: dict[str, str] | None,
    notes: str,
) -> None:
    map_row = map_row or {}
    trace_ws.append(
        [
            review_row_id,
            review_sheet_name,
            str(review_row_number),
            sheet_type,
            map_row.get("mapping_status", "NOT_COST_ITEM"),
            "PENDING",
            source_file_id,
            import_batch_id,
            budget_version_id,
            source_sheet_name,
            source_row_number,
            map_row.get("preserved_row_id", ""),
            map_row.get("cost_item_id", ""),
            "|".join(part for part in [notes, map_row.get("map_notes", "")] if part),
        ]
    )


def _write_type_row(ws: Any, row_idx: int, sheet_type: str, row_data: dict[str, str], review_row_id: str) -> None:
    if sheet_type == SHEET_BUDGET_CLASSIC:
        values = [
            row_data["code"],
            row_data["description"],
            row_data["unit"],
            row_data["quantity"],
            row_data["unit_price"],
            row_data["amount"],
        ]
    elif sheet_type == SHEET_BUDGET_SUMMARY:
        values = [row_data["code"], row_data["description"], row_data["formula"], row_data["amount"]]
    elif sheet_type == SHEET_SPACE_BREAKDOWN:
        values = [row_data["code"], row_data["info"], row_data["description"], row_data["amount"]]
    elif sheet_type == SHEET_COMPARISON_TABLE:
        values = [
            row_data["code"],
            row_data["description"],
            row_data["amount"],
            row_data["equivalent_name"],
            row_data["equivalent_amount"],
            row_data["difference"],
        ]
    elif sheet_type == SHEET_FORMULA_CALCULATION:
        values = [row_data["description"] or row_data["code"], row_data["formula"], row_data["amount"]]
    else:
        values = [row_data["description"] or row_data["code"], row_data["formula"] or row_data["amount"]]
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)
    ws.cell(row=row_idx, column=len(values) + 1, value=review_row_id)


def _create_home_sheet(
    workbook: Any,
    sheet_name: str,
    source_file_id: str,
    mode_label: str,
    source_rows: list[dict[str, str]],
) -> Any:
    ws = workbook.create_sheet(title=sheet_name, index=0)
    ws["A1"] = "Presupuesto de obra - Revision profesional (Home)"
    ws["A2"] = f"ID sanitizado: {source_file_id} | Modo: {mode_label}"
    ws["A3"] = "Seleccione una vista por hoja semantica. No se mezclan hojas incompatibles en una sola tabla."
    headers = ["Hoja origen", "Tipo semantico", "Confianza", "Vista profesional", "Estado", "Advertencias"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=header)
    for row_idx, row in enumerate(source_rows, start=5):
        ws.cell(row=row_idx, column=1, value=row["source_sheet_name"])
        ws.cell(row=row_idx, column=2, value=row["sheet_type"])
        ws.cell(row=row_idx, column=3, value=row["confidence"])
        ws.cell(row=row_idx, column=4, value=row["review_sheet_name"])
        ws.cell(row=row_idx, column=5, value=row["status"])
        ws.cell(row=row_idx, column=6, value=row["warnings"])
    _style_review_sheet(ws, visible_columns=6)
    return ws


def append_professional_budget_review(
    workbook: Any,
    extractions: list[object],  # Kept for backward compatibility with existing callers.
    source_file_id: str,
    import_batch_id: str,
    budget_version_id: str,
    mode_label: str = "PREVIEW_ONLY",
    sheet_semantics: dict[str, SheetSemanticClassification] | None = None,
) -> dict[str, str]:
    seq = _next_sequence(workbook, "BUDGET_REVIEW")
    home_sheet_name = f"BUDGET_REVIEW_{seq:03d}"
    trace_sheet_name = f"BUDGET_REVIEW_TRACE_{seq:03d}"
    trace_ws = workbook.create_sheet(title=trace_sheet_name, index=1)
    trace_ws.append(TRACE_HEADERS)

    preserved_index = _source_to_preserved_sheet(workbook)
    preserved_trace = _build_preserved_trace_index(workbook, import_batch_id=import_batch_id)
    semantics = sheet_semantics or {}
    source_rows: list[dict[str, str]] = []
    generated_view_count = 0
    trace_rows = 0

    for source_sheet_name, preserved_sheet_name in preserved_index.items():
        if preserved_sheet_name not in workbook.sheetnames:
            continue
        preserved_ws = workbook[preserved_sheet_name]
        semantic = semantics.get(
            source_sheet_name,
            SheetSemanticClassification(
                sheet_name=source_sheet_name,
                sheet_type=SHEET_UNKNOWN,
                confidence="LOW",
                reasons=("semantic_profile_missing",),
                warnings=("MANUAL_REVIEW_REQUIRED",),
                header_row=1,
                detected_columns={},
            ),
        )
        review_sheet_name = f"{home_sheet_name}_{_sanitize_suffix(source_sheet_name)}"
        review_ws = workbook.create_sheet(title=review_sheet_name, index=generated_view_count + 2)
        headers = TYPE_HEADERS.get(semantic.sheet_type, TYPE_HEADERS[SHEET_UNKNOWN])
        review_ws["A1"] = f"Vista profesional: {semantic.sheet_type}"
        review_ws["A2"] = f"Hoja origen: {source_sheet_name} | Preservada: {preserved_sheet_name}"
        review_ws["A3"] = f"Confianza: {semantic.confidence} | Razones: {', '.join(semantic.reasons) or 'n/a'}"
        for col_idx, header in enumerate(headers, start=1):
            review_ws.cell(row=4, column=col_idx, value=header)
        review_ws.cell(row=4, column=len(headers) + 1, value="_review_row_id")

        row_counter = 5
        extracted_rows = _extract_view_rows(preserved_ws, semantic)
        for row_data in extracted_rows:
            review_row_id = f"brv_{uuid4().hex[:10]}"
            _write_type_row(review_ws, row_counter, semantic.sheet_type, row_data, review_row_id)
            map_key = (source_sheet_name, row_data["row_number"])
            map_row = preserved_trace.get(map_key)
            _append_trace_row(
                trace_ws=trace_ws,
                review_row_id=review_row_id,
                review_sheet_name=review_sheet_name,
                review_row_number=row_counter,
                sheet_type=semantic.sheet_type,
                source_file_id=source_file_id,
                import_batch_id=import_batch_id,
                budget_version_id=budget_version_id,
                source_sheet_name=source_sheet_name,
                source_row_number=row_data["row_number"],
                map_row=map_row,
                notes="|".join(semantic.warnings),
            )
            row_counter += 1
            trace_rows += 1

        if row_counter == 5:
            review_ws.cell(row=5, column=1, value="Sin filas interpretables. Requiere revision manual.")
            review_ws.cell(row=5, column=len(headers) + 1, value=f"brv_{uuid4().hex[:10]}")
            row_counter += 1
        _style_review_sheet(review_ws, visible_columns=len(headers))
        generated_view_count += 1

        source_rows.append(
            {
                "source_sheet_name": source_sheet_name,
                "sheet_type": semantic.sheet_type,
                "confidence": semantic.confidence,
                "review_sheet_name": review_sheet_name,
                "status": "MANUAL_REVIEW_REQUIRED" if semantic.sheet_type in {SHEET_UNKNOWN, SHEET_AUXILIARY, SHEET_METADATA} else "READY",
                "warnings": ", ".join(semantic.warnings),
            }
        )

    home_ws = _create_home_sheet(
        workbook=workbook,
        sheet_name=home_sheet_name,
        source_file_id=source_file_id,
        mode_label=mode_label,
        source_rows=source_rows,
    )
    home_ws.sheet_properties.tabColor = "2F75B5"

    return {
        "review_sheet_name": home_sheet_name,
        "trace_sheet_name": trace_sheet_name,
        "review_row_count": str(max(home_ws.max_row - 4, 0)),
        "trace_row_count": str(trace_rows),
        "adaptive_review_views": str(generated_view_count),
    }
