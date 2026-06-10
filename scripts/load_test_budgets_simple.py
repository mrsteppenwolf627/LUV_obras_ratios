#!/usr/bin/env python3
"""Load test budgets and validate the complete flow (simplified)."""

import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.db.models import get_session
from src.db.schema import ItemMaster, Budget, LineItem
from src.core.auditor import compute_file_hash
from src.db.queries import get_budget_by_hash


def load_budget_from_excel(session, excel_path):
    """Load a single budget from Excel file."""
    excel_path = Path(excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel no encontrado: {excel_path}")

    # Compute hash
    file_hash = compute_file_hash(excel_path)

    # Check if already imported
    existing = get_budget_by_hash(session, file_hash)
    if existing:
        print(f"  ⚠️  Ya importado (hash: {file_hash[:16]}...)")
        return 0

    # Load Excel
    wb = load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Skip header
    data_rows = rows[1:] if len(rows) > 1 else []

    # Create budget
    budget = Budget(
        filename=excel_path.name,
        file_hash=file_hash,
        surface_m2=None,
        building_type="RESIDENTIAL",
        source_format="excel",
        total_cost=None,
        raw_data_json="",
    )
    session.add(budget)
    session.flush()

    # Add line items
    items_count = 0
    for row in data_rows:
        if not row or not row[0]:
            continue

        try:
            chapter, code, desc, unitario, qty, total = row[:6]
            line_item = LineItem(
                budget_id=budget.id,
                chapter_code=f"{chapter}.{code}" if chapter and code else str(code),
                chapter_name=desc,
                description=desc,
                quantity=float(qty) if qty else 0,
                unit="ud",
                unit_cost=float(unitario) if unitario else 0,
                total_cost=float(total) if total else 0,
                validation_status="VALID",
            )
            session.add(line_item)
            items_count += 1
        except (ValueError, TypeError, IndexError):
            continue

    session.commit()
    print(f"  ✅ Importado ({items_count} items)")
    return items_count


def get_db_stats(session):
    """Get current database statistics."""
    solidez_items = session.query(ItemMaster).filter(ItemMaster.muestras_count >= 5).count()
    total_items = session.query(ItemMaster).count()
    budgets = session.query(Budget).count()
    line_items = session.query(LineItem).count()

    return {
        "total_items": total_items,
        "solidez_items": solidez_items,
        "budgets": budgets,
        "line_items": line_items,
    }


def main():
    session = get_session()
    budget_dir = Path("data/samples/PRESUPUESTOS")

    print("\n" + "=" * 60)
    print("PRUEBA DE FUEGO - CARGAR PRESUPUESTOS NUEVOS")
    print("=" * 60)

    # ANTES
    stats_before = get_db_stats(session)
    print("\nESTADO ANTES:")
    print(f"  • Items totales: {stats_before['total_items']}")
    print(f"  • Items SÓLIDO (N≥5): {stats_before['solidez_items']}")
    print(f"  • Budgets: {stats_before['budgets']}")
    print(f"  • Line items: {stats_before['line_items']}")

    # CARGAR PRESUPUESTO 1
    print("\n" + "-" * 60)
    print("CARGANDO: TEST_PRESUPUESTO_EXTRA_1.xlsx")
    budget_file_1 = budget_dir / "TEST_PRESUPUESTO_EXTRA_1.xlsx"
    try:
        items_1 = load_budget_from_excel(session, budget_file_1)
    except Exception as e:
        print(f"  ❌ Error: {e}")
        session.close()
        return False

    stats_after_1 = get_db_stats(session)
    print(f"  Items SÓLIDO ahora: {stats_after_1['solidez_items']}")

    # CARGAR PRESUPUESTO 2
    print("\n" + "-" * 60)
    print("CARGANDO: TEST_PRESUPUESTO_EXTRA_2.xlsx")
    budget_file_2 = budget_dir / "TEST_PRESUPUESTO_EXTRA_2.xlsx"
    try:
        items_2 = load_budget_from_excel(session, budget_file_2)
    except Exception as e:
        print(f"  ❌ Error: {e}")
        session.close()
        return False

    stats_after_2 = get_db_stats(session)
    print(f"  Items SÓLIDO ahora: {stats_after_2['solidez_items']}")

    # VALIDACIÓN
    print("\n" + "=" * 60)
    print("VALIDACIÓN DE RESULTADOS")
    print("=" * 60)

    print("\n📊 CAMBIOS EN ESTADÍSTICAS:")
    print(f"  Items totales:       {stats_before['total_items']} → {stats_after_2['total_items']} (+{stats_after_2['total_items'] - stats_before['total_items']})")
    print(f"  Items SÓLIDO:        {stats_before['solidez_items']} → {stats_after_2['solidez_items']} (+{stats_after_2['solidez_items'] - stats_before['solidez_items']})")
    print(f"  Budgets:             {stats_before['budgets']} → {stats_after_2['budgets']} (+{stats_after_2['budgets'] - stats_before['budgets']})")
    print(f"  Line items:          {stats_before['line_items']} → {stats_after_2['line_items']} (+{stats_after_2['line_items'] - stats_before['line_items']})")

    # GENERAR MASTER
    print("\n" + "=" * 60)
    print("GENERANDO MASTER ACTUALIZADO")
    print("=" * 60)

    from src.export.excel_master_generator import generate_master_excel

    master_path = generate_master_excel(session)
    master_file = Path(master_path)

    print(f"✅ {master_file.name}")
    print(f"   Tamaño: {master_file.stat().st_size / 1024:.1f} KB")

    # Validar Master
    from openpyxl import load_workbook as load_wb

    wb = load_wb(master_path)
    ws = wb["ITEM_MASTER"]
    rows = list(ws.iter_rows(values_only=True))
    master_count = len(rows) - 1

    print(f"   Items en Master: {master_count}")

    assert master_count == stats_after_2['total_items'], f"Master tiene {master_count} items, esperaba {stats_after_2['total_items']}"

    session.close()

    print("\n" + "=" * 60)
    print("✅ PRUEBA DE FUEGO COMPLETADA EXITOSAMENTE")
    print("=" * 60 + "\n")

    return True


if __name__ == "__main__":
    try:
        success = main()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ ERROR: {e}", file=sys.stderr)
        import traceback

        traceback.print_exc()
        sys.exit(1)
