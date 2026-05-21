#!/usr/bin/env python3
"""Phase 9.18+ helper: run local XLSX dry-run generalization with sanitized IDs.

Phase 9.20 hardening: this wrapper now supports unambiguous human-review delivery.
- The output directory may be any subfolder under ``outputs/live_excel_master`` (not
  only ``xlsx_generalization``), so forensic review runs can use a fresh, never-reused
  folder such as ``manual_review_phase_9_20``.
- Output filenames are configurable via ``--name-template`` so review artifacts can be
  named ``phase_9_20_review_001.xlsx`` instead of reusing legacy preview names.
- Each artifact is re-opened from disk after saving to compute its SHA-256 and to
  re-read sheets/phase/active-sheet, guaranteeing that the file validated is byte-for-byte
  the same file delivered to the user (no in-memory-only validation).
- An optional integrity manifest binds filename <-> SHA-256 <-> phase <-> active sheet.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook  # type: ignore

try:
    from scripts.generate_live_excel_master import (
        PREVIEW_PIPELINE_PHASE,
        generate_preview_from_real_xlsx,
        validate_generated_xlsx_preview,
    )
    from scripts.live_excel_dry_run_evaluator import evaluate_dry_run_workbook
except ModuleNotFoundError:
    from generate_live_excel_master import (  # type: ignore
        PREVIEW_PIPELINE_PHASE,
        generate_preview_from_real_xlsx,
        validate_generated_xlsx_preview,
    )
    from live_excel_dry_run_evaluator import evaluate_dry_run_workbook  # type: ignore


ALLOWED_INPUT_ROOT = Path("data/samples")
# Backwards-compatible default output directory (legacy preview runs).
ALLOWED_OUTPUT_ROOT = Path("outputs/live_excel_master/xlsx_generalization")
# Phase 9.20: security boundary widened so forensic review folders are allowed,
# while still preventing writes outside the controlled outputs tree.
ALLOWED_OUTPUT_BOUNDARY = Path("outputs/live_excel_master")
DEFAULT_NAME_TEMPLATE = "xlsx_generalization_{index:03d}_preview"
SANITIZED_IDS = [
    "REAL_XLSX_GENERALIZATION_001",
    "REAL_XLSX_GENERALIZATION_002",
    "REAL_XLSX_GENERALIZATION_003",
    "REAL_XLSX_GENERALIZATION_004",
    "REAL_XLSX_GENERALIZATION_005",
]
_HOME_PATTERN = re.compile(r"BUDGET_REVIEW_\d{3}$")
_ADAPTIVE_PATTERN = re.compile(r"BUDGET_REVIEW_\d{3}_.+")


def _resolve_under(root: Path, value: Path) -> Path:
    return value.resolve() if value.is_absolute() else (root / value).resolve()


def _ensure_under(path: Path, allowed_root: Path, label: str) -> None:
    try:
        path.resolve().relative_to(allowed_root.resolve())
    except ValueError as exc:
        raise ValueError(f"{label} must stay under {allowed_root.as_posix()}") from exc


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8192), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_readme_phase(wb: Any) -> str:
    if "README_MASTER" not in wb.sheetnames:
        return ""
    ws = wb["README_MASTER"]
    for row_idx in range(2, ws.max_row + 1):
        key = str(ws.cell(row=row_idx, column=1).value or "").strip().lower()
        if key == "phase":
            return str(ws.cell(row=row_idx, column=2).value or "").strip()
    return ""


def _inspect_from_disk(path: Path) -> dict[str, Any]:
    """Re-open the saved workbook from disk and read its delivery-critical facts.

    Everything reported here is read from the on-disk artifact, never from an
    in-memory workbook, so the manifest cannot drift from the delivered file.
    """
    wb = load_workbook(path, data_only=False)
    try:
        sheets = list(wb.sheetnames)
        active_sheet = wb.active.title if wb.active is not None else ""
        readme_phase = _read_readme_phase(wb)
    finally:
        wb.close()
    home_sheets = [name for name in sheets if _HOME_PATTERN.fullmatch(name)]
    adaptive_views = [
        name
        for name in sheets
        if _ADAPTIVE_PATTERN.fullmatch(name) and not name.startswith("BUDGET_REVIEW_TRACE_")
    ]
    trace_sheets = [name for name in sheets if name.startswith("BUDGET_REVIEW_TRACE_")]
    required_sheets_present = bool(home_sheets) and "INDEX" in sheets and bool(trace_sheets)
    stat = path.stat()
    return {
        "sha256": _sha256(path),
        "size_bytes": stat.st_size,
        "modified_at_local": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
        "readme_phase": readme_phase,
        "active_sheet": active_sheet,
        "sheets": sheets,
        "home_sheets": home_sheets,
        "adaptive_views": adaptive_views,
        "trace_sheets": trace_sheets,
        "required_sheets_present": required_sheets_present,
        "adaptive_views_present": bool(adaptive_views),
    }


def _build_manifest_entry(
    artifact_index: int,
    run_id: str,
    input_id: str,
    preview_path: Path,
    repo_root: Path,
    required_phase: str,
) -> dict[str, Any]:
    """Validate the saved artifact from disk and bind its identity into a manifest entry.

    The validation in :func:`validate_generated_xlsx_preview` re-opens the file from
    disk; we additionally assert the SHA-256 is unchanged across validation, proving
    the validated bytes are exactly the delivered bytes.
    """
    sha_before = _sha256(preview_path)
    validation = validate_generated_xlsx_preview(preview_path, required_phase=required_phase)
    sha_after = _sha256(preview_path)
    if sha_before != sha_after:
        raise RuntimeError(
            f"Artifact mutated during validation (in-memory/disk mismatch): {preview_path.name}"
        )
    facts = _inspect_from_disk(preview_path)
    if facts["sha256"] != sha_after:
        raise RuntimeError(f"SHA-256 mismatch between read passes for {preview_path.name}")
    if facts["readme_phase"] != required_phase:
        raise RuntimeError(
            f"README phase {facts['readme_phase']!r} != required {required_phase!r} for {preview_path.name}"
        )
    if facts["active_sheet"] != "INDEX":
        raise RuntimeError(f"Active sheet {facts['active_sheet']!r} != INDEX for {preview_path.name}")
    try:
        relative_path = preview_path.resolve().relative_to(repo_root.resolve()).as_posix()
    except ValueError:
        relative_path = preview_path.as_posix()
    return {
        "artifact_id": f"PHASE_9_20_REVIEW_{artifact_index:03d}",
        "input_id": input_id,
        "run_id": run_id,
        "output_relative_path": relative_path,
        "output_absolute_path": str(preview_path.resolve()),
        "sha256": facts["sha256"],
        "size_bytes": facts["size_bytes"],
        "modified_at_local": facts["modified_at_local"],
        "readme_phase": facts["readme_phase"],
        "active_sheet": facts["active_sheet"],
        "sheets": facts["sheets"],
        "home_sheets": facts["home_sheets"],
        "adaptive_views": facts["adaptive_views"],
        "trace_sheets": facts["trace_sheets"],
        "required_sheets_present": facts["required_sheets_present"],
        "adaptive_views_present": facts["adaptive_views_present"],
        "validation_status": "PASSED" if validation.get("post_generation_validation") == "passed" else "FAILED",
        "human_review_start_sheet": facts["active_sheet"],
    }


def run_xlsx_generalization_dry_run(
    files: list[Path],
    output_dir: Path,
    repo_root: Path,
    name_template: str = DEFAULT_NAME_TEMPLATE,
    manifest_path: Path | None = None,
) -> dict[str, Any]:
    allowed_input = (repo_root / ALLOWED_INPUT_ROOT).resolve()
    allowed_boundary = (repo_root / ALLOWED_OUTPUT_BOUNDARY).resolve()
    output_dir_abs = _resolve_under(repo_root, output_dir)
    _ensure_under(output_dir_abs, allowed_boundary, "output_dir")
    output_dir_abs.mkdir(parents=True, exist_ok=True)

    if not files:
        raise ValueError("At least one input XLSX file is required.")
    if len(files) > len(SANITIZED_IDS):
        raise ValueError(f"Maximum supported files for this phase: {len(SANITIZED_IDS)}")

    results: list[dict[str, Any]] = []
    manifest_entries: list[dict[str, Any]] = []
    for idx, file_arg in enumerate(files):
        run_id = SANITIZED_IDS[idx]
        input_abs = _resolve_under(repo_root, file_arg)
        _ensure_under(input_abs, allowed_input, "input file")
        if not input_abs.exists():
            raise RuntimeError(f"Input file does not exist: {file_arg}")
        if input_abs.suffix.lower() != ".xlsx":
            raise ValueError(f"Only .xlsx is allowed for phase {PREVIEW_PIPELINE_PHASE} generalization: {file_arg}")

        preview_name = name_template.format(index=idx + 1) + ".xlsx"
        preview_path = output_dir_abs / preview_name
        if preview_path.exists():
            preview_path.unlink()

        generate_preview_from_real_xlsx(
            input_xlsx_path=input_abs,
            output_path=preview_path,
            source_file_id=run_id,
        )
        # Post-save validation + identity binding: re-opens the file from disk.
        manifest_entry = _build_manifest_entry(
            artifact_index=idx + 1,
            run_id=run_id,
            input_id=f"SANITIZED_INPUT_{idx + 1:03d}",
            preview_path=preview_path,
            repo_root=repo_root,
            required_phase=PREVIEW_PIPELINE_PHASE,
        )
        manifest_entries.append(manifest_entry)
        evaluation = evaluate_dry_run_workbook(preview_path, run_id=run_id)
        results.append(
            {
                "run_id": run_id,
                "state": evaluation.state,
                "format": "XLSX",
                "reasons": evaluation.reasons,
                "metrics": evaluation.metrics,
                "preview_output": str(preview_path.as_posix()),
                "sha256": manifest_entry["sha256"],
                "readme_phase": manifest_entry["readme_phase"],
                "active_sheet": manifest_entry["active_sheet"],
                "validation_status": manifest_entry["validation_status"],
            }
        )

    report: dict[str, Any] = {
        "phase": PREVIEW_PIPELINE_PHASE,
        "auto_promotion_enabled": False,
        "output_dir": output_dir_abs.as_posix(),
        "results": results,
        "manifest": manifest_entries,
    }

    if manifest_path is not None:
        manifest_abs = _resolve_under(repo_root, manifest_path)
        _ensure_under(manifest_abs, allowed_boundary, "manifest_json")
        manifest_abs.parent.mkdir(parents=True, exist_ok=True)
        manifest_doc = {
            "phase": PREVIEW_PIPELINE_PHASE,
            "generated_at_local": datetime.now().isoformat(timespec="seconds"),
            "auto_promotion_enabled": False,
            "artifacts": manifest_entries,
        }
        manifest_abs.write_text(json.dumps(manifest_doc, indent=2, ensure_ascii=True), encoding="utf-8")
        report["manifest_path"] = manifest_abs.as_posix()

    return report


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=f"Run local XLSX dry-run generalization (phase {PREVIEW_PIPELINE_PHASE})."
    )
    parser.add_argument("--files", nargs="+", type=Path, required=True, help="Input XLSX files under data/samples.")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=ALLOWED_OUTPUT_ROOT,
        help="Output directory under outputs/live_excel_master (any subfolder).",
    )
    parser.add_argument(
        "--report-json",
        type=Path,
        default=ALLOWED_OUTPUT_ROOT / "xlsx_generalization_report_sanitized.json",
        help="Sanitized report path under outputs/live_excel_master.",
    )
    parser.add_argument(
        "--name-template",
        type=str,
        default=DEFAULT_NAME_TEMPLATE,
        help="Output filename template (without extension); supports {index}. "
        "Example: phase_9_20_review_{index:03d}",
    )
    parser.add_argument(
        "--manifest-json",
        type=Path,
        default=None,
        help="Optional integrity manifest path under outputs/live_excel_master.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    allowed_boundary = (repo_root / ALLOWED_OUTPUT_BOUNDARY).resolve()
    report_path = _resolve_under(repo_root, args.report_json)
    _ensure_under(report_path, allowed_boundary, "report_json")
    report_path.parent.mkdir(parents=True, exist_ok=True)

    report = run_xlsx_generalization_dry_run(
        files=args.files,
        output_dir=args.output_dir,
        repo_root=repo_root,
        name_template=args.name_template,
        manifest_path=args.manifest_json,
    )
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=True), encoding="utf-8")
    print(
        json.dumps(
            {
                "report_json": str(report_path.as_posix()),
                "results": len(report["results"]),
                "manifest_path": report.get("manifest_path", ""),
            }
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
