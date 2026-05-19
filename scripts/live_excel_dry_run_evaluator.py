"""Dry-run evaluator for combined promotion + preservation contracts (phase 9.9)."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore

try:
    from scripts.live_excel_integrity import (
        ALLOWED_VALIDATION_STATUSES,
        ReferentialValidationError,
        SchemaValidationError,
        validate_referential_integrity,
        validate_workbook_schema,
    )
except ModuleNotFoundError:
    from live_excel_integrity import (  # type: ignore
        ALLOWED_VALIDATION_STATUSES,
        ReferentialValidationError,
        SchemaValidationError,
        validate_referential_integrity,
        validate_workbook_schema,
    )


STATE_OPERATIVE_CANDIDATE = "OPERATIVE_CANDIDATE"
STATE_PROMOTION_BLOCKED = "PROMOTION_BLOCKED"
STATE_MANUAL_REVIEW_REQUIRED = "MANUAL_REVIEW_REQUIRED"
STATE_PRESERVATION_INCOMPLETE = "PRESERVATION_INCOMPLETE"

PRELIM_TRACEABILITY_MIN = 0.95
PRELIM_AMOUNT_SEPARATION_MIN = 0.85
PRELIM_MANUAL_REVIEW_MAX = 0.25
PRELIM_BLOCKED_MAX = 0.0

_KNOWN_REASONS = {
    "missing_preserved_sheet",
    "missing_preserved_budgets_index",
    "missing_preserved_budget_sheets",
    "missing_preserved_to_cost_item_mapping",
    "insufficient_traceability",
    "ambiguous_mapping",
    "blocked_validation_status",
    "error_validation_status",
    "unknown_validation_status",
    "broken_relationship",
    "duplicate_ids",
    "amount_mixed_in_description",
    "insufficient_amount_separation",
    "manual_review_ratio_exceeded",
    "ratio_inputs_not_allowed",
    "ratios_calculated_not_allowed",
}


@dataclass
class DryRunEvaluation:
    run_id: str
    state: str
    reasons: list[str]
    metrics: dict[str, float]
    thresholds: dict[str, float]
    auto_promotion_enabled: bool = False

    def to_dict(self) -> dict[str, Any]:
        return {
            "run_id": self.run_id,
            "state": self.state,
            "reasons": list(self.reasons),
            "metrics": dict(self.metrics),
            "thresholds": dict(self.thresholds),
            "auto_promotion_enabled": self.auto_promotion_enabled,
        }


def _idx(headers: list[str], column: str) -> int:
    return headers.index(column) + 1


def _sheet_rows(ws: object, headers: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        row: dict[str, str] = {}
        any_value = False
        for col_idx, header in enumerate(headers, start=1):
            value = ws.cell(row=row_idx, column=col_idx).value
            text = "" if value is None else str(value).strip()
            row[header] = text
            if text:
                any_value = True
        if any_value:
            rows.append(row)
    return rows


def evaluate_dry_run_workbook(path: Path, run_id: str = "dry_run_eval") -> DryRunEvaluation:
    workbook = load_workbook(path)
    reasons: list[str] = []
    manual_reasons: list[str] = []
    blocked_reasons: list[str] = []
    preserve_reasons: list[str] = []

    metrics: dict[str, float] = {
        "total_preview_rows": 0.0,
        "total_preserved_rows": 0.0,
        "mapped_rows": 0.0,
        "unmapped_rows": 0.0,
        "mapping_rate": 0.0,
        "traceability_complete_rows": 0.0,
        "traceability_rate": 0.0,
        "manual_review_rows": 0.0,
        "manual_review_rate": 0.0,
        "blocked_rows": 0.0,
        "blocked_rate": 0.0,
        "amount_separated_rows": 0.0,
        "amount_separation_rate": 0.0,
        "ratio_input_rows": 0.0,
        "ratio_calculated_rows": 0.0,
    }
    thresholds = {
        "traceability_rate_min": PRELIM_TRACEABILITY_MIN,
        "amount_separation_rate_min": PRELIM_AMOUNT_SEPARATION_MIN,
        "manual_review_rate_max": PRELIM_MANUAL_REVIEW_MAX,
        "blocked_rate_max": PRELIM_BLOCKED_MAX,
    }

    try:
        validate_workbook_schema(workbook)
    except SchemaValidationError:
        blocked_reasons.append("broken_relationship")

    try:
        validate_referential_integrity(workbook)
    except ReferentialValidationError as exc:
        message = str(exc).lower()
        if "duplicate" in message:
            blocked_reasons.append("duplicate_ids")
        else:
            blocked_reasons.append("broken_relationship")

    if "PRESERVED_BUDGETS_INDEX" not in workbook.sheetnames:
        preserve_reasons.append("missing_preserved_budgets_index")
    if "PRESERVED_BUDGET_SHEETS" not in workbook.sheetnames:
        preserve_reasons.append("missing_preserved_budget_sheets")
    if "PRESERVED_TO_COST_ITEMS_MAP" not in workbook.sheetnames:
        preserve_reasons.append("missing_preserved_to_cost_item_mapping")
    if not any(name.startswith("PRES_") for name in workbook.sheetnames):
        preserve_reasons.append("missing_preserved_sheet")

    ratio_inputs_rows = workbook["RATIO_INPUTS"].max_row - 1 if "RATIO_INPUTS" in workbook.sheetnames else 0
    ratio_calculated_rows = (
        workbook["RATIOS_CALCULATED"].max_row - 1 if "RATIOS_CALCULATED" in workbook.sheetnames else 0
    )
    metrics["ratio_input_rows"] = float(max(ratio_inputs_rows, 0))
    metrics["ratio_calculated_rows"] = float(max(ratio_calculated_rows, 0))
    if ratio_inputs_rows > 0:
        blocked_reasons.append("ratio_inputs_not_allowed")
    if ratio_calculated_rows > 0:
        blocked_reasons.append("ratios_calculated_not_allowed")

    if "IMPORTED_BUDGET_VIEW" in workbook.sheetnames:
        preview_headers = [
            str(workbook["IMPORTED_BUDGET_VIEW"].cell(row=1, column=i).value or "").strip()
            for i in range(1, workbook["IMPORTED_BUDGET_VIEW"].max_column + 1)
        ]
        preview_rows = _sheet_rows(workbook["IMPORTED_BUDGET_VIEW"], preview_headers)
        metrics["total_preview_rows"] = float(len(preview_rows))
        trace_ok = 0
        amount_applicable = 0
        amount_ok = 0
        manual_count = 0
        blocked_count = 0
        for row in preview_rows:
            is_trace = all(
                row.get(key, "")
                for key in [
                    "source_file_id",
                    "import_batch_id",
                    "budget_version_id",
                    "source_sheet_name",
                    "source_row_number",
                ]
            )
            if is_trace:
                trace_ok += 1
            status = row.get("validation_status", "").upper()
            if status == "MANUAL_REVIEW_REQUIRED":
                manual_count += 1
            if status == "BLOCKED":
                blocked_count += 1
                blocked_reasons.append("blocked_validation_status")
            if status == "ERROR":
                blocked_count += 1
                blocked_reasons.append("error_validation_status")
            if status and status not in ALLOWED_VALIDATION_STATUSES:
                blocked_reasons.append("unknown_validation_status")
            if not status:
                manual_count += 1
            desc = row.get("item_description", "")
            amount = row.get("amount", "")
            if amount:
                amount_applicable += 1
                if any(char.isdigit() for char in desc):
                    blocked_reasons.append("amount_mixed_in_description")
                else:
                    amount_ok += 1
        total_preview = max(len(preview_rows), 1)
        metrics["traceability_complete_rows"] = float(trace_ok)
        metrics["traceability_rate"] = trace_ok / total_preview
        metrics["manual_review_rows"] = float(manual_count)
        metrics["manual_review_rate"] = manual_count / total_preview
        metrics["blocked_rows"] = float(blocked_count)
        metrics["blocked_rate"] = blocked_count / total_preview
        if amount_applicable > 0:
            metrics["amount_separated_rows"] = float(amount_ok)
            metrics["amount_separation_rate"] = amount_ok / amount_applicable
        else:
            metrics["amount_separated_rows"] = 0.0
            metrics["amount_separation_rate"] = 1.0

        if metrics["traceability_rate"] < PRELIM_TRACEABILITY_MIN:
            preserve_reasons.append("insufficient_traceability")
        if amount_applicable > 0 and metrics["amount_separation_rate"] < PRELIM_AMOUNT_SEPARATION_MIN:
            blocked_reasons.append("insufficient_amount_separation")
        if metrics["manual_review_rate"] > PRELIM_MANUAL_REVIEW_MAX:
            manual_reasons.append("manual_review_ratio_exceeded")
    else:
        preserve_reasons.append("missing_preserved_sheet")

    if "PRESERVED_TO_COST_ITEMS_MAP" in workbook.sheetnames:
        map_headers = [
            str(workbook["PRESERVED_TO_COST_ITEMS_MAP"].cell(row=1, column=i).value or "").strip()
            for i in range(1, workbook["PRESERVED_TO_COST_ITEMS_MAP"].max_column + 1)
        ]
        map_rows = _sheet_rows(workbook["PRESERVED_TO_COST_ITEMS_MAP"], map_headers)
        metrics["total_preserved_rows"] = float(len(map_rows))
        mapped = 0
        unmapped = 0
        for row in map_rows:
            status = row.get("mapping_status", "").upper()
            if status == "MAPPED":
                mapped += 1
            elif status == "UNMAPPED":
                unmapped += 1
            else:
                manual_reasons.append("ambiguous_mapping")
        metrics["mapped_rows"] = float(mapped)
        metrics["unmapped_rows"] = float(unmapped)
        metrics["mapping_rate"] = (mapped / len(map_rows)) if map_rows else 0.0
        if not map_rows:
            preserve_reasons.append("missing_preserved_to_cost_item_mapping")
    else:
        preserve_reasons.append("missing_preserved_to_cost_item_mapping")

    workbook.close()

    for reason in preserve_reasons + blocked_reasons + manual_reasons:
        if reason in _KNOWN_REASONS and reason not in reasons:
            reasons.append(reason)

    if preserve_reasons:
        state = STATE_PRESERVATION_INCOMPLETE
    elif blocked_reasons:
        state = STATE_PROMOTION_BLOCKED
    elif manual_reasons:
        state = STATE_MANUAL_REVIEW_REQUIRED
    else:
        state = STATE_OPERATIVE_CANDIDATE

    return DryRunEvaluation(
        run_id=run_id,
        state=state,
        reasons=reasons,
        metrics=metrics,
        thresholds=thresholds,
        auto_promotion_enabled=False,
    )

