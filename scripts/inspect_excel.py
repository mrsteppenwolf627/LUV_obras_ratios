#!/usr/bin/env python3
"""Diagnose Excel sample files without importing data into master."""

from __future__ import annotations

from datetime import datetime, timezone
import json
from pathlib import Path
import re
import sys
from typing import Any
import unicodedata

from openpyxl import load_workbook  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

SAMPLES_DIR = Path("data/samples")
REPORT_DIR = Path("reports/excel_diagnostics")
JSON_REPORT = REPORT_DIR / "excel_diagnostics_inventory.json"
MD_REPORT = REPORT_DIR / "excel_diagnostics_inventory_report.md"

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
LEGACY_OR_UNSUPPORTED_EXTENSIONS = {".xls", ".xlsb"}
ALL_EXCEL_EXTENSIONS = SUPPORTED_EXTENSIONS | LEGACY_OR_UNSUPPORTED_EXTENSIONS

COLUMN_PATTERNS = {
    "codigo": [r"\bcod(igo)?\b", r"\bref(erencia)?\b", r"\bpartida\b"],
    "descripcion": [r"\bdesc(ripcion)?\b", r"\bconcepto\b", r"\btexto\b"],
    "unidad": [r"\bunidad\b", r"\bud\b", r"\bu\.?\b"],
    "cantidad": [r"\bcantidad\b", r"\bmedicion\b", r"\bqty\b"],
    "precio": [r"\bprecio\b", r"\bp\.?\s*unit\b", r"\bunitario\b"],
    "importe": [r"\bimporte\b", r"\btotal\b", r"\bcoste\b", r"\bpresupuesto\b"],
}

HEADER_KEYWORDS = [
    "codigo",
    "cod",
    "partida",
    "descripcion",
    "unidad",
    "ud",
    "cantidad",
    "medicion",
    "precio",
    "importe",
    "total",
    "presupuesto",
]
UNIT_TOKEN_RE = re.compile(r"\b(m2|m3|ml|ud|kg|l|cm|mm)\b", re.IGNORECASE)


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_text(text: str) -> str:
    value = unicodedata.normalize("NFKD", text or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return value.lower().strip()


def _detect_sheet_type(sheet: Any) -> str:
    cls = sheet.__class__.__name__.lower()
    if "chartsheet" in cls:
        return "CHARTSHEET"
    if hasattr(sheet, "iter_rows"):
        return "WORKSHEET"
    return "UNKNOWN"


def _truncate_cells(cells: list[str], max_cells: int = 8, max_len: int = 60) -> list[str]:
    out: list[str] = []
    for value in cells[:max_cells]:
        if len(value) > max_len:
            out.append(value[: max_len - 3] + "...")
        else:
            out.append(value)
    return out


def _scan_header_rows(rows_text: list[list[str]]) -> tuple[list[int], list[list[str]], dict[str, list[str]]]:
    scored_rows: list[tuple[int, int, int, list[str]]] = []
    max_rows = min(50, len(rows_text))

    for ridx in range(max_rows):
        row = rows_text[ridx]
        non_empty = [cell for cell in row if cell]
        if len(non_empty) < 2:
            continue

        text_cells = sum(1 for cell in non_empty if any(ch.isalpha() for ch in cell))
        kw_hits = 0
        for cell in non_empty:
            cell_n = _normalize_text(cell)
            if any(kw in cell_n for kw in HEADER_KEYWORDS):
                kw_hits += 1

        score = kw_hits * 3 + text_cells
        if kw_hits > 0 or text_cells >= 3:
            scored_rows.append((ridx + 1, score, kw_hits, non_empty))

    scored_rows.sort(key=lambda x: (x[1], x[2]), reverse=True)
    candidate_row_numbers = [item[0] for item in scored_rows[:3]]
    candidate_rows_cells = [_truncate_cells(item[3], max_cells=10) for item in scored_rows[:3]]

    candidate_columns: dict[str, list[str]] = {k: [] for k in COLUMN_PATTERNS}
    if scored_rows:
        best_row_idx = scored_rows[0][0] - 1
        best_row = rows_text[best_row_idx]
        for cidx, cell in enumerate(best_row, start=1):
            cell_n = _normalize_text(cell)
            if not cell_n:
                continue
            for field, patterns in COLUMN_PATTERNS.items():
                if any(re.search(p, cell_n) for p in patterns):
                    label = _as_text(cell) or get_column_letter(cidx)
                    if label not in candidate_columns[field]:
                        candidate_columns[field].append(label)

    return candidate_row_numbers, candidate_rows_cells, candidate_columns


def _derive_columns_without_clear_headers(
    rows_text: list[list[str]],
    formula_positions: set[tuple[int, int]],
    number_formats_map: dict[tuple[int, int], str],
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {k: [] for k in COLUMN_PATTERNS}
    if min_row == 0 or min_col == 0:
        return out

    for cidx in range(min_col, max_col + 1):
        text_count = 0
        numeric_count = 0
        formula_count = 0
        unit_hits = 0
        large_numeric_count = 0

        for ridx in range(min_row, max_row + 1):
            cell = rows_text[ridx - 1][cidx - 1] if ridx - 1 < len(rows_text) and cidx - 1 < len(rows_text[ridx - 1]) else ""
            if not cell:
                continue
            if (ridx, cidx) in formula_positions:
                formula_count += 1

            cell_n = _normalize_text(cell)
            if UNIT_TOKEN_RE.search(cell_n):
                unit_hits += 1

            num_candidate = cell_n.replace(".", "").replace(",", "")
            if num_candidate.isdigit():
                numeric_count += 1
                try:
                    parsed = float(cell_n.replace(".", "").replace(",", ".")) if "," in cell_n else float(cell_n)
                except Exception:
                    parsed = 0.0
                if abs(parsed) >= 100:
                    large_numeric_count += 1
            else:
                text_count += 1

        label = get_column_letter(cidx)
        active_cells = text_count + numeric_count
        if active_cells == 0:
            continue

        if unit_hits > 0:
            out["unidad"].append(label)

        if formula_count > 0 and (numeric_count >= 2 or large_numeric_count >= 1):
            if label not in out["importe"]:
                out["importe"].append(label)

        if numeric_count >= max(2, text_count) and large_numeric_count >= 1:
            if label not in out["importe"]:
                out["importe"].append(label)

        fmt_hits = 0
        for ridx in range(min_row, max_row + 1):
            fmt = _normalize_text(number_formats_map.get((ridx, cidx), ""))
            if any(token in fmt for token in ["#", "0.00", "$", "€", "_-"]):
                fmt_hits += 1
        if fmt_hits >= 2 and label not in out["precio"]:
            out["precio"].append(label)

        if numeric_count >= 2 and formula_count == 0 and label not in out["cantidad"]:
            out["cantidad"].append(label)

        if text_count >= 3 and numeric_count <= 1:
            if label not in out["descripcion"]:
                out["descripcion"].append(label)

        if text_count >= 2 and numeric_count <= 2:
            if label not in out["codigo"]:
                out["codigo"].append(label)

    return out


def _merge_candidate_columns(primary: dict[str, list[str]], secondary: dict[str, list[str]]) -> dict[str, list[str]]:
    merged: dict[str, list[str]] = {k: [] for k in COLUMN_PATTERNS}
    for key in merged:
        for value in primary.get(key, []) + secondary.get(key, []):
            if value and value not in merged[key]:
                merged[key].append(value)
    return merged


def inspect_excel_file(path: Path, relative_path: str) -> dict[str, Any]:
    ext = path.suffix.lower()
    file_ref = {
        "relative_path": relative_path,
        "extension": ext,
        "size_bytes": path.stat().st_size,
        "supported_extension": ext in SUPPORTED_EXTENSIONS,
    }

    if ext in LEGACY_OR_UNSUPPORTED_EXTENSIONS:
        return {
            "file_ref": file_ref,
            "workbook_status": "UNSUPPORTED_LEGACY_EXCEL",
            "sheets": [],
            "summary": {
                "worksheets_count": 0,
                "chartsheets_count": 0,
                "empty_sheets_count": 0,
                "non_tabular_sheets_count": 0,
                "possible_tables_count": 0,
            },
            "risks": ["LEGACY_OR_UNSUPPORTED_EXTENSION"],
            "manual_review": ["CONVERT_TO_SUPPORTED_EXCEL_FORMAT"],
        }

    try:
        wb = load_workbook(path, read_only=False, data_only=False)
    except Exception as exc:
        return {
            "file_ref": file_ref,
            "workbook_status": "WORKBOOK_OPEN_ERROR",
            "sheets": [],
            "summary": {
                "worksheets_count": 0,
                "chartsheets_count": 0,
                "empty_sheets_count": 0,
                "non_tabular_sheets_count": 0,
                "possible_tables_count": 0,
            },
            "risks": ["WORKBOOK_OPEN_ERROR"],
            "manual_review": [f"OPEN_ERROR:{exc}"],
        }

    sheets_out: list[dict[str, Any]] = []
    risks: list[str] = []
    manual_review: list[str] = []

    worksheets_count = 0
    chartsheets_count = 0
    empty_sheets_count = 0
    non_tabular_sheets_count = 0
    possible_tables_count = 0

    hidden_rows_total = 0
    hidden_columns_total = 0

    for sheet in wb.worksheets + wb.chartsheets:  # type: ignore[attr-defined]
        sheet_type = _detect_sheet_type(sheet)
        sheet_name = sheet.title
        out: dict[str, Any] = {
            "sheet_name": sheet_name,
            "sheet_type": sheet_type,
        }

        if sheet_type != "WORKSHEET":
            if sheet_type == "CHARTSHEET":
                chartsheets_count += 1
                manual_review.append(f"CHARTSHEET_PRESENT:{sheet_name}")
            else:
                manual_review.append(f"UNKNOWN_SHEET_TYPE:{sheet_name}")
            out.update(
                {
                    "dimensions": {"max_row": None, "max_column": None},
                    "used_range": {"min_row": None, "max_row": None, "min_column": None, "max_column": None},
                    "non_empty_rows": 0,
                    "non_empty_columns": 0,
                    "density_profile": {
                        "non_empty_cells": 0,
                        "used_cells_total": 0,
                        "non_empty_pct": 0.0,
                        "top_non_empty_rows": [],
                        "top_non_empty_columns": [],
                    },
                    "possible_tables": [],
                    "candidate_headers": [],
                    "candidate_header_rows": [],
                    "candidate_columns": {k: [] for k in COLUMN_PATTERNS},
                    "merged_cells_count": 0,
                    "formula_cells_count": 0,
                    "numeric_format_samples": [],
                    "hidden_rows_count": 0,
                    "hidden_columns_count": 0,
                    "column_width_samples": {},
                    "format_density": {"styled_cells": 0, "style_vs_data_ratio": 0.0},
                    "sanitized_samples": {
                        "first_non_empty_rows": [],
                        "dense_rows": [],
                        "possible_header_rows": [],
                    },
                    "is_empty_sheet": True,
                    "is_likely_tabular": False,
                }
            )
            sheets_out.append(out)
            continue

        worksheets_count += 1
        max_row = getattr(sheet, "max_row", 0) or 0
        max_column = getattr(sheet, "max_column", 0) or 0

        rows_text: list[list[str]] = []
        non_empty_rows = 0
        non_empty_cols_set: set[int] = set()
        formula_cells_count = 0
        numeric_formats: set[str] = set()
        formula_positions: set[tuple[int, int]] = set()
        number_formats_map: dict[tuple[int, int], str] = {}

        non_empty_cells = 0
        styled_cells = 0
        min_used_row = 0
        max_used_row = 0
        min_used_col = 0
        max_used_col = 0
        row_counts: dict[int, int] = {}
        col_counts: dict[int, int] = {}

        for ridx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column), start=1):
            row_values: list[str] = []
            has_data = False
            for cidx, cell in enumerate(row, start=1):
                txt = _as_text(cell.value)
                row_values.append(txt)
                number_formats_map[(ridx, cidx)] = _as_text(cell.number_format)

                if cell.style_id not in (None, 0):
                    styled_cells += 1

                if txt:
                    has_data = True
                    non_empty_cells += 1
                    non_empty_cols_set.add(cidx)
                    row_counts[ridx] = row_counts.get(ridx, 0) + 1
                    col_counts[cidx] = col_counts.get(cidx, 0) + 1
                    if min_used_row == 0 or ridx < min_used_row:
                        min_used_row = ridx
                    if ridx > max_used_row:
                        max_used_row = ridx
                    if min_used_col == 0 or cidx < min_used_col:
                        min_used_col = cidx
                    if cidx > max_used_col:
                        max_used_col = cidx

                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells_count += 1
                    formula_positions.add((ridx, cidx))
                fmt = _as_text(cell.number_format)
                if fmt and fmt not in {"General", "@"} and len(numeric_formats) < 30:
                    numeric_formats.add(fmt)
            if has_data:
                non_empty_rows += 1
            rows_text.append(row_values)

        used_cells_total = 0
        if min_used_row and min_used_col:
            used_cells_total = (max_used_row - min_used_row + 1) * (max_used_col - min_used_col + 1)

        density = (non_empty_cells / used_cells_total) if used_cells_total else 0.0

        candidate_header_rows, candidate_header_samples, headers_based_columns = _scan_header_rows(rows_text)
        inferred_columns = _derive_columns_without_clear_headers(
            rows_text,
            formula_positions,
            number_formats_map,
            min_used_row,
            max_used_row,
            min_used_col,
            max_used_col,
        )
        candidate_columns = _merge_candidate_columns(headers_based_columns, inferred_columns)

        possible_tables = sorted(list(getattr(sheet, "tables", {}).keys()))
        merged_cells_count = len(getattr(sheet, "merged_cells", {}).ranges)

        hidden_rows_count = sum(1 for dim in sheet.row_dimensions.values() if getattr(dim, "hidden", False))
        hidden_columns_count = sum(1 for dim in sheet.column_dimensions.values() if getattr(dim, "hidden", False))
        hidden_rows_total += hidden_rows_count
        hidden_columns_total += hidden_columns_count

        width_samples: dict[str, float] = {}
        for cidx in range(min_used_col or 1, min((max_used_col or 1), (min_used_col or 1) + 7) + 1):
            letter = get_column_letter(cidx)
            width = getattr(sheet.column_dimensions.get(letter), "width", None)
            if isinstance(width, (int, float)):
                width_samples[letter] = float(width)

        top_rows = sorted(row_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        top_cols = sorted(col_counts.items(), key=lambda x: x[1], reverse=True)[:5]

        first_non_empty_rows = []
        for ridx in sorted(row_counts.keys())[:3]:
            row = rows_text[ridx - 1]
            row_vals = [cell for cell in row if cell]
            first_non_empty_rows.append({"row": ridx, "cells": _truncate_cells(row_vals, max_cells=6)})

        dense_rows = []
        for ridx, cnt in top_rows[:3]:
            row = rows_text[ridx - 1]
            row_vals = [cell for cell in row if cell]
            dense_rows.append({"row": ridx, "non_empty_cells": cnt, "cells": _truncate_cells(row_vals, max_cells=6)})

        if non_empty_rows == 0:
            empty_sheets_count += 1
            risks.append("EMPTY_WORKSHEET_PRESENT")
            manual_review.append(f"EMPTY_WORKSHEET:{sheet_name}")

        is_likely_tabular = non_empty_rows > 1 and len(non_empty_cols_set) > 1 and density >= 0.03
        if not is_likely_tabular:
            non_tabular_sheets_count += 1
            manual_review.append(f"NON_TABULAR_WORKSHEET:{sheet_name}")

        if not candidate_header_rows:
            manual_review.append(f"NO_CLEAR_HEADERS:{sheet_name}")

        if formula_cells_count > 0:
            risks.append("FORMULAS_PRESENT")

        if merged_cells_count > 0:
            risks.append("MERGED_CELLS_PRESENT")

        if hidden_rows_count > 0 or hidden_columns_count > 0:
            risks.append("HIDDEN_STRUCTURE_PRESENT")

        style_ratio = (styled_cells / non_empty_cells) if non_empty_cells else 0.0
        if styled_cells > 0 and non_empty_cells > 0 and style_ratio > 3 and density < 0.1:
            risks.append("HIGH_FORMAT_LOW_DATA_DENSITY")

        possible_tables_count += len(possible_tables)

        out.update(
            {
                "dimensions": {"max_row": max_row, "max_column": max_column},
                "used_range": {
                    "min_row": min_used_row or None,
                    "max_row": max_used_row or None,
                    "min_column": min_used_col or None,
                    "max_column": max_used_col or None,
                },
                "non_empty_rows": non_empty_rows,
                "non_empty_columns": len(non_empty_cols_set),
                "density_profile": {
                    "non_empty_cells": non_empty_cells,
                    "used_cells_total": used_cells_total,
                    "non_empty_pct": round(density * 100, 2),
                    "top_non_empty_rows": [{"row": r, "count": c} for r, c in top_rows],
                    "top_non_empty_columns": [{"column": get_column_letter(c), "count": n} for c, n in top_cols],
                },
                "possible_tables": possible_tables,
                "candidate_headers": candidate_header_samples,
                "candidate_header_rows": candidate_header_rows,
                "candidate_columns": candidate_columns,
                "merged_cells_count": merged_cells_count,
                "formula_cells_count": formula_cells_count,
                "numeric_format_samples": sorted(numeric_formats),
                "hidden_rows_count": hidden_rows_count,
                "hidden_columns_count": hidden_columns_count,
                "column_width_samples": width_samples,
                "format_density": {
                    "styled_cells": styled_cells,
                    "style_vs_data_ratio": round(style_ratio, 2),
                },
                "sanitized_samples": {
                    "first_non_empty_rows": first_non_empty_rows,
                    "dense_rows": dense_rows,
                    "possible_header_rows": [
                        {"row": row_no, "cells": cells}
                        for row_no, cells in zip(candidate_header_rows, candidate_header_samples)
                    ],
                },
                "is_empty_sheet": non_empty_rows == 0,
                "is_likely_tabular": is_likely_tabular,
            }
        )
        sheets_out.append(out)

    wb.close()

    if chartsheets_count > 0:
        risks.append("CHARTSHEETS_PRESENT")

    return {
        "file_ref": file_ref,
        "workbook_status": "EXCEL_DIAGNOSED",
        "sheets": sheets_out,
        "summary": {
            "worksheets_count": worksheets_count,
            "chartsheets_count": chartsheets_count,
            "empty_sheets_count": empty_sheets_count,
            "non_tabular_sheets_count": non_tabular_sheets_count,
            "possible_tables_count": possible_tables_count,
            "hidden_rows_total": hidden_rows_total,
            "hidden_columns_total": hidden_columns_total,
        },
        "risks": sorted(set(risks)),
        "manual_review": sorted(set(manual_review)),
    }


def inspect_excel_samples(root: Path) -> dict[str, Any]:
    samples_path = root / SAMPLES_DIR
    generated_at = datetime.now(timezone.utc).isoformat()

    if not samples_path.exists():
        return {
            "generated_at": generated_at,
            "samples_dir": str(SAMPLES_DIR).replace("\\", "/"),
            "exists": False,
            "files_total": 0,
            "excel_files_detected": 0,
            "files": [],
            "global_summary": {},
        }

    all_files = sorted([p for p in samples_path.rglob("*") if p.is_file()])
    excel_files = [p for p in all_files if p.suffix.lower() in ALL_EXCEL_EXTENSIONS]

    out_files: list[dict[str, Any]] = []
    worksheets_total = 0
    chartsheets_total = 0

    for path in excel_files:
        rel = str(path.relative_to(root)).replace("\\", "/")
        diagnosed = inspect_excel_file(path, rel)
        out_files.append(diagnosed)
        summary = diagnosed.get("summary", {})
        worksheets_total += int(summary.get("worksheets_count", 0) or 0)
        chartsheets_total += int(summary.get("chartsheets_count", 0) or 0)

    return {
        "generated_at": generated_at,
        "samples_dir": str(SAMPLES_DIR).replace("\\", "/"),
        "exists": True,
        "files_total": len(all_files),
        "excel_files_detected": len(excel_files),
        "files": out_files,
        "global_summary": {
            "worksheets_total": worksheets_total,
            "chartsheets_total": chartsheets_total,
            "excel_files_with_manual_review": sum(1 for f in out_files if f.get("manual_review")),
            "excel_files_with_risks": sum(1 for f in out_files if f.get("risks")),
        },
    }


def write_reports(root: Path, payload: dict[str, Any]) -> tuple[Path, Path]:
    json_path = root / JSON_REPORT
    md_path = root / MD_REPORT
    json_path.parent.mkdir(parents=True, exist_ok=True)

    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    lines = [
        "# Excel Diagnostics Inventory",
        "",
        "> Local report. Real-data output may be sensitive and must remain outside Git.",
        "",
        f"- Generated at (UTC): {payload.get('generated_at')}",
        f"- Samples dir: {payload.get('samples_dir')}",
        f"- Files total: {payload.get('files_total', 0)}",
        f"- Excel files detected: {payload.get('excel_files_detected', 0)}",
        "",
        "## Global summary",
        "",
        f"- Worksheets total: {payload.get('global_summary', {}).get('worksheets_total', 0)}",
        f"- Chartsheets total: {payload.get('global_summary', {}).get('chartsheets_total', 0)}",
        f"- Excel files with risks: {payload.get('global_summary', {}).get('excel_files_with_risks', 0)}",
        f"- Excel files with manual review: {payload.get('global_summary', {}).get('excel_files_with_manual_review', 0)}",
        "",
        "## Files",
        "",
    ]

    for entry in payload.get("files", []):
        ref = entry.get("file_ref", {})
        lines.append(f"### {ref.get('relative_path')}")
        lines.append(f"- Workbook status: {entry.get('workbook_status')}")
        lines.append(f"- Worksheets: {entry.get('summary', {}).get('worksheets_count', 0)}")
        lines.append(f"- Chartsheets: {entry.get('summary', {}).get('chartsheets_count', 0)}")
        lines.append(f"- Risks: {', '.join(entry.get('risks', [])) or 'none'}")
        lines.append(f"- Manual review: {', '.join(entry.get('manual_review', [])) or 'none'}")
        for sheet in entry.get("sheets", []):
            if sheet.get("sheet_type") != "WORKSHEET":
                continue
            lines.append(f"  - Sheet {sheet.get('sheet_name')}: headers_rows={sheet.get('candidate_header_rows', [])}, columns={sheet.get('candidate_columns', {})}")
        lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")
    return json_path, md_path


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    payload = inspect_excel_samples(root)
    json_path, md_path = write_reports(root, payload)

    print("Excel diagnostics summary")
    print(f"- Samples dir: {payload.get('samples_dir')}")
    print(f"- Files total: {payload.get('files_total', 0)}")
    print(f"- Excel files detected: {payload.get('excel_files_detected', 0)}")
    print(f"- Worksheets total: {payload.get('global_summary', {}).get('worksheets_total', 0)}")
    print(f"- Chartsheets total: {payload.get('global_summary', {}).get('chartsheets_total', 0)}")
    print(f"- JSON report: {json_path.relative_to(root).as_posix()}")
    print(f"- Markdown report: {md_path.relative_to(root).as_posix()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
